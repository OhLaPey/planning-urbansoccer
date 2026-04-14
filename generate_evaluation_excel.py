#!/usr/bin/env python3
"""Génère la fiche d'évaluation arbitre UrbanSoccer League au format Excel."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORANGE = "FF6600"
DARK_BG = "1A1A1A"
DARK_ROW = "2A2A2A"
LIGHT_ROW = "333333"
WHITE = "FFFFFF"
GRAY = "888888"
GREEN = "4CAF50"
AMBER = "FF9800"
BLUE = "42A5F5"

thin_border = Border(
    left=Side(style="thin", color="444444"),
    right=Side(style="thin", color="444444"),
    top=Side(style="thin", color="444444"),
    bottom=Side(style="thin", color="444444"),
)

orange_border = Border(
    left=Side(style="thin", color=ORANGE),
    right=Side(style="thin", color=ORANGE),
    top=Side(style="thin", color=ORANGE),
    bottom=Side(style="thin", color=ORANGE),
)


def style_cell(ws, row, col, value="", font_size=10, bold=False, color=WHITE,
               fill=None, alignment=None, border=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=font_size, bold=bold, color=color)
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.alignment = alignment or Alignment(vertical="center", wrap_text=wrap)
    if border:
        cell.border = border
    return cell


def section_title(ws, row, title, col_start=1, col_end=8):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        cell.border = Border(
            bottom=Side(style="medium", color=ORANGE),
            top=Side(style="medium", color=ORANGE),
        )
    style_cell(ws, row, col_start, title, font_size=11, bold=True, color=WHITE,
               fill=ORANGE, alignment=Alignment(vertical="center"))
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)


def info_field(ws, row, col_label, label, col_value, merge_end=None):
    style_cell(ws, row, col_label, label, font_size=9, bold=True, color=GRAY,
               fill=DARK_BG, border=thin_border)
    cell = style_cell(ws, row, col_value, "", font_size=10, bold=False, color=WHITE,
                       fill=DARK_ROW, border=orange_border)
    cell.alignment = Alignment(vertical="center")
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col_value, end_row=row, end_column=merge_end)


CRITERES = [
    ("PLACEMENT & DÉPLACEMENTS", [
        ("Placement & Déplacements",
         "Positionnement sur le terrain, déplacements fluides, bons angles de vision"),
    ]),
    ("SIFFLET & SIGNALÉTIQUE", [
        ("Clarté du sifflet", "Coups de sifflet nets, audibles et bien différenciés"),
        ("Gestuelle", "Signale clairement les décisions (fautes, touches, corners, coups francs)"),
        ("Timing du sifflet", "Siffle au bon moment, ni trop tôt ni trop tard"),
    ]),
    ("DÉCISIONS & LECTURE DU JEU", [
        ("Justesse des décisions", "Les décisions prises sont correctes et cohérentes"),
        ("Cohérence", "Applique les mêmes critères du début à la fin du match"),
        ("Gestion de l'avantage", "Laisse jouer quand c'est pertinent, sait laisser le jeu couler"),
        ("Connaissance des règles", "Maîtrise les règles spécifiques UrbanSoccer / futsal"),
    ]),
    ("COMMUNICATION & GESTION", [
        ("Communication avec les joueurs", "Échange calmement, explique ses décisions"),
        ("Autorité & Présence", "Inspire le respect, contrôle le match sans être autoritaire"),
        ("Gestion des tensions", "Désamorce les situations conflictuelles avec sang-froid"),
        ("Utilisation des sanctions", "Avertissements et cartons de façon juste et proportionnée"),
    ]),
    ("ATTITUDE & PROFESSIONNALISME", [
        ("Présentation", "Tenue correcte, ponctualité, équipement d'arbitre"),
        ("Impartialité", "Traite les deux équipes de façon équitable"),
        ("Concentration", "Reste attentif du début à la fin, même sur les temps morts"),
        ("Timing & Gestion du temps", "Début des matchs à l'heure, durées de mi-temps, enchaînement"),
    ]),
    ("ANIMATION", [
        ("Jeux à la mi-temps", "Animation proposée quand programmée (quiz, défis, jeux…)"),
        ("Relation clients", "Discussions avec les joueurs/équipes, accueil, disponibilité"),
        ("Ambiance de la soirée", "Contribue à une bonne atmosphère, dynamise la soirée"),
    ]),
]

NOTE_LABELS = ["N/N", "1", "2", "3", "4", "5"]


def generate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Évaluation Arbitre"
    ws.sheet_properties.tabColor = ORANGE

    # Fond sombre global
    for row in range(1, 80):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")

    # Largeurs de colonnes
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 6
    ws.column_dimensions["H"].width = 6

    # ── HEADER ──
    r = 1
    ws.merge_cells("A1:H1")
    style_cell(ws, 1, 1, "ÉVALUATION ARBITRE", font_size=18, bold=True, color=WHITE,
               fill=DARK_BG, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 36

    r = 2
    ws.merge_cells("A2:H2")
    style_cell(ws, 2, 1, "UrbanSoccer League", font_size=10, bold=True, color=ORANGE,
               fill=DARK_BG, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 20

    # Ligne séparatrice orange
    r = 3
    for c in range(1, 9):
        ws.cell(row=3, column=c).fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    ws.row_dimensions[3].height = 4

    # ── INFORMATIONS ──
    r = 5
    section_title(ws, r, "INFORMATIONS")
    ws.row_dimensions[r].height = 24

    r = 6
    info_field(ws, r, 1, "Arbitre", 2, merge_end=4)
    info_field(ws, r, 5, "Date", 6, merge_end=8)
    ws.row_dimensions[r].height = 28

    r = 7
    info_field(ws, r, 1, "Match", 2, merge_end=8)
    ws.row_dimensions[r].height = 28

    r = 8
    info_field(ws, r, 1, "Créneau", 2, merge_end=4)
    info_field(ws, r, 5, "Observateur", 6, merge_end=8)
    ws.row_dimensions[r].height = 28

    # ── ÉVALUATION ──
    r = 10
    section_title(ws, r, "ÉVALUATION")
    ws.row_dimensions[r].height = 24

    # En-têtes colonnes notes
    r = 11
    style_cell(ws, r, 1, "", fill=DARK_BG)
    style_cell(ws, r, 2, "Critère", font_size=9, bold=True, color=GRAY, fill=DARK_BG,
               alignment=Alignment(vertical="center"))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    for i, label in enumerate(NOTE_LABELS):
        col = 3 + i
        style_cell(ws, r, col, label, font_size=9, bold=True, color=GRAY, fill=DARK_BG,
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=thin_border)
    ws.row_dimensions[r].height = 20

    r = 12
    alt = False
    for cat_name, items in CRITERES:
        # Sous-titre catégorie
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_cell(ws, r, 1, cat_name, font_size=9, bold=True, color=ORANGE,
                   fill="222222", alignment=Alignment(vertical="center"))
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = Border(
                bottom=Side(style="thin", color=ORANGE),
                top=Side(style="thin", color=ORANGE),
            )
            if c > 1:
                ws.cell(row=r, column=c).fill = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
        ws.row_dimensions[r].height = 22
        r += 1

        for nom, desc in items:
            bg = LIGHT_ROW if alt else DARK_ROW
            alt = not alt

            # Nom du critère
            style_cell(ws, r, 1, "", fill=bg)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            style_cell(ws, r, 1, nom, font_size=10, bold=True, color=WHITE, fill=bg,
                       border=thin_border, alignment=Alignment(vertical="center", indent=1))

            # Cases de notation N/N, 1-5
            for i in range(6):
                col = 3 + i
                style_cell(ws, r, col, "", font_size=10, color=WHITE, fill=bg,
                           border=orange_border,
                           alignment=Alignment(horizontal="center", vertical="center"))
            ws.row_dimensions[r].height = 26
            r += 1

            # Description + commentaire
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            style_cell(ws, r, 1, desc, font_size=8, color=GRAY, fill=bg,
                       alignment=Alignment(vertical="center", indent=1, wrap_text=True))
            ws.row_dimensions[r].height = 16
            r += 1

            # Ligne commentaire
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            style_cell(ws, r, 1, "", font_size=9, color="CCCCCC", fill=bg,
                       border=Border(bottom=Side(style="dotted", color="444444")),
                       alignment=Alignment(vertical="center", indent=1))
            ws.row_dimensions[r].height = 22
            r += 1

    # ── SITUATIONS À REVOIR ──
    r += 1
    section_title(ws, r, "SITUATIONS À REVOIR (VIDÉO)")
    ws.row_dimensions[r].height = 24
    r += 1

    # En-têtes
    style_cell(ws, r, 1, "Min.", font_size=9, bold=True, color=GRAY, fill=DARK_BG,
               border=thin_border, alignment=Alignment(horizontal="center", vertical="center"))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    style_cell(ws, r, 2, "Description de la situation", font_size=9, bold=True, color=GRAY,
               fill=DARK_BG, border=thin_border, alignment=Alignment(vertical="center"))
    ws.row_dimensions[r].height = 20
    r += 1

    for i in range(6):
        bg = LIGHT_ROW if i % 2 else DARK_ROW
        style_cell(ws, r, 1, "", font_size=10, color=WHITE, fill=bg, border=orange_border,
                   alignment=Alignment(horizontal="center", vertical="center"))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        style_cell(ws, r, 2, "", font_size=10, color=WHITE, fill=bg, border=orange_border,
                   alignment=Alignment(vertical="center"))
        ws.row_dimensions[r].height = 24
        r += 1

    # ── BILAN ──
    r += 1
    section_title(ws, r, "BILAN")
    ws.row_dimensions[r].height = 24
    r += 1

    for label, color in [("POINTS POSITIFS", GREEN), ("AXES D'AMÉLIORATION", AMBER), ("COMMENTAIRE GÉNÉRAL", GRAY)]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_cell(ws, r, 1, label, font_size=9, bold=True, color=color, fill=DARK_BG,
                   alignment=Alignment(vertical="center"))
        ws.row_dimensions[r].height = 18
        r += 1

        for line in range(4):
            bg = LIGHT_ROW if line % 2 else DARK_ROW
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            style_cell(ws, r, 1, "", font_size=10, color=WHITE, fill=bg,
                       border=Border(bottom=Side(style="dotted", color="444444")),
                       alignment=Alignment(vertical="center", indent=1))
            ws.row_dimensions[r].height = 22
            r += 1

        r += 1

    # ── Mise en page impression ──
    ws.print_area = f"A1:H{r}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    output = "Evaluation Arbitre - UrbanSoccer League.xlsx"
    wb.save(output)
    print(f"Fichier généré : {output}")


if __name__ == "__main__":
    generate()
