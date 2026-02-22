#!/usr/bin/env python3
"""
Application web UrbanSoccer — Gestion des anniversaires.

Interface web (Flask + SQLite) qui remplace le fichier Excel d'entrée.
Permet de saisir les anniversaires et de générer :
  - Un récapitulatif Excel + PDF pour le suivi week-end
  - Les affiches prénoms (PPTX + PDF) à imprimer pour les tables

Usage :
    python app.py                # lance le serveur sur http://localhost:5000
    python app.py --port 8080    # port personnalisé
"""

import argparse
import io
import os
import sqlite3
from datetime import datetime, date, timedelta

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, send_file, jsonify, g)

# Import du générateur d'affiches existant
from generate_birthday_posters import (
    FORMULE_MAP, TEMPLATE_PATH, OUTPUT_DIR, TEMPLATE_ORANGE,
    TEMPLATE_FFF_BIENVENUE, TEMPLATE_FFF_CERTIFICAT, TEMPLATE_BUMP,
    split_prenoms, generate_pptx, convert_to_pdf,
)

# ── Config ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "Anniversaires", "anniversaires.db")

app = Flask(__name__,
            template_folder=os.path.join(SCRIPT_DIR, "templates"),
            static_folder=os.path.join(SCRIPT_DIR, "static"))
app.secret_key = "urbansoccer-anniversaires-2026"

FORMULES = list(FORMULE_MAP.keys())


# ── Base de données ───────────────────────────────────────────────────────────

def get_db():
    """Connexion SQLite avec row_factory."""
    if "db" not in g:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    """Crée la table si elle n'existe pas."""
    db = get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS birthdays (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date_anniv  TEXT NOT NULL,
            horaire     TEXT NOT NULL DEFAULT '',
            animateur   TEXT NOT NULL DEFAULT '',
            formule     TEXT NOT NULL,
            nb_enfants  TEXT NOT NULL DEFAULT '',
            prenom      TEXT NOT NULL,
            boisson     TEXT NOT NULL DEFAULT '',
            cadeau      TEXT NOT NULL DEFAULT '',
            options     TEXT NOT NULL DEFAULT '',
            gateau      TEXT NOT NULL DEFAULT '',
            commentaires TEXT NOT NULL DEFAULT '',
            created_at  TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            updated_at  TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
    """)
    db.execute("""
        CREATE INDEX IF NOT EXISTS idx_birthdays_date
        ON birthdays(date_anniv)
    """)
    db.commit()


@app.before_request
def before_request():
    init_db()


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_dates_with_birthdays():
    """Retourne la liste des dates qui ont au moins un anniversaire."""
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT date_anniv FROM birthdays ORDER BY date_anniv DESC"
    ).fetchall()
    return [row["date_anniv"] for row in rows]


def get_birthdays_for_date(date_str):
    """Retourne tous les anniversaires pour une date donnée, triés par horaire."""
    db = get_db()
    rows = db.execute(
        "SELECT * FROM birthdays WHERE date_anniv = ? ORDER BY horaire, prenom",
        (date_str,)
    ).fetchall()
    return [dict(row) for row in rows]


def format_date_display(date_str):
    """Convertit '2026-02-22' en 'Dimanche 22 février 2026'."""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi",
                 "Samedi", "Dimanche"]
        mois = ["janvier", "février", "mars", "avril", "mai", "juin",
                "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
        return f"{jours[d.weekday()]} {d.day} {mois[d.month - 1]} {d.year}"
    except (ValueError, IndexError):
        return date_str


def next_weekend_dates():
    """Retourne les dates du prochain samedi et dimanche."""
    today = date.today()
    # Samedi = 5, Dimanche = 6
    days_until_saturday = (5 - today.weekday()) % 7
    if days_until_saturday == 0 and today.weekday() == 5:
        saturday = today
    elif today.weekday() == 6:
        saturday = today - timedelta(days=1)
    else:
        saturday = today + timedelta(days=days_until_saturday or 7)
    sunday = saturday + timedelta(days=1)
    return saturday, sunday


def weekend_for_date(date_str):
    """Retourne (samedi, dimanche) pour le week-end contenant la date donnée."""
    d = datetime.strptime(date_str, "%Y-%m-%d").date()
    if d.weekday() == 6:  # Dimanche
        saturday = d - timedelta(days=1)
    elif d.weekday() == 5:  # Samedi
        saturday = d
    else:
        days_until = (5 - d.weekday()) % 7
        saturday = d + timedelta(days=days_until or 7)
    sunday = saturday + timedelta(days=1)
    return saturday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")


def get_birthdays_for_dates(date_list):
    """Retourne tous les anniversaires pour plusieurs dates."""
    db = get_db()
    placeholders = ",".join("?" for _ in date_list)
    rows = db.execute(
        f"SELECT * FROM birthdays WHERE date_anniv IN ({placeholders}) "
        f"ORDER BY date_anniv, horaire, prenom",
        date_list
    ).fetchall()
    return [dict(row) for row in rows]


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    """Page d'accueil : sélection de date + liste des anniversaires."""
    # Date sélectionnée (param ou prochain samedi)
    selected_date = request.args.get("date")
    if not selected_date:
        sat, _ = next_weekend_dates()
        selected_date = sat.strftime("%Y-%m-%d")

    birthdays = get_birthdays_for_date(selected_date)
    all_dates = get_dates_with_birthdays()

    # Ajouter la date sélectionnée si elle n'est pas dans la liste
    if selected_date not in all_dates:
        all_dates.append(selected_date)
        all_dates.sort(reverse=True)

    # Stats
    formule_counts = {}
    total_enfants = 0
    for b in birthdays:
        f = b["formule"]
        formule_counts[f] = formule_counts.get(f, 0) + 1
        try:
            total_enfants += int(b["nb_enfants"])
        except (ValueError, TypeError):
            pass

    return render_template("index.html",
                           birthdays=birthdays,
                           selected_date=selected_date,
                           all_dates=all_dates,
                           format_date=format_date_display,
                           formule_counts=formule_counts,
                           total_enfants=total_enfants,
                           formules=FORMULES)


@app.route("/add", methods=["GET", "POST"])
def add_birthday():
    """Ajouter un anniversaire."""
    if request.method == "POST":
        date_anniv = request.form["date_anniv"]
        prenom = request.form["prenom"].strip()
        formule = request.form["formule"]

        if not prenom:
            flash("Le prénom est obligatoire.", "error")
            return redirect(url_for("add_birthday", date=date_anniv))

        if formule not in FORMULE_MAP:
            flash(f"Formule inconnue : {formule}", "error")
            return redirect(url_for("add_birthday", date=date_anniv))

        db = get_db()
        db.execute("""
            INSERT INTO birthdays
                (date_anniv, horaire, animateur, formule, nb_enfants,
                 prenom, boisson, cadeau, options, gateau, commentaires)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            date_anniv,
            request.form.get("horaire", "").strip(),
            request.form.get("animateur", "").strip(),
            formule,
            request.form.get("nb_enfants", "").strip(),
            prenom,
            request.form.get("boisson", "").strip(),
            request.form.get("cadeau", "").strip(),
            request.form.get("options", "").strip(),
            request.form.get("gateau", "").strip(),
            request.form.get("commentaires", "").strip(),
        ))
        db.commit()

        flash(f"Anniversaire de {prenom} ajouté !", "success")
        return redirect(url_for("index", date=date_anniv))

    # GET
    date_anniv = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    return render_template("form.html",
                           birthday=None,
                           date_anniv=date_anniv,
                           formules=FORMULES,
                           action="add")


@app.route("/edit/<int:birthday_id>", methods=["GET", "POST"])
def edit_birthday(birthday_id):
    """Modifier un anniversaire."""
    db = get_db()

    if request.method == "POST":
        prenom = request.form["prenom"].strip()
        formule = request.form["formule"]
        date_anniv = request.form["date_anniv"]

        if not prenom:
            flash("Le prénom est obligatoire.", "error")
            return redirect(url_for("edit_birthday", birthday_id=birthday_id))

        db.execute("""
            UPDATE birthdays SET
                date_anniv=?, horaire=?, animateur=?, formule=?, nb_enfants=?,
                prenom=?, boisson=?, cadeau=?, options=?, gateau=?, commentaires=?,
                updated_at=datetime('now', 'localtime')
            WHERE id=?
        """, (
            date_anniv,
            request.form.get("horaire", "").strip(),
            request.form.get("animateur", "").strip(),
            formule,
            request.form.get("nb_enfants", "").strip(),
            prenom,
            request.form.get("boisson", "").strip(),
            request.form.get("cadeau", "").strip(),
            request.form.get("options", "").strip(),
            request.form.get("gateau", "").strip(),
            request.form.get("commentaires", "").strip(),
            birthday_id,
        ))
        db.commit()

        flash(f"Anniversaire de {prenom} mis à jour !", "success")
        return redirect(url_for("index", date=date_anniv))

    # GET
    row = db.execute("SELECT * FROM birthdays WHERE id=?", (birthday_id,)).fetchone()
    if not row:
        flash("Anniversaire introuvable.", "error")
        return redirect(url_for("index"))

    return render_template("form.html",
                           birthday=dict(row),
                           date_anniv=row["date_anniv"],
                           formules=FORMULES,
                           action="edit")


@app.route("/delete/<int:birthday_id>", methods=["POST"])
def delete_birthday(birthday_id):
    """Supprimer un anniversaire."""
    db = get_db()
    row = db.execute("SELECT prenom, date_anniv FROM birthdays WHERE id=?",
                     (birthday_id,)).fetchone()
    if row:
        db.execute("DELETE FROM birthdays WHERE id=?", (birthday_id,))
        db.commit()
        flash(f"Anniversaire de {row['prenom']} supprimé.", "success")
        return redirect(url_for("index", date=row["date_anniv"]))
    flash("Anniversaire introuvable.", "error")
    return redirect(url_for("index"))


@app.route("/duplicate/<int:birthday_id>", methods=["POST"])
def duplicate_birthday(birthday_id):
    """Dupliquer un anniversaire existant."""
    db = get_db()
    row = db.execute("SELECT * FROM birthdays WHERE id=?", (birthday_id,)).fetchone()
    if not row:
        flash("Anniversaire introuvable.", "error")
        return redirect(url_for("index"))

    db.execute("""
        INSERT INTO birthdays
            (date_anniv, horaire, animateur, formule, nb_enfants,
             prenom, boisson, cadeau, options, gateau, commentaires)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        row["date_anniv"], row["horaire"], row["animateur"], row["formule"],
        row["nb_enfants"], row["prenom"] + " (copie)", row["boisson"],
        row["cadeau"], row["options"], row["gateau"], row["commentaires"],
    ))
    db.commit()
    flash(f"Copie de {row['prenom']} créée !", "success")
    return redirect(url_for("index", date=row["date_anniv"]))


# ── Vue week-end ─────────────────────────────────────────────────────────────

@app.route("/weekend")
def weekend_view():
    """Vue combinée samedi + dimanche du week-end."""
    selected_date = request.args.get("date")
    if not selected_date:
        sat, _ = next_weekend_dates()
        selected_date = sat.strftime("%Y-%m-%d")

    sat_str, sun_str = weekend_for_date(selected_date)
    sat_birthdays = get_birthdays_for_date(sat_str)
    sun_birthdays = get_birthdays_for_date(sun_str)
    all_dates = get_dates_with_birthdays()

    def compute_stats(birthdays):
        fc = {}
        te = 0
        for b in birthdays:
            fc[b["formule"]] = fc.get(b["formule"], 0) + 1
            try:
                te += int(b["nb_enfants"])
            except (ValueError, TypeError):
                pass
        return fc, te

    sat_fc, sat_te = compute_stats(sat_birthdays)
    sun_fc, sun_te = compute_stats(sun_birthdays)

    return render_template("weekend.html",
                           sat_date=sat_str, sun_date=sun_str,
                           sat_birthdays=sat_birthdays,
                           sun_birthdays=sun_birthdays,
                           sat_formule_counts=sat_fc, sun_formule_counts=sun_fc,
                           sat_total_enfants=sat_te, sun_total_enfants=sun_te,
                           all_dates=all_dates,
                           format_date=format_date_display,
                           formules=FORMULES)


# ── Génération fichiers ──────────────────────────────────────────────────────

def birthdays_to_entries(birthdays):
    """Convertit les rows SQLite en entrées compatibles avec generate_pptx."""
    return [{
        "prenom": b["prenom"],
        "formule": b["formule"],
        "horaire": b["horaire"],
        "nb_enfants": b["nb_enfants"],
        "boisson": b["boisson"],
        "cadeau": b["cadeau"],
        "options": b["options"],
        "gateau": b["gateau"],
        "commentaires": b["commentaires"],
        "animateur": b["animateur"],
    } for b in birthdays]


@app.route("/generate/posters/<date_str>")
def generate_posters(date_str):
    """Génère les affiches PPTX + PDF et les télécharge."""
    birthdays = get_birthdays_for_date(date_str)
    if not birthdays:
        flash("Aucun anniversaire pour cette date.", "error")
        return redirect(url_for("index", date=date_str))

    entries = birthdays_to_entries(birthdays)
    safe_name = date_str.replace("-", ".")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    pptx_path = os.path.join(OUTPUT_DIR, f"Affiches_{safe_name}.pptx")

    generate_pptx(entries, TEMPLATE_PATH, pptx_path)
    pdf_path = convert_to_pdf(pptx_path)

    # Envoyer le PDF s'il existe, sinon le PPTX
    out_path = pdf_path or pptx_path
    return send_file(out_path, as_attachment=True,
                     download_name=os.path.basename(out_path))


@app.route("/generate/recap/<date_str>")
def generate_recap(date_str):
    """Génère le récapitulatif Excel pour le suivi week-end."""
    birthdays = get_birthdays_for_date(date_str)
    if not birthdays:
        flash("Aucun anniversaire pour cette date.", "error")
        return redirect(url_for("index", date=date_str))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = format_date_display(date_str)

    # ── Styles ──
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF7832", end_color="FF7832",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Formule couleurs
    formule_fills = {
        "Ligue 1":         PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"),
        "Champions League": PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid"),
        "FFF":             PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
        "Bump":            PatternFill(start_color="F5E6FF", end_color="F5E6FF", fill_type="solid"),
    }

    # ── Titre ──
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"ANNIVERSAIRES — {format_date_display(date_str)}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FF7832")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ── Stats ──
    formule_counts = {}
    total_enfants = 0
    for b in birthdays:
        formule_counts[b["formule"]] = formule_counts.get(b["formule"], 0) + 1
        try:
            total_enfants += int(b["nb_enfants"])
        except (ValueError, TypeError):
            pass

    stats_parts = [f"{len(birthdays)} anniv"]
    for f, c in sorted(formule_counts.items()):
        stats_parts.append(f"{c} {f}")
    stats_parts.append(f"{total_enfants} enfants")

    ws.merge_cells("A2:J2")
    ws["A2"].value = " | ".join(stats_parts)
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # Ligne vide
    ws.row_dimensions[3].height = 8

    # ── En-têtes ──
    headers = ["Horaire", "Animateur", "Formule", "Nb enfants",
               "Prénom", "Boisson", "Cadeau", "Options", "Gâteau", "Commentaires"]
    col_widths = [10, 14, 18, 12, 22, 12, 12, 14, 14, 25]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.row_dimensions[4].height = 30

    # ── Données ──
    fields = ["horaire", "animateur", "formule", "nb_enfants",
              "prenom", "boisson", "cadeau", "options", "gateau", "commentaires"]

    for row_idx, b in enumerate(birthdays, 5):
        formule_fill = formule_fills.get(b["formule"])
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=b[field])
            cell.alignment = cell_align
            cell.border = thin_border
            if formule_fill:
                cell.fill = formule_fill

        # Mettre le prénom en gras
        ws.cell(row=row_idx, column=5).font = Font(name="Calibri", bold=True, size=11)
        ws.row_dimensions[row_idx].height = 24

    # ── Freeze panes ──
    ws.freeze_panes = "A5"

    # Sauvegarder en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = date_str.replace("-", ".")
    filename = f"Recap_Anniversaires_{safe_name}.xlsx"

    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/generate/recap-pdf/<date_str>")
def generate_recap_pdf(date_str):
    """Génère le récapitulatif PDF pour le suivi week-end."""
    birthdays = get_birthdays_for_date(date_str)
    if not birthdays:
        flash("Aucun anniversaire pour cette date.", "error")
        return redirect(url_for("index", date=date_str))

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)

    styles = getSampleStyleSheet()
    elements = []

    # Titre
    title_style = styles["Title"]
    title_style.textColor = colors.HexColor("#FF7832")
    title_style.fontSize = 18
    elements.append(Paragraph(
        f"ANNIVERSAIRES — {format_date_display(date_str)}", title_style))
    elements.append(Spacer(1, 0.3 * cm))

    # Stats
    formule_counts = {}
    total_enfants = 0
    for b in birthdays:
        formule_counts[b["formule"]] = formule_counts.get(b["formule"], 0) + 1
        try:
            total_enfants += int(b["nb_enfants"])
        except (ValueError, TypeError):
            pass

    stats = f"{len(birthdays)} anniversaires | {total_enfants} enfants | "
    stats += " | ".join(f"{c} {f}" for f, c in sorted(formule_counts.items()))
    sub_style = styles["Normal"]
    sub_style.textColor = colors.HexColor("#666666")
    sub_style.alignment = 1  # center
    elements.append(Paragraph(stats, sub_style))
    elements.append(Spacer(1, 0.5 * cm))

    # Tableau
    headers = ["Horaire", "Animateur", "Formule", "Enfants",
               "Prénom", "Boisson", "Cadeau", "Options", "Gâteau", "Commentaires"]

    data = [headers]
    for b in birthdays:
        data.append([
            b["horaire"], b["animateur"], b["formule"], b["nb_enfants"],
            b["prenom"], b["boisson"], b["cadeau"], b["options"],
            b["gateau"], b["commentaires"],
        ])

    col_widths = [2.2*cm, 3*cm, 3.5*cm, 2*cm, 4*cm,
                  2.2*cm, 2.2*cm, 2.5*cm, 2.5*cm, 4*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Couleurs par formule
    formule_colors = {
        "Ligue 1":          colors.HexColor("#FFF2E6"),
        "Champions League":  colors.HexColor("#E6F0FF"),
        "FFF":              colors.HexColor("#E6FFE6"),
        "Bump":             colors.HexColor("#F5E6FF"),
    }

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF7832")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
        ("FONTNAME", (4, 1), (4, -1), "Helvetica-Bold"),
    ]

    # Appliquer couleur par formule
    for row_idx, b in enumerate(birthdays, 1):
        bg = formule_colors.get(b["formule"])
        if bg:
            style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), bg))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    safe_name = date_str.replace("-", ".")
    filename = f"Recap_Anniversaires_{safe_name}.pdf"

    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype="application/pdf")


# ── Génération week-end combiné ───────────────────────────────────────────────

@app.route("/generate/weekend-posters/<date_str>")
def generate_weekend_posters(date_str):
    """Génère les affiches PPTX+PDF pour tout le week-end."""
    sat_str, sun_str = weekend_for_date(date_str)
    birthdays = get_birthdays_for_dates([sat_str, sun_str])
    if not birthdays:
        flash("Aucun anniversaire ce week-end.", "error")
        return redirect(url_for("weekend_view", date=date_str))

    entries = birthdays_to_entries(birthdays)
    safe_name = sat_str.replace("-", ".")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    pptx_path = os.path.join(OUTPUT_DIR, f"Affiches_WE_{safe_name}.pptx")

    generate_pptx(entries, TEMPLATE_PATH, pptx_path)
    pdf_path = convert_to_pdf(pptx_path)

    out_path = pdf_path or pptx_path
    return send_file(out_path, as_attachment=True,
                     download_name=os.path.basename(out_path))


@app.route("/generate/weekend-recap/<date_str>")
def generate_weekend_recap(date_str):
    """Génère le récap Excel week-end (un onglet par jour)."""
    sat_str, sun_str = weekend_for_date(date_str)
    sat_b = get_birthdays_for_date(sat_str)
    sun_b = get_birthdays_for_date(sun_str)

    if not sat_b and not sun_b:
        flash("Aucun anniversaire ce week-end.", "error")
        return redirect(url_for("weekend_view", date=date_str))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF7832", end_color="FF7832",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    formule_fills = {
        "Ligue 1":         PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"),
        "Champions League": PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid"),
        "FFF":             PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
        "Bump":            PatternFill(start_color="F5E6FF", end_color="F5E6FF", fill_type="solid"),
    }

    headers = ["Horaire", "Animateur", "Formule", "Nb enfants",
               "Prénom", "Boisson", "Cadeau", "Options", "Gâteau", "Commentaires"]
    col_widths = [10, 14, 18, 12, 22, 12, 12, 14, 14, 25]
    fields = ["horaire", "animateur", "formule", "nb_enfants",
              "prenom", "boisson", "cadeau", "options", "gateau", "commentaires"]

    for day_str, day_b in [(sat_str, sat_b), (sun_str, sun_b)]:
        day_label = format_date_display(day_str)
        ws = wb.create_sheet(title=day_label[:31])

        # Titre
        ws.merge_cells("A1:J1")
        ws["A1"].value = f"ANNIVERSAIRES — {day_label}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FF7832")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Stats
        fc = {}
        te = 0
        for b in day_b:
            fc[b["formule"]] = fc.get(b["formule"], 0) + 1
            try:
                te += int(b["nb_enfants"])
            except (ValueError, TypeError):
                pass
        parts = [f"{len(day_b)} anniv"]
        for f, c in sorted(fc.items()):
            parts.append(f"{c} {f}")
        parts.append(f"{te} enfants")

        ws.merge_cells("A2:J2")
        ws["A2"].value = " | ".join(parts)
        ws["A2"].font = Font(name="Calibri", size=10, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 8

        # En-têtes
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[chr(64 + ci)].width = w
        ws.row_dimensions[4].height = 30

        # Données
        for ri, b in enumerate(day_b, 5):
            ff = formule_fills.get(b["formule"])
            for ci, field in enumerate(fields, 1):
                cell = ws.cell(row=ri, column=ci, value=b[field])
                cell.alignment = cell_align
                cell.border = thin_border
                if ff:
                    cell.fill = ff
            ws.cell(row=ri, column=5).font = Font(name="Calibri", bold=True, size=11)
            ws.row_dimensions[ri].height = 24

        ws.freeze_panes = "A5"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe = sat_str.replace("-", ".")
    return send_file(buffer, as_attachment=True,
                     download_name=f"Recap_WeekEnd_{safe}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/generate/weekend-recap-pdf/<date_str>")
def generate_weekend_recap_pdf(date_str):
    """Génère le récap PDF week-end (samedi + dimanche)."""
    sat_str, sun_str = weekend_for_date(date_str)
    sat_b = get_birthdays_for_date(sat_str)
    sun_b = get_birthdays_for_date(sun_str)

    if not sat_b and not sun_b:
        flash("Aucun anniversaire ce week-end.", "error")
        return redirect(url_for("weekend_view", date=date_str))

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, PageBreak)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("WETitle", parent=styles["Title"],
                                  textColor=colors.HexColor("#FF7832"),
                                  fontSize=18)
    sub_style = ParagraphStyle("WESub", parent=styles["Normal"],
                                textColor=colors.HexColor("#666666"),
                                alignment=1)

    formule_colors = {
        "Ligue 1":          colors.HexColor("#FFF2E6"),
        "Champions League":  colors.HexColor("#E6F0FF"),
        "FFF":              colors.HexColor("#E6FFE6"),
        "Bump":             colors.HexColor("#F5E6FF"),
    }
    col_widths = [2.2*cm, 3*cm, 3.5*cm, 2*cm, 4*cm,
                  2.2*cm, 2.2*cm, 2.5*cm, 2.5*cm, 4*cm]
    pdf_headers = ["Horaire", "Animateur", "Formule", "Enfants",
                   "Prénom", "Boisson", "Cadeau", "Options", "Gâteau", "Commentaires"]

    elements = []

    for i, (day_str, day_b) in enumerate([(sat_str, sat_b), (sun_str, sun_b)]):
        if i > 0:
            elements.append(PageBreak())

        day_label = format_date_display(day_str)
        elements.append(Paragraph(f"ANNIVERSAIRES — {day_label}", title_style))
        elements.append(Spacer(1, 0.3*cm))

        fc = {}
        te = 0
        for b in day_b:
            fc[b["formule"]] = fc.get(b["formule"], 0) + 1
            try:
                te += int(b["nb_enfants"])
            except (ValueError, TypeError):
                pass

        stats = f"{len(day_b)} anniversaires | {te} enfants"
        if fc:
            stats += " | " + " | ".join(f"{c} {f}" for f, c in sorted(fc.items()))
        elements.append(Paragraph(stats, sub_style))
        elements.append(Spacer(1, 0.5*cm))

        if day_b:
            data = [pdf_headers]
            for b in day_b:
                data.append([
                    b["horaire"], b["animateur"], b["formule"], b["nb_enfants"],
                    b["prenom"], b["boisson"], b["cadeau"], b["options"],
                    b["gateau"], b["commentaires"],
                ])

            table = Table(data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF7832")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("FONTNAME", (4, 1), (4, -1), "Helvetica-Bold"),
            ]
            for ri, b in enumerate(day_b, 1):
                bg = formule_colors.get(b["formule"])
                if bg:
                    style_cmds.append(("BACKGROUND", (0, ri), (-1, ri), bg))
            table.setStyle(TableStyle(style_cmds))
            elements.append(table)
        else:
            elements.append(Paragraph("Aucun anniversaire.", sub_style))

    doc.build(elements)
    buffer.seek(0)

    safe = sat_str.replace("-", ".")
    return send_file(buffer, as_attachment=True,
                     download_name=f"Recap_WeekEnd_{safe}.pdf",
                     mimetype="application/pdf")


# ── API JSON (pour AJAX) ─────────────────────────────────────────────────────

@app.route("/api/birthdays/<date_str>")
def api_birthdays(date_str):
    """Retourne les anniversaires en JSON."""
    birthdays = get_birthdays_for_date(date_str)
    return jsonify(birthdays)


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="UrbanSoccer Anniversaires Web App")
    parser.add_argument("--port", type=int, default=5000)
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    print(f"\n  UrbanSoccer Anniversaires")
    print(f"  http://localhost:{args.port}\n")

    app.run(host=args.host, port=args.port, debug=args.debug)
