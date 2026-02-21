#!/usr/bin/env python3
"""Convertit le fichier Excel S10 en fichiers ICS + HTML + JSON."""

import openpyxl
import json
import os
import re
from datetime import datetime, timedelta

# === Configuration ===
EXCEL_FILE = "Plannings 2026 S10.xlsx"
WEEK_NUM = 10
YEAR = 2026
# Dates de la semaine (du lundi au dimanche)
WEEK_DATES = {
    "B": datetime(2026, 3, 2),   # Lundi
    "C": datetime(2026, 3, 3),   # Mardi
    "D": datetime(2026, 3, 4),   # Mercredi
    "E": datetime(2026, 3, 5),   # Jeudi
    "F": datetime(2026, 3, 6),   # Vendredi
    "G": datetime(2026, 3, 7),   # Samedi
    "H": datetime(2026, 3, 8),   # Dimanche
}

# Mapping codes -> noms lisibles
CODE_NAMES = {
    "VDC": "Vie de centre",
    "L-REG": "Régisseur League",
    "CUP-L": "Cup League",
    "CUP-R": "Cup Régisseur",
    "STAGE": "Stage",
    "STA-E": "Stage encadrement",
    "C-PAD": "Cours Padel",
    "PAD-A": "Padel animation",
    "ANNIV": "Anniversaire",
    "INVEN": "Inventaire",
    "MAL": "Maladie",
    "REU": "Réunion",
    "EV-RE": "Événement régisseur",
    "AIDE": "Aide",
    "P25M": "P25M",
    "PSG": "PSG Academy",
}

COLS = ["B", "C", "D", "E", "F", "G", "H"]
COL_INDEX = {c: i for i, c in enumerate(COLS)}


def slug(name):
    """BONILLO Matthieu -> bonillo-matthieu"""
    s = name.lower()
    s = s.replace("ï", "i").replace("é", "e").replace("è", "e").replace("ê", "e")
    s = s.replace("ô", "o").replace("ü", "u").replace("ù", "u").replace("û", "u")
    s = s.replace("à", "a").replace("â", "a").replace("ç", "c")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def parse_time(time_str, base_date):
    """Parse '08:00/10:00' or '19:00/00:30+' into (start_dt, end_dt)."""
    parts = time_str.strip().split("/")
    if len(parts) != 2:
        return None

    start_str = parts[0].strip()
    end_str = parts[1].strip()

    next_day = end_str.endswith("+")
    if next_day:
        end_str = end_str.rstrip("+")

    sh, sm = int(start_str.split(":")[0]), int(start_str.split(":")[1])
    eh, em = int(end_str.split(":")[0]), int(end_str.split(":")[1])

    # Handle 24:00 as midnight next day
    start_extra = 0
    if sh >= 24:
        sh -= 24
        start_extra = 1
    end_extra = 0
    if eh >= 24:
        eh -= 24
        end_extra = 1

    start_dt = base_date.replace(hour=sh, minute=sm, second=0) + timedelta(days=start_extra)
    end_dt = base_date.replace(hour=eh, minute=em, second=0) + timedelta(days=end_extra)

    if next_day or (end_dt <= start_dt):
        end_dt += timedelta(days=1)

    return (start_dt, end_dt)


def get_cell(ws, row, col):
    """Get cell value at given row and column letter."""
    return ws[f"{col}{row}"].value


def parse_employees(ws):
    """Parse all employees and their shifts from the Planning sheet."""
    employees = {}
    current_name = None
    current_rows = []

    for row in range(5, ws.max_row + 1):
        name_cell = get_cell(ws, row, "A")
        if name_cell and isinstance(name_cell, str) and name_cell.strip():
            # Save previous employee
            if current_name:
                employees[current_name] = parse_shifts(ws, current_rows)
            current_name = name_cell.strip()
            current_rows = [row]
        elif current_name:
            current_rows.append(row)

    # Save last employee
    if current_name:
        employees[current_name] = parse_shifts(ws, current_rows)

    return employees


def parse_shifts(ws, rows):
    """Parse shifts from a list of rows for one employee."""
    events = []

    # Process rows in pairs: code row, then time row
    i = 0
    while i < len(rows):
        row = rows[i]
        # Check if this row has codes (non-time values)
        has_codes = False
        codes = {}
        for col in COLS:
            val = get_cell(ws, row, col)
            if val and isinstance(val, str):
                val = val.strip()
                if val and not re.match(r"^\d{2}:\d{2}/\d{2}:\d{2}", val):
                    has_codes = True
                    codes[col] = val

        if has_codes:
            # Next row should have times
            times = {}
            if i + 1 < len(rows):
                time_row = rows[i + 1]
                for col in COLS:
                    val = get_cell(ws, time_row, col)
                    if val and isinstance(val, str) and re.match(r"^\d{2}:\d{2}/\d{2}:\d{2}", val.strip()):
                        times[col] = val.strip()

            # Create events from code+time pairs
            for col, code in codes.items():
                if col in times and col in WEEK_DATES:
                    parsed = parse_time(times[col], WEEK_DATES[col])
                    if parsed:
                        label = CODE_NAMES.get(code, code)
                        events.append({
                            "code": code,
                            "label": label,
                            "start": parsed[0],
                            "end": parsed[1],
                        })

            i += 2  # Skip code row + time row
        else:
            i += 1

    # Sort events by start time
    events.sort(key=lambda e: e["start"])
    return events


def generate_ics(name, events, week_num):
    """Generate ICS content for an employee."""
    s = slug(name)
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Planning Urban 7D//FR",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:Planning {name}",
        "X-WR-TIMEZONE:Europe/Paris",
        "BEGIN:VTIMEZONE",
        "TZID:Europe/Paris",
        "BEGIN:STANDARD",
        "TZOFFSETFROM:+0200",
        "TZOFFSETTO:+0100",
        "TZNAME:CET",
        "DTSTART:19701025T030000",
        "RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=-1SU",
        "END:STANDARD",
        "BEGIN:DAYLIGHT",
        "TZOFFSETFROM:+0100",
        "TZOFFSETTO:+0200",
        "TZNAME:CEST",
        "DTSTART:19700329T020000",
        "RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=-1SU",
        "END:DAYLIGHT",
        "END:VTIMEZONE",
    ]

    for i, evt in enumerate(events, 1):
        dt_fmt = lambda d: d.strftime("%Y%m%dT%H%M%S")
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{s}-s{week_num}-{i}@urban7d",
            f"DTSTAMP:{dt_fmt(evt['start'])}",
            f"DTSTART:{dt_fmt(evt['start'])}",
            f"DTEND:{dt_fmt(evt['end'])}",
            f"SUMMARY:{evt['label']}",
            f"DESCRIPTION:{evt['label']}",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    return "\r\n".join(lines)


def generate_html(employees, week_num, all_weeks):
    """Generate the HTML page for a given week."""
    active_names = [name for name, evts in employees.items() if len(evts) > 0]

    week_tabs = ""
    for w in all_weeks:
        if w == week_num:
            week_tabs += f'            <a href="#" class="week-tab active">S{w}</a>\n'
        else:
            week_tabs += f'            <a href="S{w}.html" class="week-tab ">S{w}</a>\n'

    employee_lines = ""
    for name in employees:
        s = slug(name)
        if name in active_names:
            employee_lines += f'            <a href="ics/{s}.ics" class="employee">{name}</a>\n'
        else:
            employee_lines += f'            <div class="employee repos">{name} <span class="badge">Repos</span></div>\n'

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Planning Urban 7D - S{week_num}</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Inter', sans-serif; background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); min-height: 100vh; padding: 20px; }}
        .container {{ max-width: 500px; margin: 0 auto; }}
        .header {{ text-align: center; margin-bottom: 20px; padding: 20px; }}
        .logo {{ font-size: 48px; margin-bottom: 10px; }}
        h1 {{ color: #FF6B35; font-size: 28px; font-weight: 700; margin-bottom: 8px; }}
        .subtitle {{ color: #888; font-size: 14px; margin-bottom: 5px; }}
        .dates {{ color: #FF6B35; font-size: 18px; font-weight: 600; background: rgba(255, 107, 53, 0.1); padding: 10px 20px; border-radius: 20px; display: inline-block; margin-top: 10px; }}
        .week-selector {{ display: flex; justify-content: center; gap: 8px; margin-bottom: 25px; flex-wrap: wrap; }}
        .week-tab {{ padding: 10px 18px; background: rgba(255, 255, 255, 0.05); border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 25px; color: #888; text-decoration: none; font-weight: 500; font-size: 14px; transition: all 0.2s ease; }}
        .week-tab:hover {{ background: rgba(255, 107, 53, 0.1); border-color: rgba(255, 107, 53, 0.3); color: #FF6B35; }}
        .week-tab.active {{ background: #FF6B35; border-color: #FF6B35; color: white; }}
        .employees {{ display: flex; flex-direction: column; gap: 8px; }}
        .employee {{ display: flex; align-items: center; justify-content: space-between; padding: 16px 20px; background: rgba(255, 255, 255, 0.05); border-radius: 12px; color: white; text-decoration: none; font-weight: 500; transition: all 0.2s ease; border: 1px solid rgba(255, 255, 255, 0.1); }}
        a.employee:hover {{ background: rgba(255, 107, 53, 0.2); border-color: #FF6B35; transform: translateX(5px); }}
        a.employee::after {{ content: '\U0001F4C5'; font-size: 20px; }}
        .employee.repos {{ color: #555; background: rgba(255, 255, 255, 0.02); border-color: rgba(255, 255, 255, 0.05); }}
        .badge {{ font-size: 11px; padding: 4px 10px; background: rgba(255, 255, 255, 0.1); border-radius: 20px; color: #555; font-weight: 400; }}
        .footer {{ text-align: center; margin-top: 30px; color: #666; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="logo">\u26bd</div>
            <h1>Planning Urban 7D</h1>
            <p class="subtitle">Semaine {week_num}</p>
            <div class="dates">2 Mars \u2192 8 Mars</div>
        </div>
        <div class="week-selector">
{week_tabs.rstrip()}
        </div>
        <div class="employees">
{employee_lines.rstrip()}
        </div>
        <div class="footer"><p>Cliquez sur votre nom pour ajouter le planning \u00e0 votre calendrier</p></div>
    </div>
</body>
</html>"""


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Planning"]

    # Parse employees
    employees = parse_employees(ws)

    print(f"Employés trouvés: {len(employees)}")
    active = {name: evts for name, evts in employees.items() if len(evts) > 0}
    print(f"Employés actifs (avec shifts): {len(active)}")

    for name, evts in employees.items():
        if evts:
            print(f"\n  {name} ({len(evts)} événements):")
            for e in evts:
                print(f"    {e['start'].strftime('%a %d/%m %H:%M')} - {e['end'].strftime('%H:%M')} : {e['label']}")

    # Generate ICS files
    os.makedirs("ics", exist_ok=True)
    for name, evts in employees.items():
        if evts:
            ics_content = generate_ics(name, evts, WEEK_NUM)
            filename = f"ics/{slug(name)}.ics"
            with open(filename, "w", encoding="utf-8") as f:
                f.write(ics_content)
            print(f"  Écrit: {filename}")

    # Generate JSON
    os.makedirs("data", exist_ok=True)
    json_data = {
        "semaine": WEEK_NUM,
        "annee": YEAR,
        "date_debut": "2 Mars",
        "date_fin": "8 Mars",
        "employesActifs": sorted([name for name in active])
    }
    with open(f"data/S{WEEK_NUM}.json", "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    print(f"\nÉcrit: data/S{WEEK_NUM}.json")

    # Generate HTML
    all_weeks = [3, 5, 6, 10]
    html_content = generate_html(employees, WEEK_NUM, all_weeks)
    with open(f"S{WEEK_NUM}.html", "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Écrit: S{WEEK_NUM}.html")

    print("\nTerminé !")


if __name__ == "__main__":
    main()
