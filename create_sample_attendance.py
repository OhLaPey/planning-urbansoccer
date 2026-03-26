#!/usr/bin/env python3
"""Crée un fichier Excel de présences PSG Academy de démonstration."""

import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(SCRIPT_DIR, "PSG Academy Présences 2025-2026.xlsx")

# Couleurs
PINK_FILL = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
LIGHT_PINK = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
CYAN_FILL = PatternFill(start_color="00CED1", end_color="00CED1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
TOTAL_FONT = Font(bold=True, size=12)

# Données de démo — 10 créneaux comme dans le vrai fichier EDF
CRENEAUX = [
    {
        "name": "LUNDI 17H30 Présence",
        "title": "LUNDI 17H30",
        "start_date": datetime(2025, 9, 8),
        "day_offset": 0,  # lundi
        "groups": [
            {
                "kids": [
                    ("BERNARD Ethan", "U11"),
                    ("THOMAS Axel", "U11"),
                    ("LEROY Mathis", "U11"),
                    ("GIRARD Louis", "U12"),
                    ("BONNET Enzo", "U12"),
                    ("FONTAINE Adam", "U12"),
                ]
            },
        ],
    },
    {
        "name": "MARDI 17H30 Présence",
        "title": "MARDI 17H30",
        "start_date": datetime(2025, 9, 9),
        "day_offset": 1,  # mardi
        "groups": [
            {
                "kids": [
                    ("DUPONT Mathéo", "U10"),
                    ("MOREL Nathan", "U10"),
                    ("FOURNIER Enzo", "U10"),
                    ("LAMBERT Théo", "U11"),
                    ("ROUSSEAU Lucas", "U11"),
                    ("BLANC Ethan", "U11"),
                    ("CHEVALIER Liam", "U12"),
                ]
            },
        ],
    },
    {
        "name": "MERCREDI 14h Présence",
        "title": "MERCREDI 14H",
        "start_date": datetime(2025, 9, 10),
        "day_offset": 2,  # mercredi
        "groups": [
            {
                "kids": [
                    ("MARTIN Lucas", "U10"),
                    ("DUBOIS Nathan", "U10"),
                    ("PETIT Arthur", "U10"),
                    ("ROBERT Léo", "U10"),
                    ("RICHARD Hugo", "U11"),
                    ("MOREAU Théo", "U11"),
                    ("SIMON Gabriel", "U11"),
                    ("LAURENT Raphaël", "U11"),
                ]
            },
        ],
    },
    {
        "name": "MERCREDI 15H45 Présence",
        "title": "MERCREDI 15H45",
        "start_date": datetime(2025, 9, 10),
        "day_offset": 2,  # mercredi
        "groups": [
            {
                "kids": [
                    ("ABOU LEYLA Youssef", "U10"),
                    ("LEGRAND LHOPITAULT Gabriel", "U10"),
                    ("NGUEMA MEDANG Seth", "U10"),
                    ("MEDALE Joaquim", "U10"),
                    ("BESNARD Noah", "U10"),
                    ("DESPOISSE Louis", "U10"),
                    ("SACADURA DOS SANTOS Thomas", "U10"),
                    ("KARIMI Joal-Liam", "U10"),
                    ("DELGADO Emiliano", "U10"),
                    ("MANNA Gabriel", "U10"),
                    ("FARSI Amir", "U10"),
                    ("AMRI Souleymane", "U10"),
                    ("MAHAMOUD IBRAHIM Zaïn", "U10"),
                    ("NICOLLE Noah", "U10"),
                    ("AMER Ayman", "U10"),
                    ("AZIZA Sam", "U10"),
                ]
            },
            {
                "kids": [
                    ("MOREIRA Liam", "U11"),
                    ("DEVOS Adam", "U11"),
                    ("MOLINA NEGRE Ruben", "U11"),
                    ("MOUMEN Maher", "U11"),
                    ("MOUMEN Waël", "U11"),
                    ("BEAUMONT Elie", "U11"),
                    ("GOMIS Noam", "U11"),
                    ("BAMBA Karim Christ Emmanuel", "U12"),
                    ("ARBONA Florian", "U12"),
                    ("MAGOSSOU Melphy Antoine", "U12"),
                    ("ABOU LEYA Adam", "U12"),
                    ("MARTIN CHOPINAUD Louis", "U12"),
                    ("MOREIRA Lucas", "U12"),
                    ("GUERRERO Nino", "U12"),
                    ("GONCALVES Gabin", "U12"),
                    ("GARBAA Ilyas", "U12"),
                ]
            },
        ],
    },
    {
        "name": "VENDREDI 17h30 Présence",
        "title": "VENDREDI 17H30",
        "start_date": datetime(2025, 9, 12),
        "day_offset": 4,  # vendredi
        "groups": [
            {
                "kids": [
                    ("PEREIRA Nolan", "U11"),
                    ("DA SILVA Yanis", "U11"),
                    ("COSTA Léandro", "U11"),
                    ("FERREIRA Mattéo", "U12"),
                    ("SANTOS Rafael", "U12"),
                    ("OLIVEIRA Ilian", "U12"),
                    ("RODRIGUES Noam", "U12"),
                    ("ALVES Ayoub", "U13"),
                ]
            },
        ],
    },
    {
        "name": "SAMEDI 9h30 Présence",
        "title": "SAMEDI 9H30",
        "start_date": datetime(2025, 9, 13),
        "day_offset": 5,  # samedi
        "groups": [
            {
                "kids": [
                    ("LECLERC Noé", "U10"),
                    ("GARNIER Rayan", "U10"),
                    ("FAURE Amine", "U10"),
                    ("MERCIER Timéo", "U10"),
                    ("GAUTHIER Ilyes", "U10"),
                    ("MARCHAND Léon", "U11"),
                    ("DUVAL Nolan", "U11"),
                    ("DENIS Aymen", "U11"),
                    ("LEMOINE Soan", "U11"),
                ]
            },
        ],
    },
    {
        "name": "SAMEDI 11h15 Présence",
        "title": "SAMEDI 11H15",
        "start_date": datetime(2025, 9, 13),
        "day_offset": 5,  # samedi
        "groups": [
            {
                "kids": [
                    ("DOYHARCABAL Andoni", "U10"),
                    ("ABOU LEYLA Youssef", "U10"),
                    ("HAMZA Yanis", "U10"),
                    ("MEDALE Joaquim", "U10"),
                    ("AFCHARD MONTAIGNE Victor", "U10"),
                    ("DZIEDZIC Théo", "U10"),
                    ("CAMY Simon", "U10"),
                    ("RENJITH Ritwik", "U10"),
                    ("CALSAT Abel", "U10"),
                    ("AHMED Khalil", "U10"),
                    ("MECALIFF BERTRAND Martin", "U10"),
                ]
            },
            {
                "kids": [
                    ("DEVOS Adam", "U11"),
                    ("RAMANANTOSA Nolan", "U11"),
                    ("BLANCO Rayane", "U11"),
                    ("CASTEL Emile", "U11"),
                    ("LEPREUX Louis", "U12"),
                    ("FAGOO DOREMUS Nolan", "U12"),
                    ("ALBERT Kenan", "U12"),
                    ("ABOU LEYA Adam", "U12"),
                    ("GRANFILS Stanislas", "U12"),
                    ("MAGOSSOU Melphy Antoine", "U12"),
                    ("BERTHELOT Adel", "U12"),
                    ("BREHAULT Camille", "U12"),
                    ("DELCROIX Raphaël", "U13"),
                    ("VACHON LAFFORGUE Matthieu", "U13"),
                    ("MAXENCE", "U13"),
                ]
            },
        ],
    },
    {
        "name": "SAMEDI 14h00 Présence",
        "title": "SAMEDI 14H00",
        "start_date": datetime(2025, 9, 13),
        "day_offset": 5,  # samedi
        "groups": [
            {
                "kids": [
                    ("HENRY Raphaël", "U11"),
                    ("MICHEL Noah", "U11"),
                    ("LEFEBVRE Mathis", "U11"),
                    ("DAVID Liam", "U12"),
                    ("BERTRAND Enzo", "U12"),
                    ("ROUX Timéo", "U12"),
                    ("VINCENT Lucas", "U13"),
                    ("MULLER Yanis", "U13"),
                ]
            },
        ],
    },
    {
        "name": "SAMEDI 15h45 Présence",
        "title": "SAMEDI 15H45",
        "start_date": datetime(2025, 9, 13),
        "day_offset": 5,  # samedi
        "groups": [
            {
                "kids": [
                    ("POULAIN Gabriel", "U12"),
                    ("COLLET Noé", "U12"),
                    ("MASSON Rayan", "U12"),
                    ("AUBERT Ilyes", "U13"),
                    ("RENAUD Soan", "U13"),
                    ("PICARD Nolan", "U13"),
                    ("ROGER Aymen", "U13"),
                ]
            },
        ],
    },
    {
        "name": "BABY Présence",
        "title": "BABY",
        "start_date": datetime(2025, 9, 13),
        "day_offset": 5,  # samedi
        "groups": [
            {
                "kids": [
                    ("LOPEZ Timéo", "BABY"),
                    ("GARCIA Noé", "BABY"),
                    ("MARTINEZ Sacha", "BABY"),
                    ("ANDERSON Milo", "BABY"),
                    ("WILSON Jade", "BABY"),
                ]
            },
        ],
    },
]

# Vacances scolaires (semaines de vacances - approximatives)
VACATION_WEEKS = [
    (datetime(2025, 10, 18), datetime(2025, 11, 3)),   # Toussaint
    (datetime(2025, 12, 20), datetime(2026, 1, 5)),     # Noël
    (datetime(2026, 2, 14), datetime(2026, 3, 2)),      # Hiver
]

import random
random.seed(42)


def is_vacation(d):
    for start, end in VACATION_WEEKS:
        if start <= d <= end:
            return True
    return False


def create_sheet(wb, creneau):
    ws = wb.create_sheet(title=creneau["name"])

    # Row 2: Title
    ws.merge_cells("A2:Z2")
    ws["A2"].value = creneau["title"]
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    # Generate session dates
    start = creneau["start_date"]
    sessions = []
    current = start
    session_num = 1
    while current < datetime(2026, 7, 1):
        if is_vacation(current):
            sessions.append({"date": current, "label": "V", "is_vac": True})
        else:
            sessions.append({"date": current, "label": f"S{session_num}", "is_vac": False})
            session_num += 1
        current += timedelta(days=7)

    # Limit to ~32 sessions
    sessions = sessions[:32]

    first_data_col = 5  # Column E

    # Row 5: Dates
    for i, sess in enumerate(sessions):
        col = first_data_col + i
        ws.cell(row=5, column=col).value = sess["date"]
        ws.cell(row=5, column=col).number_format = "DD/MM/YY"
        ws.cell(row=5, column=col).font = Font(size=9, bold=True)
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="center",
                                                          text_rotation=90)
        if sess["is_vac"]:
            ws.cell(row=5, column=col).fill = CYAN_FILL

    # Row 6: Session labels (S1, S2, V, ...)
    ws.cell(row=6, column=3).value = f"MAJ LISTE {datetime.now().strftime('%d/%m')} : {datetime.now().strftime('%Hh%M')}"
    ws.cell(row=6, column=3).font = Font(bold=True, color="FF0000", size=9)

    for i, sess in enumerate(sessions):
        col = first_data_col + i
        ws.cell(row=6, column=col).value = sess["label"]
        ws.cell(row=6, column=col).font = Font(bold=True, size=10)
        ws.cell(row=6, column=col).alignment = Alignment(horizontal="center")
        if sess["is_vac"]:
            ws.cell(row=6, column=col).fill = CYAN_FILL

    # Column widths
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 5

    current_row = 7

    for g_idx, group in enumerate(creneau["groups"]):
        for k_idx, (name, cat) in enumerate(group["kids"]):
            row = current_row + k_idx
            ws.cell(row=row, column=2).value = k_idx + 1
            ws.cell(row=row, column=3).value = name
            ws.cell(row=row, column=3).font = Font(bold=True)
            ws.cell(row=row, column=4).value = cat

            # Fill in attendance: past sessions get 1/0
            # Only fill for sessions up to "current" date (March 2026)
            for i, sess in enumerate(sessions):
                col = first_data_col + i
                if sess["is_vac"]:
                    ws.cell(row=row, column=col).fill = CYAN_FILL
                    continue
                if sess["date"] < datetime(2026, 3, 10):
                    # Random attendance (weighted towards present ~80%)
                    val = 1 if random.random() < 0.8 else 0
                    ws.cell(row=row, column=col).value = val
                    if val == 1:
                        ws.cell(row=row, column=col).fill = CYAN_FILL
                    else:
                        ws.cell(row=row, column=col).fill = LIGHT_PINK

            # Pink background for name rows
            ws.cell(row=row, column=2).fill = PINK_FILL
            ws.cell(row=row, column=3).fill = PINK_FILL
            ws.cell(row=row, column=4).fill = PINK_FILL

        # TOTAL row
        total_row = current_row + len(group["kids"])
        ws.cell(row=total_row, column=2).value = len(group["kids"])
        ws.cell(row=total_row, column=3).value = f"TOTAL G{g_idx + 1}"
        ws.cell(row=total_row, column=3).font = TOTAL_FONT
        ws.cell(row=total_row, column=3).fill = PINK_FILL

        for i, sess in enumerate(sessions):
            col = first_data_col + i
            if sess["is_vac"]:
                continue
            # Sum for total
            total = 0
            for k_idx in range(len(group["kids"])):
                v = ws.cell(row=current_row + k_idx, column=col).value
                if v and isinstance(v, int):
                    total += v
            ws.cell(row=total_row, column=col).value = total
            ws.cell(row=total_row, column=col).font = Font(bold=True)

        current_row = total_row + 3  # Skip some rows before next group

    return ws


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for creneau in CRENEAUX:
        create_sheet(wb, creneau)

    # Add a Listing sheet (non-presence)
    ws = wb.create_sheet(title="Listing Utilisateurs")
    ws["A1"].value = "Listing des inscrits"

    wb.save(OUTPUT)
    print(f"Fichier créé : {OUTPUT}")


if __name__ == "__main__":
    main()
