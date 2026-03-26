#!/usr/bin/env python3
"""
Module de parsing et écriture du fichier Excel de présences EDF / PSG Academy.

Gère automatiquement les différents layouts de colonnes :
  - Layout A : Col B=numéro, Col C=nom, Col D=catégorie (fichier sample)
  - Layout B : Col B=nom, Col C=catégorie (fichier "Suivi des présences EDF.xlsx")

Détecte automatiquement les onglets « …Présence », les groupes d'enfants,
les séances (S1, S2…, V pour vacances, F pour férié) et les lignes TOTAL.
"""

import os
import re
import glob as glob_mod
from datetime import datetime

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Regex pour détecter les catégories d'âge
_CAT_RE = re.compile(r"^(U\d+|BABY|MINI|POUSSIN)$", re.IGNORECASE)


def find_attendance_file():
    """Trouve le fichier Excel de présences."""
    patterns = [
        # Fichier EDF (prioritaire — c'est le fichier de suivi réel)
        os.path.join(SCRIPT_DIR, "*Suivi*présences*EDF*xlsx"),
        os.path.join(SCRIPT_DIR, "*Suivi*presences*EDF*xlsx"),
        os.path.join(SCRIPT_DIR, "*suivi*présences*xlsx"),
        os.path.join(SCRIPT_DIR, "*suivi*presences*xlsx"),
        # Fichier PSG Academy
        os.path.join(SCRIPT_DIR, "*PSG*résence*xlsx"),
        os.path.join(SCRIPT_DIR, "*PSG*resence*xlsx"),
        os.path.join(SCRIPT_DIR, "*PSG*Academy*xlsx"),
        os.path.join(SCRIPT_DIR, "PSG*.xlsx"),
    ]
    for pat in patterns:
        matches = glob_mod.glob(pat, recursive=False)
        # Exclure les fichiers temporaires Excel (~$...)
        matches = [m for m in matches if not os.path.basename(m).startswith("~$")]
        if matches:
            return matches[0]
    # Fallback : tout fichier avec "Présence" ou "Presence"
    for pat in [
        os.path.join(SCRIPT_DIR, "*résence*xlsx"),
        os.path.join(SCRIPT_DIR, "*resence*xlsx"),
        os.path.join(SCRIPT_DIR, "*Presence*xlsx"),
    ]:
        matches = glob_mod.glob(pat, recursive=False)
        matches = [m for m in matches
                   if "Plannings 2026" not in m
                   and not os.path.basename(m).startswith("~$")]
        if matches:
            return matches[0]
    return None


def parse_attendance(filepath=None):
    """Parse le fichier Excel de présences. Retourne un dict structuré."""
    if filepath is None:
        filepath = find_attendance_file()
    if not filepath or not os.path.exists(filepath):
        return {"file": None, "creneaux": []}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"file": filepath, "creneaux": []}

    for sheet_name in wb.sheetnames:
        if not _is_presence_sheet(sheet_name):
            continue
        ws = wb[sheet_name]
        creneau = _parse_sheet(ws, sheet_name)
        if creneau and creneau["groups"]:
            result["creneaux"].append(creneau)

    wb.close()
    return result


def _is_presence_sheet(name):
    """Détermine si un onglet est un onglet de présence."""
    name_lower = name.lower()
    return "présence" in name_lower or "presence" in name_lower


def _slugify(text):
    """Crée un slug URL-friendly."""
    import unicodedata
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text.lower())
    return re.sub(r"[-\s]+", "-", text).strip("-")


def _detect_columns(ws, session_row, first_session_col, max_row):
    """
    Auto-détecte les colonnes nom et catégorie dans les lignes de données.

    Examine les premières lignes de données après le header pour déterminer
    quel layout est utilisé.

    Retourne (name_col, category_col, has_num_col).
    """
    data_start = session_row + 1
    scan_end = min(data_start + 15, max_row + 1)

    # Colonnes candidates (avant la première colonne de session)
    candidate_cols = list(range(1, first_session_col))

    # Compter les types de valeurs par colonne
    col_stats = {}
    for col in candidate_cols:
        col_stats[col] = {"nums": 0, "names": 0, "cats": 0, "empty": 0, "total_text": 0}

    for row in range(data_start, scan_end):
        for col in candidate_cols:
            val = ws.cell(row=row, column=col).value
            if val is None or (isinstance(val, str) and not val.strip()):
                col_stats[col]["empty"] += 1
                continue
            s = str(val).strip()
            if _CAT_RE.match(s):
                col_stats[col]["cats"] += 1
            elif re.match(r"^\d+$", s):
                col_stats[col]["nums"] += 1
            elif len(s) > 2 and any(c.isalpha() for c in s):
                col_stats[col]["names"] += 1
                col_stats[col]["total_text"] += len(s)

    # Identifier la colonne nom (la plus de noms + texte le plus long)
    name_col = None
    best_name_score = 0
    for col, stats in col_stats.items():
        score = stats["names"] * 100 + stats["total_text"]
        if score > best_name_score:
            best_name_score = score
            name_col = col

    # Identifier la colonne catégorie (la plus de catégories)
    cat_col = None
    best_cat = 0
    for col, stats in col_stats.items():
        if col != name_col and stats["cats"] > best_cat:
            best_cat = stats["cats"]
            cat_col = col

    # Y a-t-il une colonne de numérotation ?
    has_num = False
    num_col = None
    for col, stats in col_stats.items():
        if col != name_col and col != cat_col and stats["nums"] > 3:
            has_num = True
            num_col = col

    # Fallbacks si la détection échoue
    if name_col is None:
        name_col = 3 if first_session_col > 4 else 2
    if cat_col is None:
        cat_col = name_col + 1 if name_col + 1 < first_session_col else None

    return name_col, cat_col, has_num


def _parse_sheet(ws, sheet_name):
    """Parse un onglet de présence."""
    max_row = ws.max_row or 50
    max_col = ws.max_column or 50

    # 1. Trouver le titre (ligne 1-3, cellule non vide > 3 chars, hors MAJ LISTE)
    title = ""
    for row in range(1, min(5, max_row + 1)):
        for col in range(1, min(15, max_col + 1)):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and len(val) > 3:
                v = val.strip()
                if "MAJ LISTE" not in v.upper() and "TOTAL" not in v.upper():
                    title = v
                    break
        if title:
            break
    if not title:
        title = sheet_name.replace(" Présence", "").replace(" Presence", "").strip()

    slug = _slugify(title or sheet_name)

    # 2. Trouver la ligne des sessions (première occurrence de S1)
    session_row = None
    date_row = None
    for row in range(1, min(12, max_row + 1)):
        for col in range(1, min(max_col + 1, 100)):
            v = ws.cell(row=row, column=col).value
            if v is not None and str(v).strip() == "S1":
                session_row = row
                date_row = row - 1 if row > 1 else None
                break
        if session_row:
            break

    if not session_row:
        return None

    # 3. Parser les sessions
    sessions = []
    first_session_col = None
    for col in range(1, min(max_col + 1, 100)):
        val = ws.cell(row=session_row, column=col).value
        if val is None:
            continue
        label = str(val).strip()
        if re.match(r"^S\d+$", label) or label in ("V", "F"):
            is_vacation = label in ("V", "F")
            # Date correspondante
            date_str = None
            if date_row:
                date_val = ws.cell(row=date_row, column=col).value
                if date_val:
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime("%d/%m/%y")
                    else:
                        date_str = str(date_val).strip()
            sessions.append({
                "col": col,
                "label": label,
                "date": date_str,
                "is_vacation": is_vacation,
            })
            if first_session_col is None:
                first_session_col = col

    if not sessions or not first_session_col:
        return None

    # 4. Auto-détecter les colonnes nom/catégorie
    name_col, cat_col, has_num = _detect_columns(ws, session_row, first_session_col, max_row)

    # 5. Parser les enfants et groupes
    groups = []
    current_group = {"name": "Groupe 1", "kids": [], "total_row": None}
    group_num = 1
    kid_num = 0
    data_start_row = session_row + 1

    for row in range(data_start_row, max_row + 1):
        # Chercher TOTAL dans toutes les colonnes avant les sessions
        is_total = False
        for col in range(1, first_session_col):
            v = ws.cell(row=row, column=col).value
            if v and isinstance(v, str) and "TOTAL" in v.upper():
                is_total = True
                break

        if is_total:
            current_group["total_row"] = row
            if current_group["kids"]:
                groups.append(current_group)
            group_num += 1
            kid_num = 0
            current_group = {"name": f"Groupe {group_num}", "kids": [], "total_row": None}
            continue

        # Lire le nom
        name_val = ws.cell(row=row, column=name_col).value
        if not name_val or not isinstance(name_val, str) or not name_val.strip():
            continue
        name = name_val.strip()
        # Ignorer les lignes qui ne sont pas des noms (headers répétés, etc.)
        if name.startswith("S") and re.match(r"^S\d+$", name):
            continue
        if "MAJ LISTE" in name.upper() or "TOTAL" in name.upper():
            continue
        if "INSCRIPTION" in name.upper():
            continue

        # Catégorie
        category = ""
        if cat_col:
            cat_val = ws.cell(row=row, column=cat_col).value
            if cat_val:
                cs = str(cat_val).strip()
                if _CAT_RE.match(cs):
                    category = cs

        # Numérotation auto
        kid_num += 1

        # Lire les présences
        attendance = {}
        for sess in sessions:
            cell_val = ws.cell(row=row, column=sess["col"]).value
            if cell_val is not None:
                try:
                    attendance[sess["label"]] = int(cell_val)
                except (ValueError, TypeError):
                    attendance[sess["label"]] = None
            else:
                attendance[sess["label"]] = None

        current_group["kids"].append({
            "row": row,
            "num": kid_num,
            "name": name,
            "category": category,
            "attendance": attendance,
        })

    # Ajouter le dernier groupe s'il a des enfants
    if current_group["kids"]:
        groups.append(current_group)

    return {
        "sheet_name": sheet_name,
        "title": title,
        "slug": slug,
        "sessions": sessions,
        "groups": groups,
    }


def save_attendance(filepath, sheet_name, updates):
    """
    Écrit les présences dans le fichier Excel.
    updates : liste de {"row": int, "col": int, "value": int}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    for u in updates:
        cell = ws.cell(row=u["row"], column=u["col"])
        cell.value = u["value"]

    _update_totals(ws)
    wb.save(filepath)
    wb.close()


def _update_totals(ws):
    """Recalcule les lignes TOTAL après mise à jour."""
    max_row = ws.max_row or 50
    max_col = ws.max_column or 50

    # Trouver la ligne session
    session_row = None
    first_session_col = None
    for row in range(1, min(12, max_row + 1)):
        for col in range(1, min(max_col + 1, 100)):
            if str(ws.cell(row=row, column=col).value or "").strip() == "S1":
                session_row = row
                first_session_col = col
                break
        if session_row:
            break

    if not session_row:
        return

    # Colonnes de sessions
    session_cols = []
    for col in range(1, min(max_col + 1, 100)):
        val = str(ws.cell(row=session_row, column=col).value or "").strip()
        if re.match(r"^S\d+$", val) or val in ("V", "F"):
            session_cols.append(col)

    # Recalculer les TOTAL
    data_start = session_row + 1
    group_start = data_start

    for row in range(data_start, max_row + 1):
        is_total = False
        for col in range(1, first_session_col):
            v = ws.cell(row=row, column=col).value
            if v and isinstance(v, str) and "TOTAL" in v.upper():
                is_total = True
                break

        if is_total:
            for col in session_cols:
                total = 0
                for r in range(group_start, row):
                    val = ws.cell(row=r, column=col).value
                    if val is not None:
                        try:
                            total += int(val)
                        except (ValueError, TypeError):
                            pass
                ws.cell(row=row, column=col).value = total
            group_start = row + 1


def get_current_session(sessions):
    """Détermine la séance en cours. Retourne le label (ex: "S12")."""
    today = datetime.now()
    last_session = None
    for sess in sessions:
        if sess["is_vacation"]:
            continue
        if sess["date"]:
            for fmt in ("%d/%m/%y", "%d/%m/%Y"):
                try:
                    d = datetime.strptime(sess["date"], fmt)
                    if d.date() <= today.date():
                        last_session = sess
                    break
                except ValueError:
                    continue

    if last_session:
        return last_session["label"]

    # Fallback : première séance
    for sess in sessions:
        if not sess["is_vacation"]:
            return sess["label"]
    return sessions[0]["label"] if sessions else None


def get_session_col(sessions, label):
    """Retourne le numéro de colonne pour un label de séance."""
    for sess in sessions:
        if sess["label"] == label:
            return sess["col"]
    return None
