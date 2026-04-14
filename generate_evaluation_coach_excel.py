#!/usr/bin/env python3
"""Génère la fiche d'évaluation coach PSG Academy au format Excel."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PSG_RED = "E30613"
PSG_BLUE = "0C1C3E"
DARK_ROW = "152A50"
LIGHT_ROW = "1A3260"
WHITE = "FFFFFF"
GRAY = "888888"
GREEN = "4CAF50"
AMBER = "FF9800"

thin_border = Border(
    left=Side(style="thin", color="2A3F6A"),
    right=Side(style="thin", color="2A3F6A"),
    top=Side(style="thin", color="2A3F6A"),
    bottom=Side(style="thin", color="2A3F6A"),
)

red_border = Border(
    left=Side(style="thin", color=PSG_RED),
    right=Side(style="thin", color=PSG_RED),
    top=Side(style="thin", color=PSG_RED),
    bottom=Side(style="thin", color=PSG_RED),
)


def style_cell(ws, row, col, value="", font_size=10, bold=False, color=WHITE,
               fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=font_size, bold=bold, color=color)
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.alignment = alignment or Alignment(vertical="center")
    if border:
        cell.border = border
    return cell


def section_title(ws, row, title, col_end=8):
    for c in range(1, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color=PSG_RED, end_color=PSG_RED, fill_type="solid")
        cell.border = Border(
            bottom=Side(style="medium", color=PSG_RED),
            top=Side(style="medium", color=PSG_RED),
        )
    style_cell(ws, row, 1, title, font_size=9, bold=True, color=WHITE,
               fill=PSG_RED, alignment=Alignment(vertical="center"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)


def info_field(ws, row, col_label, label, col_value, merge_end=None):
    style_cell(ws, row, col_label, label, font_size=8, bold=True, color=GRAY,
               fill=PSG_BLUE, border=thin_border)
    style_cell(ws, row, col_value, "", font_size=9, bold=False, color=WHITE,
               fill=DARK_ROW, border=red_border,
               alignment=Alignment(vertical="center"))
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col_value, end_row=row, end_column=merge_end)


CRITERES = [
    ("PRÉPARATION & ORGANISATION", [
        ("Préparation de la séance", "Séance planifiée, structurée, fil conducteur clair"),
        ("Matériel & Terrain", "Terrain prêt à l'arrivée, matériel installé et adapté"),
        ("Gestion du temps", "Début/fin à l'heure, temps de jeu optimisé, transitions fluides"),
    ]),
    ("PÉDAGOGIE & CONTENU", [
        ("Clarté des consignes", "Explications claires, adaptées à l'âge, démonstrations"),
        ("Progression des exercices", "Enchaînement logique, difficulté progressive"),
        ("Adaptation au niveau", "Ajuste selon le niveau et le nombre d'enfants"),
        ("Place du jeu", "Temps de jeu suffisant, situations ludiques, plaisir"),
    ]),
    ("COMMUNICATION & RELATION", [
        ("Communication enfants", "Ton adapté, encourage, prénoms, feedback positif"),
        ("Relation parents", "Accueil, échange début/fin de séance, disponibilité"),
        ("Énergie & Enthousiasme", "Dynamique, motivant, transmet sa passion"),
    ]),
    ("GESTION DU GROUPE", [
        ("Captation de l'attention", "Les enfants écoutent et sont engagés"),
        ("Discipline & Cadre", "Règles claires, gestion des comportements"),
        ("Inclusion de tous", "Chaque enfant participe, gestion des niveaux"),
    ]),
    ("SÉCURITÉ & ATTITUDE", [
        ("Vigilance & Sécurité", "Attentif aux risques, surveille tous les enfants"),
        ("Présentation", "Tenue PSG Academy, ponctualité, attitude pro"),
        ("Comportement exemplaire", "Langage adapté, fair-play, montre l'exemple"),
    ]),
]

NOTE_LABELS = ["N/N", "1", "2", "3", "4", "5"]


def generate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Évaluation Coach"
    ws.sheet_properties.tabColor = PSG_RED

    for row in range(1, 65):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=PSG_BLUE, end_color=PSG_BLUE, fill_type="solid")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 5
    ws.column_dimensions["H"].width = 5

    # ── HEADER ──
    ws.merge_cells("A1:H1")
    style_cell(ws, 1, 1, "ÉVALUATION COACH", font_size=14, bold=True, color=WHITE,
               fill=PSG_BLUE, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    style_cell(ws, 2, 1, "PSG Academy", font_size=8, bold=True, color=PSG_RED,
               fill=PSG_BLUE, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 13

    for c in range(1, 9):
        ws.cell(row=3, column=c).fill = PatternFill(start_color=PSG_RED, end_color=PSG_RED, fill_type="solid")
    ws.row_dimensions[3].height = 2

    # ── INFORMATIONS ──
    r = 4
    section_title(ws, r, "INFORMATIONS")
    ws.row_dimensions[r].height = 14

    r = 5
    info_field(ws, r, 1, "Coach", 2, merge_end=4)
    info_field(ws, r, 5, "Date", 6, merge_end=8)
    ws.row_dimensions[r].height = 18

    r = 6
    info_field(ws, r, 1, "Créneau", 2, merge_end=4)
    info_field(ws, r, 5, "Catégorie", 6, merge_end=8)
    ws.row_dimensions[r].height = 18

    r = 7
    info_field(ws, r, 1, "Observ.", 2, merge_end=8)
    ws.row_dimensions[r].height = 18

    # ── ÉVALUATION ──
    r = 8
    section_title(ws, r, "ÉVALUATION")
    ws.row_dimensions[r].height = 14

    r = 9
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    style_cell(ws, r, 1, "Critère", font_size=7, bold=True, color=GRAY, fill=PSG_BLUE,
               alignment=Alignment(vertical="center"))
    for i, label in enumerate(NOTE_LABELS):
        col = 3 + i
        style_cell(ws, r, col, label, font_size=7, bold=True, color=GRAY, fill=PSG_BLUE,
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=thin_border)
    ws.row_dimensions[r].height = 13

    r = 10
    alt = False
    cat_bg = "0A1630"
    for cat_name, items in CRITERES:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_cell(ws, r, 1, cat_name, font_size=7, bold=True, color=PSG_RED,
                   fill=cat_bg, alignment=Alignment(vertical="center"))
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = Border(
                bottom=Side(style="thin", color=PSG_RED),
                top=Side(style="thin", color=PSG_RED),
            )
            if c > 1:
                ws.cell(row=r, column=c).fill = PatternFill(start_color=cat_bg, end_color=cat_bg, fill_type="solid")
        ws.row_dimensions[r].height = 13
        r += 1

        for nom, desc in items:
            bg = LIGHT_ROW if alt else DARK_ROW
            alt = not alt

            style_cell(ws, r, 1, "", fill=bg)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            style_cell(ws, r, 1, nom, font_size=8, bold=True, color=WHITE, fill=bg,
                       border=thin_border, alignment=Alignment(vertical="center", indent=1))

            for i in range(6):
                col = 3 + i
                style_cell(ws, r, col, "", font_size=8, color=WHITE, fill=bg,
                           border=red_border,
                           alignment=Alignment(horizontal="center", vertical="center"))
            ws.row_dimensions[r].height = 15
            r += 1

            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            style_cell(ws, r, 1, desc, font_size=6, color=GRAY, fill=bg,
                       alignment=Alignment(vertical="center", indent=1))
            ws.row_dimensions[r].height = 10
            r += 1

    # ── BILAN ──
    section_title(ws, r, "BILAN")
    ws.row_dimensions[r].height = 14
    r += 1

    for label, color in [("POINTS POSITIFS", GREEN), ("AXES D'AMÉLIORATION", AMBER), ("COMMENTAIRE GÉNÉRAL", GRAY)]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_cell(ws, r, 1, label, font_size=7, bold=True, color=color, fill=PSG_BLUE,
                   alignment=Alignment(vertical="center"))
        ws.row_dimensions[r].height = 12
        r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_cell(ws, r, 1, "", font_size=8, color=WHITE, fill=DARK_ROW,
                   border=Border(bottom=Side(style="dotted", color="2A3F6A")),
                   alignment=Alignment(vertical="center", indent=1))
        ws.row_dimensions[r].height = 30
        r += 1

    # ── Mise en page impression ──
    ws.print_area = f"A1:H{r - 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    output = "Evaluation Coach - PSG Academy.xlsx"
    wb.save(output)
    print(f"Fichier généré : {output}")


if __name__ == "__main__":
    generate()
