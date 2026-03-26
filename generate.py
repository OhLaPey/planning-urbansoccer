#!/usr/bin/env python3
"""
Générateur de plannings Urban 7D — Abonnement calendrier.

Convertit les fichiers Excel « Plannings YYYY SXX.xlsx » en :
  - fichiers ICS (un par employé, cumulatif toutes semaines)
  - pages HTML (une par semaine, avec liens webcal://)
  - fichiers JSON (métadonnées par semaine)

Usage :
    python generate.py

Architecture :
    Excel (source) ──► generate.py ──► ics/ + HTML + data/
    Héberger sur un serveur web (GitHub Pages, Netlify, etc.)
    Les employés s'abonnent via webcal://domaine/ics/nom.ics

Le script détecte automatiquement tous les fichiers
« Plannings YYYY SXX.xlsx » présents dans le répertoire courant.
"""

import openpyxl
import json
import os
import re
from datetime import datetime, timedelta

# ── Mapping codes → noms lisibles + couleurs néon (basées sur l'Excel) ────

CODE_NAMES = {
    "VDC":   "Vie de centre",
    "L-REG": "Régisseur League",
    "CUP-L": "Cup League",
    "CUP-R": "Cup Régisseur",
    "STAGE": "Stage",
    "STA-E": "Stage encadrement",
    "C-PAD": "Cours Padel",
    "PAD-A": "Padel animation",
    "ANNIV": "Anniversaire",
    "INVEN": "Inventaire",
    "MAL":   "Maladie",
    "REU":   "Réunion",
    "EV-RE": "Événement régisseur",
    "AIDE":  "Aide",
    "P25M":  "P25M",
    "PSG":   "PSG Academy",
    "FOR-E": "Formation théorique",
    "FOR-P": "Formation pratique",
    "STA-P": "Stage Padel",
    "ENT-M": "Entretien Ménage",
    "EDF-A": "EDF-A",
    "EDF-B": "EDF-B",
    "EDF-C": "EDF-C",
    "COMMD": "Commercialisation",
    "EV-LO": "Événement logistique",
    "VIS-M": "Visite médicale",
    "FORE":  "Formation",
    "FORP":  "Formation pratique",
}

# Couleurs néon par code — inspirées de l'Excel avec effet glow
CODE_COLORS = {
    "VDC":   {"bg": "rgba(255,102,0,0.35)", "border": "#ff7832",  "text": "#ffb080"},
    "L-REG": {"bg": "rgba(180,100,255,0.35)", "border": "#b464ff",  "text": "#d4a0ff"},
    "CUP-L": {"bg": "rgba(255,255,255,0.25)", "border": "#ffffff",  "text": "#ffffff"},
    "CUP-R": {"bg": "rgba(100,220,60,0.35)",  "border": "#64dc3c",  "text": "#90ff70"},
    "STAGE": {"bg": "rgba(100,230,255,0.30)", "border": "#64e6ff",  "text": "#a0f0ff"},
    "STA-E": {"bg": "rgba(100,230,255,0.30)", "border": "#64e6ff",  "text": "#a0f0ff"},
    "C-PAD": {"bg": "rgba(180,180,180,0.30)", "border": "#b4b4b4",  "text": "#d0d0d0"},
    "PAD-A": {"bg": "rgba(180,180,180,0.30)", "border": "#b4b4b4",  "text": "#d0d0d0"},
    "ANNIV": {"bg": "rgba(0,176,240,0.40)",   "border": "#00b0f0",  "text": "#60d0ff"},
    "INVEN": {"bg": "rgba(255,192,0,0.40)",   "border": "#ffc000",  "text": "#ffd060"},
    "MAL":   {"bg": "rgba(255,80,80,0.35)",   "border": "#ff5050",  "text": "#ff8080"},
    "REU":   {"bg": "rgba(255,192,0,0.40)",   "border": "#ffc000",  "text": "#ffd060"},
    "EV-RE": {"bg": "rgba(100,220,60,0.35)",  "border": "#64dc3c",  "text": "#90ff70"},
    "AIDE":  {"bg": "rgba(0,176,240,0.40)",   "border": "#00b0f0",  "text": "#60d0ff"},
    "P25M":  {"bg": "rgba(255,102,0,0.40)",  "border": "#ff7832",  "text": "#ff9850"},
    "PSG":   {"bg": "rgba(255,102,0,0.40)",  "border": "#ff7832",  "text": "#ff9850"},
    "L-ARB": {"bg": "rgba(180,100,255,0.35)", "border": "#b464ff",  "text": "#d4a0ff"},
    "FOR-E": {"bg": "rgba(255,200,50,0.35)",  "border": "#ffc832",  "text": "#ffe080"},
    "FOR-P": {"bg": "rgba(50,200,120,0.35)",  "border": "#32c878",  "text": "#80ffb0"},
    "STA-P": {"bg": "rgba(100,230,255,0.30)", "border": "#64e6ff",  "text": "#a0f0ff"},
    "ENT-M": {"bg": "rgba(255,192,0,0.40)",   "border": "#ffc000",  "text": "#ffd060"},
    "EDF-A": {"bg": "rgba(0,112,192,0.40)",   "border": "#0070c0",  "text": "#60b0ff"},
    "EDF-B": {"bg": "rgba(0,112,192,0.40)",   "border": "#0070c0",  "text": "#60b0ff"},
    "EDF-C": {"bg": "rgba(0,112,192,0.40)",   "border": "#0070c0",  "text": "#60b0ff"},
    "COMMD": {"bg": "rgba(255,192,0,0.40)",   "border": "#ffc000",  "text": "#ffd060"},
    "EV-LO": {"bg": "rgba(100,220,60,0.35)",  "border": "#64dc3c",  "text": "#90ff70"},
    "VIS-M": {"bg": "rgba(180,100,255,0.35)", "border": "#b464ff",  "text": "#d4a0ff"},
    "FORE":  {"bg": "rgba(255,200,50,0.35)",  "border": "#ffc832",  "text": "#ffe080"},
    "FORP":  {"bg": "rgba(50,200,120,0.35)",  "border": "#32c878",  "text": "#80ffb0"},
}
DEFAULT_COLOR = {"bg": "rgba(255,255,255,0.20)", "border": "#888888", "text": "#cccccc"}

COLS = ["B", "C", "D", "E", "F", "G", "H"]

FRENCH_MONTHS = {
    1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
    9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre",
}

# ── Utilitaires ────────────────────────────────────────────────────────────


def first_name(name):
    """DE NOUEL Maxime -> Maxime, HEBERT Jean Baptiste -> Jean Baptiste"""
    parts = name.split()
    for i, p in enumerate(parts):
        if p != p.upper():
            return ' '.join(parts[i:])
    return parts[-1]


def slug(name):
    """BONILLO Matthieu -> bonillo-matthieu"""
    s = name.lower()
    for old, new in [("ï", "i"), ("é", "e"), ("è", "e"), ("ê", "e"),
                     ("ô", "o"), ("ü", "u"), ("ù", "u"), ("û", "u"),
                     ("à", "a"), ("â", "a"), ("ç", "c")]:
        s = s.replace(old, new)
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")


def week_dates(year, week):
    """Calcule les dates Lundi→Dimanche à partir de l'année/semaine ISO."""
    monday = datetime.fromisocalendar(year, week, 1)
    return {col: monday + timedelta(days=i) for i, col in enumerate(COLS)}


def format_date_range(year, week):
    """Retourne « 2 → 8 Mars » ou « 28 Février → 6 Mars »."""
    monday = datetime.fromisocalendar(year, week, 1)
    sunday = monday + timedelta(days=6)
    if monday.month == sunday.month:
        return f"{monday.day} \u2192 {sunday.day} {FRENCH_MONTHS[monday.month]}"
    return (f"{monday.day} {FRENCH_MONTHS[monday.month]} \u2192 "
            f"{sunday.day} {FRENCH_MONTHS[sunday.month]}")


def discover_excel_files(directory="."):
    """Trouve tous les fichiers « Plannings YYYY SXX.xlsx ».

    Si plusieurs versions existent pour la même semaine (ex: S14.xlsx et S14 v2.xlsx),
    seule la version la plus récente (v2 > v1 > sans version) est conservée.
    """
    pattern = re.compile(r"Plannings\s+(\d{4})\s+S(\d+)(?:\s+v(\d+))?\.xlsx", re.IGNORECASE)
    # Garder la meilleure version par (year, week)
    best = {}  # (year, week) -> {filename, year, week, version}
    for f in sorted(os.listdir(directory)):
        m = pattern.match(f)
        if m:
            year = int(m.group(1))
            week = int(m.group(2))
            version = int(m.group(3)) if m.group(3) else 0
            key = (year, week)
            if key not in best or version > best[key]["version"]:
                best[key] = {
                    "filename": os.path.join(directory, f) if directory != "." else f,
                    "year": year,
                    "week": week,
                    "version": version,
                }
    files = list(best.values())
    files.sort(key=lambda x: (x["year"], x["week"]))
    return files


# ── Parsing Excel ──────────────────────────────────────────────────────────


def get_cell(ws, row, col):
    return ws[f"{col}{row}"].value


def normalize_time_str(val):
    """Normalise une valeur de cellule horaire en chaîne « HH:MM/HH:MM[+] ».

    Gère :
    - str  « 08:00/10:00 », « 8:00/10:00 », « 19:00/00:30+ »
    - datetime (Excel formate la cellule en Heure) → converti en str
    - Retourne None si non reconnu.
    """
    if isinstance(val, datetime):
        # Excel time-formatted cell: openpyxl returns datetime(1900,1,1,H,M,S)
        # On ne peut extraire qu'une seule heure, pas un intervalle → warning
        return None
    if not isinstance(val, str):
        return None
    s = val.strip()
    if not s:
        return None
    # Accepter les heures à 1 ou 2 chiffres : « 8:00/10:00 » → « 08:00/10:00 »
    m = re.match(r"^(\d{1,2}):(\d{2})/(\d{1,2}):(\d{2})(\+?)$", s)
    if not m:
        return None
    return f"{int(m.group(1)):02d}:{m.group(2)}/{int(m.group(3)):02d}:{m.group(4)}{m.group(5)}"


def parse_time(time_str, base_date):
    """Parse « 08:00/10:00 » ou « 19:00/00:30+ » en (start_dt, end_dt).

    Le suffixe « + » indique explicitement que l'heure de fin est le
    lendemain (ex : CUP-R 19:00/00:30+  →  19h → 0h30 le jour suivant).
    Même sans « + », si end ≤ start le lendemain est détecté automatiquement.
    """
    parts = time_str.strip().split("/")
    if len(parts) != 2:
        return None

    start_str = parts[0].strip()
    end_str = parts[1].strip()

    next_day = end_str.endswith("+")
    if next_day:
        end_str = end_str.rstrip("+")

    sh, sm = int(start_str.split(":")[0]), int(start_str.split(":")[1])
    eh, em = int(end_str.split(":")[0]), int(end_str.split(":")[1])

    # Gérer 24:00 comme minuit du jour suivant
    start_extra = 0
    if sh >= 24:
        sh -= 24
        start_extra = 1
    end_extra = 0
    if eh >= 24:
        eh -= 24
        end_extra = 1

    start_dt = base_date.replace(hour=sh, minute=sm, second=0) + timedelta(days=start_extra)
    end_dt = base_date.replace(hour=eh, minute=em, second=0) + timedelta(days=end_extra)

    # « + » explicite OU détection automatique si fin ≤ début
    if next_day or (end_dt <= start_dt):
        end_dt += timedelta(days=1)

    return (start_dt, end_dt)


def parse_employees(ws, dates, week_num):
    """Parse tous les employés et leurs créneaux depuis la feuille Planning."""
    employees = {}
    current_name = None
    current_rows = []

    for row in range(5, ws.max_row + 1):
        name_cell = get_cell(ws, row, "A")
        if name_cell and isinstance(name_cell, str) and name_cell.strip():
            if current_name:
                employees[current_name] = parse_shifts(ws, current_rows, dates, week_num, current_name)
            current_name = name_cell.strip()
            current_rows = [row]
        elif current_name:
            current_rows.append(row)

    if current_name:
        employees[current_name] = parse_shifts(ws, current_rows, dates, week_num, current_name)

    return employees


def parse_shifts(ws, rows, dates, week_num, employee_name=""):
    """Parse les créneaux d'un employé à partir de ses lignes."""
    events = []
    warnings = []

    i = 0
    while i < len(rows):
        row = rows[i]
        has_codes = False
        codes = {}
        for col in COLS:
            val = get_cell(ws, row, col)
            if val and isinstance(val, str):
                val = val.strip()
                if val and not re.match(r"^\d{1,2}:\d{2}/\d{1,2}:\d{2}", val):
                    has_codes = True
                    codes[col] = val

        if has_codes:
            times = {}
            if i + 1 < len(rows):
                time_row = rows[i + 1]
                for col in COLS:
                    raw_val = get_cell(ws, time_row, col)
                    if raw_val is None:
                        continue
                    normalized = normalize_time_str(raw_val)
                    if normalized:
                        times[col] = normalized
                    elif isinstance(raw_val, datetime):
                        warnings.append(
                            f"  /!\\ {employee_name} ligne {time_row} col {col} : "
                            f"cellule format\u00e9e en Heure ({raw_val.strftime('%H:%M')}), "
                            f"convertir en texte dans Excel"
                        )
            else:
                for col, code in codes.items():
                    if col in dates:
                        warnings.append(
                            f"  /!\\ {employee_name} ligne {row} col {col} : "
                            f"code \u00ab {code} \u00bb sans ligne horaire en dessous"
                        )

            for col, code in codes.items():
                if col in dates:
                    if col in times:
                        parsed = parse_time(times[col], dates[col])
                        if parsed:
                            label = CODE_NAMES.get(code, code)
                            events.append({
                                "code": code,
                                "label": label,
                                "start": parsed[0],
                                "end": parsed[1],
                                "week": week_num,
                            })
                    else:
                        # Codes journée entière (MAL, CP, etc.) : pas d'horaire
                        FULL_DAY_CODES = {"MAL", "CP", "RTT", "ABS", "FOR-E", "FOR-P"}
                        if code in FULL_DAY_CODES:
                            label = CODE_NAMES.get(code, code)
                            day_date = dates[col]
                            events.append({
                                "code": code,
                                "label": label,
                                "start": day_date.replace(hour=0, minute=0),
                                "end": day_date.replace(hour=23, minute=59),
                                "week": week_num,
                                "all_day": True,
                            })
                        else:
                            warnings.append(
                                f"  /!\\ {employee_name} ligne {row} col {col} : "
                                f"code \u00ab {code} \u00bb sans horaire trouv\u00e9"
                            )

            i += 2
        else:
            i += 1

    for w in warnings:
        print(w)

    events.sort(key=lambda e: e["start"])

    # Résoudre les chevauchements : un staff ne peut pas avoir deux items en même temps.
    # Si deux events se chevauchent, on tronque la fin du premier au début du suivant.
    resolved = []
    for ev in events:
        if resolved:
            prev = resolved[-1]
            if ev["start"] < prev["end"]:
                print(
                    f"  /!\\ {employee_name} : chevauchement détecté entre "
                    f"«{prev['code']}» (fin {prev['end'].strftime('%H:%M')}) et "
                    f"«{ev['code']}» (début {ev['start'].strftime('%H:%M')}) — "
                    f"troncature de «{prev['code']}»"
                )
                prev["end"] = ev["start"]
                # Si le précédent a une durée nulle ou négative, le supprimer
                if prev["end"] <= prev["start"]:
                    resolved.pop()
        resolved.append(ev)

    return resolved


# ── Génération ICS (abonnement calendrier) ─────────────────────────────────


def generate_ics(name, events, week_notes=None):
    """Génère le contenu ICS pour un employé (toutes semaines confondues).

    Chaque fichier ICS contient TOUS les événements de l'employé, ce qui
    permet à l'abonnement calendrier de rester à jour automatiquement.
    week_notes: dict {week_num: notes_data} pour ajouter les commentaires.
    """
    if week_notes is None:
        week_notes = {}
    s = slug(name)
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Planning Urban 7D//FR",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:Planning {name}",
        "X-WR-TIMEZONE:Europe/Paris",
        # Intervalle de rafraîchissement pour les clients calendrier
        "REFRESH-INTERVAL;VALUE=DURATION:PT12H",
        "X-PUBLISHED-TTL:PT12H",
        "BEGIN:VTIMEZONE",
        "TZID:Europe/Paris",
        "BEGIN:STANDARD",
        "TZOFFSETFROM:+0200",
        "TZOFFSETTO:+0100",
        "TZNAME:CET",
        "DTSTART:19701025T030000",
        "RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=-1SU",
        "END:STANDARD",
        "BEGIN:DAYLIGHT",
        "TZOFFSETFROM:+0100",
        "TZOFFSETTO:+0200",
        "TZNAME:CEST",
        "DTSTART:19700329T020000",
        "RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=-1SU",
        "END:DAYLIGHT",
        "END:VTIMEZONE",
    ]

    # Grouper par semaine pour des UIDs stables
    by_week = {}
    for evt in events:
        w = evt["week"]
        if w not in by_week:
            by_week[w] = []
        by_week[w].append(evt)

    # DTSTAMP must be UTC per RFC 5545
    from datetime import datetime as _dt, timezone as _tz
    dtstamp_utc = _dt.now(_tz.utc).strftime("%Y%m%dT%H%M%SZ")

    for week_num in sorted(by_week.keys()):
        # Build description with weekly notes if available
        wn = week_notes.get(week_num, {})
        week_comment = wn.get("comment", "")
        week_updates = wn.get("updates", [])
        extra_desc = ""
        if week_comment:
            extra_desc += week_comment
        for upd in week_updates:
            upd_text = upd.get("text", "")
            upd_date = upd.get("date", "")
            if upd_text:
                prefix = f"MAJ {upd_date}: " if upd_date else "MAJ: "
                if extra_desc:
                    extra_desc += "\n"
                extra_desc += prefix + upd_text

        # Build replacement lookup for this week
        week_repls = wn.get("replacements", [])

        for i, evt in enumerate(by_week[week_num], 1):
            dt_start = evt["start"].strftime("%Y%m%dT%H%M%S")
            dt_end = evt["end"].strftime("%Y%m%dT%H%M%S")
            evt_date = evt["start"].strftime("%Y-%m-%d")
            evt_sh = evt["start"].hour + evt["start"].minute / 60
            evt_eh = evt["end"].hour + evt["end"].minute / 60
            if evt_eh <= evt_sh:
                evt_eh = 24

            # Check if this event is affected by a replacement
            summary = evt['label']
            repl_note = ""
            for r in week_repls:
                if r.get("date") != evt_date:
                    continue
                r_parts = r.get("start", "0:0").split(":")
                r_start = int(r_parts[0]) + int(r_parts[1] if len(r_parts) > 1 else 0) / 60
                r_parts = r.get("end", "0:0").split(":")
                r_end = int(r_parts[0]) + int(r_parts[1] if len(r_parts) > 1 else 0) / 60
                if evt_sh < r_end and evt_eh > r_start:
                    if name == r.get("out"):
                        # Get first name of replacer
                        in_name = r.get("in", "")
                        in_first = in_name.split()[-1] if in_name else ""
                        summary = f"[Remplacé par {in_first}] " + summary
                        repl_note = f"Remplacé par {in_name}"
                    elif name == r.get("in"):
                        out_name = r.get("out", "")
                        out_first = out_name.split()[-1] if out_name else ""
                        summary = f"[Remplace {out_first}] " + summary
                        repl_note = f"Remplace {out_name}"

            # Escape for ICS
            desc = extra_desc
            if repl_note:
                desc = repl_note + ("\n" + desc if desc else "")
            desc_escaped = desc.replace("\\", "\\\\").replace("\n", "\\n").replace(",", "\\,").replace(";", "\\;")
            summary_escaped = summary.replace(chr(92), chr(92)+chr(92)).replace(',', chr(92)+',').replace(';', chr(92)+';')
            vevent = [
                "BEGIN:VEVENT",
                f"UID:{s}-s{week_num}-{i}@urban7d",
                f"DTSTAMP:{dtstamp_utc}",
                f"DTSTART;TZID=Europe/Paris:{dt_start}",
                f"DTEND;TZID=Europe/Paris:{dt_end}",
                f"SUMMARY:{summary_escaped}",
            ]
            if desc_escaped:
                vevent.append(f"DESCRIPTION:{desc_escaped}")
            vevent.append("END:VEVENT")
            lines.extend(vevent)

    lines.append("END:VCALENDAR")
    # RFC 5545 §3.1: content lines MUST NOT exceed 75 octets – fold long lines
    folded = []
    for line in lines:
        encoded = line.encode("utf-8")
        if len(encoded) <= 75:
            folded.append(line)
        else:
            # First chunk: max 75 octets, continuations: space + max 74 octets
            chunks = []
            while len(encoded) > 75:
                # Find a safe cut point (don't split multi-byte UTF-8 chars)
                cut = 75 if not chunks else 74
                pos = cut
                while pos > 0 and (encoded[pos] & 0xC0) == 0x80:
                    pos -= 1
                if pos == 0:
                    pos = cut  # fallback
                if chunks:
                    chunks.append(" " + encoded[:pos].decode("utf-8", errors="replace"))
                else:
                    chunks.append(encoded[:pos].decode("utf-8", errors="replace"))
                encoded = encoded[pos:]
            if encoded:
                rest = encoded.decode("utf-8", errors="replace")
                chunks.append((" " + rest) if chunks else rest)
            folded.extend(chunks)
    return "\r\n".join(folded) + "\r\n"


# ── Génération HTML ────────────────────────────────────────────────────────


def build_events_json(week_employees):
    """Construit les données JSON des événements pour injection dans le HTML."""
    data = {}
    for name, evts in week_employees.items():
        data[name] = {
            "slug": slug(name),
            "events": [{
                "code": e["code"],
                "label": e["label"],
                "start": e["start"].strftime("%Y-%m-%dT%H:%M"),
                "end": e["end"].strftime("%Y-%m-%dT%H:%M"),
                "day": e["start"].weekday(),
            } for e in evts],
        }
    return json.dumps(data, ensure_ascii=False)


def load_week_notes(week_num):
    """Charge les notes de semaine depuis notes/SXX.json."""
    path = f"notes/S{week_num}.json"
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"comment": "", "updates": []}


def generate_html(week_employees, week_num, year, all_weeks, excel_version=0):
    """Génère la page HTML avec preview timeline + vue individuelle + abonnement."""
    date_range = format_date_range(year, week_num)

    # Toujours utiliser les données fraîchement parsées de l'Excel
    events_json = build_events_json(week_employees)

    # Lire les métadonnées _meta depuis le JSON (si présent)
    try:
        parsed = json.loads(events_json)
        meta = parsed.get("_meta", {})
    except (json.JSONDecodeError, AttributeError):
        meta = {}
    meta_json = json.dumps(meta, ensure_ascii=False)
    colors_json = json.dumps(CODE_COLORS, ensure_ascii=False)
    default_color_json = json.dumps(DEFAULT_COLOR, ensure_ascii=False)
    notes_data = load_week_notes(week_num)
    notes_json = json.dumps(notes_data, ensure_ascii=False)

    DAYS_SHORT = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    DAYS_FULL = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    monday = datetime.fromisocalendar(year, week_num, 1)
    day_labels_json = json.dumps([
        f"{DAYS_SHORT[i]} {(monday + timedelta(days=i)).day:02d}"
        for i in range(7)
    ], ensure_ascii=False)
    day_labels_full_json = json.dumps([
        f"{DAYS_FULL[i]} {(monday + timedelta(days=i)).day:02d}/{(monday + timedelta(days=i)).month:02d}"
        for i in range(7)
    ], ensure_ascii=False)
    week_dates_json = json.dumps([
        (monday + timedelta(days=i)).strftime('%Y-%m-%d')
        for i in range(7)
    ])

    week_tabs = ""
    for w in sorted(all_weeks):
        cls = ' active' if w == week_num else ''
        href = '#' if w == week_num else f'S{w}.html'
        w_sun = datetime.fromisocalendar(year, w, 7)
        data_end = w_sun.strftime('%Y-%m-%d')
        week_tabs += f'            <a href="{href}" class="week-tab{cls}" data-end="{data_end}">S{w}</a>\n'

    employee_buttons = ""
    for name in week_employees:
        s = slug(name)
        has_events = len(week_employees[name]) > 0
        if has_events:
            employee_buttons += (
                f'            <button class="employee-btn" data-name="{name}" '
                f'data-slug="{s}">{name}</button>\n'
            )
        else:
            employee_buttons += (
                f'            <div class="employee-btn repos">{name} '
                f'<span class="badge">Repos</span></div>\n'
            )

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="Planning U7D">
    <link rel="manifest" href="manifest.json">
    <meta name="theme-color" content="#FF6600">
    <title>Planning Urban 7D - S{week_num}</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Montserrat', sans-serif;
            background: #1A1A1A;
            min-height: 100vh;
            padding: 15px;
            color: #fff;
            position: relative;
        }}
        body::before {{
            content: '';
            position: fixed;
            inset: 0;
            z-index: 0;
            pointer-events: none;
            background: url('bg-team.jpg') center center / cover no-repeat;
            opacity: 0.22;
        }}
        .container {{ position: relative; z-index: 1; max-width: 600px; margin: 0 auto;
                      background: rgba(26,26,26,0.95); border-radius: 8px;
                      padding: 2px 6px; margin-top: 6px; margin-bottom: 6px;
                      border-top: 4px solid #FF6600; overflow: hidden; }}
        .container::after {{ content: '\\276F\\276F\\276F\\276F'; position: absolute;
                             top: 12px; right: -2px; font-size: 28px; font-weight: 900;
                             color: rgba(255,102,0,0.08); letter-spacing: -4px;
                             pointer-events: none; z-index: 0; }}

        /* ── Top bar (Admin + PSG Academy) ── */
        .top-bar {{ display: flex; justify-content: space-between; align-items: center;
                     padding: 8px 4px 0; }}
        .top-bar .admin-btn, .top-bar .psg-btn {{
            padding: 10px 22px; border-radius: 8px; font-size: 14px; font-weight: 700;
            cursor: pointer; text-decoration: none; transition: all 0.25s;
            font-family: 'Montserrat', sans-serif; text-transform: uppercase;
            letter-spacing: 0.5px; border: 1px solid rgba(255,255,255,0.12);
            background: none; color: #666;
        }}
        .top-bar .admin-btn:hover {{ color: #FF6600; border-color: rgba(255,102,0,0.4); }}
        .top-bar .admin-btn.unlocked {{ color: #FF6600; border-color: rgba(255,102,0,0.4); }}
        .top-bar .psg-btn {{ background: #0C1C3E; color: #E30613;
                              border-color: rgba(227,6,19,0.4); }}
        .top-bar .psg-btn:hover {{ background: rgba(227,6,19,0.15); color: #ff3040;
                                    border-color: #E30613; }}

        /* ── Header ── */
        .header {{ text-align: center; margin-bottom: 12px; padding: 10px 10px 8px; }}
        h1 {{ font-family: 'Montserrat', sans-serif;
              color: #fff; font-size: 20px; font-weight: 900; margin-bottom: 2px;
              text-transform: uppercase; letter-spacing: 2px;
              text-shadow: 0 0 30px rgba(255,102,0,0.3); }}
        .subtitle {{ color: #FF6600; font-size: 11px; font-weight: 700; text-transform: uppercase;
                     letter-spacing: 3px; }}
        .dates {{ color: #FF6600; font-size: 14px; font-weight: 700;
                  background: rgba(255,102,0,0.1); padding: 6px 14px;
                  border-radius: 6px; display: inline-block; margin-top: 6px;
                  border: 1px solid rgba(255,102,0,0.2); }}

        /* ── Week selector ── */
        .week-selector {{ display: flex; justify-content: center; gap: 6px;
                          margin-bottom: 15px; flex-wrap: wrap; }}
        .week-tab {{ padding: 8px 14px; background: rgba(255,255,255,0.04);
                     border: 1px solid rgba(255,255,255,0.08); border-radius: 6px;
                     color: #666; text-decoration: none; font-weight: 600; font-size: 13px;
                     transition: all 0.2s; text-transform: uppercase; letter-spacing: 0.5px; }}
        .week-tab:hover {{ background: rgba(255,102,0,0.1); border-color: rgba(255,102,0,0.3); color: #FF6600; }}
        .week-tab.active {{ background: #FF6600; border-color: #FF6600; color: white;
                            box-shadow: 0 0 15px rgba(255,102,0,0.4); }}
        .week-tab.past {{ opacity: 0.4; }}
        .week-tab.past:hover {{ opacity: 0.7; }}


        /* ── View toggle ── */
        .view-toggle {{ display: flex; justify-content: center; gap: 4px; margin-bottom: 15px;
                        background: rgba(255,255,255,0.04); border-radius: 6px; padding: 4px; }}
        .view-btn {{ flex: 1; padding: 8px; border: none; background: transparent;
                     color: #666; font-size: 12px; font-weight: 700; cursor: pointer;
                     border-radius: 4px; transition: all 0.2s; font-family: inherit;
                     text-transform: uppercase; letter-spacing: 0.5px; }}
        .view-btn.active {{ background: rgba(255,102,0,0.15); color: #FF6600;
                            box-shadow: 0 0 10px rgba(255,102,0,0.2); }}

        /* ── Day tabs ── */
        .day-tabs {{ display: flex; gap: 3px; margin-bottom: 12px; overflow-x: auto;
                     padding-bottom: 4px; -webkit-overflow-scrolling: touch;
                     scrollbar-width: none; }}
        .day-tabs::-webkit-scrollbar {{ display: none; }}
        .day-tab {{ padding: 6px 8px; background: rgba(255,255,255,0.04);
                    border: 1px solid rgba(255,255,255,0.08); border-radius: 4px;
                    color: #666; font-size: 10px; font-weight: 700; cursor: pointer;
                    white-space: nowrap; transition: all 0.2s; flex: 1; min-width: 0;
                    text-align: center; text-transform: uppercase; }}
        .day-tab.active {{ background: rgba(255,102,0,0.15); border-color: rgba(255,102,0,0.3);
                           color: #FF6600; }}

        /* ── Timeline (vue Journée) ── */
        .timeline {{ position: relative; margin-bottom: 20px;
                     overflow-x: auto; -webkit-overflow-scrolling: touch; }}
        /* ── Scrollbar orange néon ── */
        .timeline::-webkit-scrollbar {{ height: 6px; }}
        .timeline::-webkit-scrollbar-track {{ background: rgba(255,255,255,0.04); border-radius: 3px; }}
        .timeline::-webkit-scrollbar-thumb {{ background: #FF6600; border-radius: 2px;
                                              box-shadow: 0 0 8px rgba(255,102,0,0.6); }}
        .timeline {{ scrollbar-width: thin; scrollbar-color: #FF6600 rgba(255,255,255,0.04); }}
        .timeline-inner {{ min-width: 500px; }}
        .time-markers {{ display: flex; justify-content: space-between; padding: 0 0 6px 0;
                         border-bottom: 1px solid rgba(255,255,255,0.06); margin-bottom: 8px; }}
        .time-marker {{ font-size: 9px; color: #555; font-weight: 500; }}
        .timeline-row {{ display: flex; align-items: center; margin-bottom: 4px; }}
        .tl-name {{ width: 70px; font-size: 10px; color: #aaa; font-weight: 500;
                    flex-shrink: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
                    padding-right: 6px; cursor: pointer; transition: color 0.2s;
                    position: sticky; left: 0; z-index: 2;
                    background: linear-gradient(90deg, rgba(26,26,26,0.98) 80%, transparent);
                    padding-right: 10px; }}
        .tl-name:hover {{ color: #FF6600; }}
        .tl-bar-container {{ flex: 1; position: relative; height: 26px;
                             background: rgba(255,255,255,0.02); border-radius: 4px; }}
        .tl-grid-line {{ position: absolute; top: 0; bottom: 0; width: 1px; pointer-events: none; z-index: 0; }}
        .tl-grid-line.hour {{ background: rgba(255,255,255,0.10); }}
        .tl-grid-line.half {{ background: rgba(255,255,255,0.05); border-left: 1px dashed rgba(255,255,255,0.08); width: 0; }}
        @keyframes nowPulse {{
            0%, 100% {{ filter: drop-shadow(0 0 4px #ffd700) drop-shadow(0 0 8px rgba(255,215,0,0.4)); opacity: 0.7; }}
            50% {{ filter: drop-shadow(0 0 10px #ffd700) drop-shadow(0 0 20px rgba(255,215,0,0.8)); opacity: 1; }}
        }}
        .tl-now-line {{ position: absolute; top: 0; bottom: 0; width: 2px; pointer-events: none; z-index: 3;
                        border-left: 2px dashed #ffd700;
                        filter: drop-shadow(0 0 4px #ffd700) drop-shadow(0 0 8px rgba(255,215,0,0.4)); opacity: 0.8; }}
        .tl-now-marker {{ position: absolute; top: 0; bottom: 0; width: 2px; pointer-events: none; z-index: 3;
                          border-left: 2px dashed #ffd700;
                          filter: drop-shadow(0 0 4px #ffd700) drop-shadow(0 0 8px rgba(255,215,0,0.4)); opacity: 0.8; }}
        .tl-bar {{ position: absolute; height: 100%; border-radius: 4px;
                   display: flex; align-items: center; justify-content: center;
                   font-size: 9px; font-weight: 600; overflow: hidden;
                   border-left: 2px solid; transition: all 0.2s;
                   cursor: default; }}
        .tl-bar:hover {{ filter: brightness(1.3); z-index: 2;
                         box-shadow: 0 0 12px var(--glow-color); }}
        .tl-bar .bar-label {{ padding: 0 4px; white-space: nowrap; }}
        .tl-bar.replaced {{ position: relative; opacity: 0.7; }}
        .tl-bar.replaced::after {{ content: ''; position: absolute; inset: 0; border-radius: inherit;
            background: repeating-linear-gradient(45deg, transparent, transparent 3px, rgba(255,60,60,0.35) 3px, rgba(255,60,60,0.35) 5px);
            pointer-events: none; }}
        .tl-bar.replacer {{ position: relative; }}
        .tl-bar.replacer::after {{ content: ''; position: absolute; inset: 0; border-radius: inherit;
            background: repeating-linear-gradient(45deg, transparent, transparent 3px, rgba(60,220,80,0.35) 3px, rgba(60,220,80,0.35) 5px);
            pointer-events: none; }}

        /* ── Employee list (vue Staff) ── */
        .employee-list {{ display: flex; flex-direction: column; gap: 6px; margin-bottom: 15px; }}
        .employee-btn {{ display: flex; align-items: center; justify-content: space-between;
                         padding: 12px 14px; background: rgba(255,255,255,0.04);
                         border-radius: 6px; color: white; font-weight: 600; font-size: 13px;
                         border: 1px solid rgba(255,255,255,0.08); cursor: pointer;
                         transition: all 0.2s; font-family: inherit; width: 100%; text-align: left; }}
        .employee-btn:hover {{ background: rgba(255,102,0,0.1); border-color: rgba(255,102,0,0.3);
                               transform: translateX(4px); }}
        .employee-btn.repos {{ color: #444; cursor: default; pointer-events: none; }}
        .badge {{ font-size: 10px; padding: 3px 8px; background: rgba(255,255,255,0.06);
                  border-radius: 4px; color: #444; font-weight: 600; text-transform: uppercase; }}
        .hours-badge {{ background: rgba(255,102,0,0.15); color: #FF6600; font-weight: 700; }}

        /* ── Individual preview (modal) ── */
        .modal-overlay {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.85);
                          z-index: 100; justify-content: center; align-items: flex-start;
                          padding: 15px 10px; overflow-y: auto; }}
        .modal-overlay.open {{ display: flex; }}
        .modal {{ background: #1f1f1f; border-radius: 8px; width: 100%; max-width: 500px;
                  border: 1px solid rgba(255,255,255,0.08); overflow: hidden; }}
        .modal-header {{ padding: 14px 16px; display: flex; justify-content: space-between;
                         align-items: center; border-bottom: 1px solid rgba(255,255,255,0.06); }}
        .modal-header h2 {{ font-size: 16px; color: #FF6600; font-weight: 700; }}
        .modal-close {{ background: none; border: none; color: #666; font-size: 24px;
                        cursor: pointer; padding: 0 5px; line-height: 1; }}
        .modal-close:hover {{ color: #fff; }}
        .modal-body {{ padding: 12px 14px; }}
        .modal-day {{ margin-bottom: 12px; }}
        .modal-day-title {{ font-size: 11px; color: #666; font-weight: 600;
                            text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 5px; }}
        .modal-event {{ display: flex; align-items: center; gap: 8px; padding: 8px 10px;
                        border-radius: 4px; margin-bottom: 3px; border-left: 3px solid; }}
        .modal-event .ev-time {{ font-size: 11px; font-weight: 600; white-space: nowrap;
                                 min-width: 80px; }}
        .modal-event .ev-label {{ font-size: 12px; font-weight: 500; }}
        .modal-event .ev-repl {{ font-size: 10px; font-weight: 600; margin-left: auto; white-space: nowrap; }}
        .modal-event.replaced {{ position: relative; opacity: 0.7; }}
        .modal-event.replaced::after {{ content: ''; position: absolute; inset: 0; border-radius: inherit;
            background: repeating-linear-gradient(45deg, transparent, transparent 3px,
            rgba(255,60,60,0.25) 3px, rgba(255,60,60,0.25) 5px); pointer-events: none; }}
        .modal-event.replacer {{ position: relative; }}
        .modal-event.replacer::after {{ content: ''; position: absolute; inset: 0; border-radius: inherit;
            background: repeating-linear-gradient(45deg, transparent, transparent 3px,
            rgba(60,220,80,0.25) 3px, rgba(60,220,80,0.25) 5px); pointer-events: none; }}
        .modal-footer {{ padding: 12px 16px; border-top: 1px solid rgba(255,255,255,0.06);
                         text-align: center; }}
        .modal-hours-total {{ margin-top: 12px; padding: 10px 14px; text-align: right;
                              font-size: 13px; color: #FF6600; font-weight: 500;
                              border-top: 1px solid rgba(255,255,255,0.06); }}
        .hours-line {{ padding: 2px 0; }}
        .hours-line.pause {{ color: #888; font-size: 11px; }}
        .hours-line.net {{ color: #64dc3c; font-size: 14px; margin-top: 4px;
                           padding-top: 6px; border-top: 1px solid rgba(255,255,255,0.06); }}
        .hours-brut {{ color: #666; font-weight: 400; font-size: 9px; }}
        .subscribe-btn {{ display: inline-flex; align-items: center; gap: 8px;
                          padding: 10px 24px; background: #FF6600; color: white;
                          border: none; border-radius: 6px; font-size: 13px; font-weight: 700;
                          cursor: pointer; font-family: inherit; transition: all 0.2s;
                          text-decoration: none; text-transform: uppercase; letter-spacing: 0.5px;
                          box-shadow: 0 0 20px rgba(255,102,0,0.3); }}
        .subscribe-btn:hover {{ background: #ff9050;
                                box-shadow: 0 0 30px rgba(255,102,0,0.5); transform: scale(1.02); }}

        /* ── Calendar chooser (bottom sheet) ── */
        .cal-chooser-overlay {{ display:none; position:fixed; inset:0; background:rgba(0,0,0,0.85);
                                z-index:200; justify-content:center; align-items:flex-end; }}
        .cal-chooser-overlay.open {{ display:flex; }}
        .cal-chooser {{ background:#1f1f1f; border-radius:8px 8px 0 0; width:100%; max-width:500px;
                        padding:20px 16px 30px; animation: slideUp 0.25s ease-out; }}
        @keyframes slideUp {{ from {{ transform:translateY(100%); }} to {{ transform:translateY(0); }} }}
        .cal-chooser h3 {{ color:#FF6600; font-size:15px; font-weight:700; margin-bottom:4px; text-align:center; }}
        .cal-chooser .cal-sub {{ color:#666; font-size:11px; text-align:center; margin-bottom:16px; }}
        .cal-option {{ display:flex; align-items:center; gap:12px; padding:14px;
                       background:rgba(255,255,255,0.04); border:1px solid rgba(255,255,255,0.08);
                       border-radius:6px; margin-bottom:8px; text-decoration:none; color:white;
                       transition:all 0.2s; cursor:pointer; }}
        .cal-option:hover {{ background:rgba(255,102,0,0.1); border-color:rgba(255,102,0,0.3); }}
        .cal-option .cal-icon {{ font-size:22px; width:36px; text-align:center; flex-shrink:0; }}
        .cal-option .cal-info {{ flex:1; }}
        .cal-option .cal-name {{ font-weight:600; font-size:13px; }}
        .cal-option .cal-desc {{ font-size:10px; color:#888; margin-top:2px; }}
        .cal-chooser-cancel {{ display:block; width:100%; padding:12px; background:none;
                               border:1px solid rgba(255,255,255,0.1); border-radius:6px;
                               color:#888; font-size:13px; cursor:pointer; margin-top:4px;
                               font-family:inherit; transition: all 0.2s; }}
        .cal-chooser-cancel:hover {{ color:#fff; border-color:rgba(255,255,255,0.3); }}
        .google-steps {{ padding: 0 4px; }}
        .step-url {{ background: rgba(0,0,0,0.3); border: 1px solid rgba(255,255,255,0.1);
                     border-radius: 4px; padding: 10px 12px; font-size: 10px; color: #FF6600;
                     word-break: break-all; margin-bottom: 10px; font-family: monospace; }}
        .copy-url-btn {{ width: 100%; padding: 10px; background: #FF6600; color: #fff; border: none;
                         border-radius: 6px; font-size: 13px; font-weight: 600; cursor: pointer;
                         font-family: inherit; transition: all 0.2s; }}
        .copy-url-btn.copied {{ background: #2a6e2a; color: #64dc3c; }}
        .step-divider {{ height: 1px; background: rgba(255,255,255,0.08); margin: 14px 0; }}
        .step-title {{ font-size: 12px; color: #aaa; font-weight: 600; margin-bottom: 6px; }}
        .step-list {{ font-size: 12px; color: #ccc; padding-left: 20px; margin: 0; }}
        .step-list li {{ margin-bottom: 6px; }}
        .step-link {{ color: #4a9eff; text-decoration: none; }}
        .step-link:hover {{ text-decoration: underline; }}
        .step-note {{ font-size: 10px; color: #64dc3c; margin-top: 10px; text-align: center; }}
        .email-btn {{ display: flex; align-items: center; justify-content: center; gap: 8px;
                      width: 100%; padding: 10px; background: rgba(255,255,255,0.06);
                      border: 1px solid rgba(255,255,255,0.12); border-radius: 6px;
                      font-size: 12px; color: #ccc; cursor: pointer; font-family: inherit;
                      text-decoration: none; margin-top: 8px; transition: all 0.2s; }}
        .email-btn:hover {{ background: rgba(255,255,255,0.1); color: #fff; }}

        .no-events {{ text-align: center; padding: 30px; color: #444; font-size: 13px; }}

        /* ── Notes de semaine ── */
        .week-notes {{ margin-bottom: 15px; }}
        .note-card {{ background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08);
                      border-radius: 6px; padding: 12px 14px; margin-bottom: 8px; }}
        .note-card.comment {{ border-left: 3px solid #FF6600; }}
        .note-card.update {{ border-left: 3px solid #ffc000; }}
        .note-header {{ display: flex; justify-content: space-between; align-items: center;
                        margin-bottom: 6px; }}
        .note-label {{ font-size: 10px; font-weight: 700; text-transform: uppercase;
                       letter-spacing: 0.5px; }}
        .note-label.comment {{ color: #FF6600; }}
        .note-label.update {{ color: #ffc000; }}
        .note-text {{ font-size: 12px; color: #ccc; line-height: 1.6; white-space: pre-line; }}
        .note-text:empty::before {{ content: 'Cliquer pour ajouter...'; color: #444; font-style: italic; }}
        .note-text[contenteditable=true] {{ outline: none; border: 1px solid rgba(255,102,0,0.2);
                                            border-radius: 6px; padding: 8px; min-height: 40px;
                                            background: rgba(0,0,0,0.2); }}
        .note-actions {{ display: flex; gap: 6px; }}
        .note-btn {{ background: none; border: none; color: #555; font-size: 14px;
                     cursor: pointer; padding: 2px 4px; transition: color 0.2s; }}
        .note-btn:hover {{ color: #FF6600; }}
        .note-btn.del:hover {{ color: #ff5050; }}
        .add-note-btn {{ display: flex; align-items: center; justify-content: center; gap: 6px;
                         padding: 8px; background: rgba(255,255,255,0.02);
                         border: 1px dashed rgba(255,255,255,0.1); border-radius: 6px;
                         color: #444; font-size: 11px; cursor: pointer; transition: all 0.2s;
                         font-family: inherit; width: 100%; margin-bottom: 8px; }}
        .add-note-btn:hover {{ border-color: rgba(255,102,0,0.3); color: #FF6600; }}
        .publish-btn {{ display: flex; align-items: center; justify-content: center; gap: 6px;
                        padding: 10px 16px; background: #FF6600; border: none; border-radius: 6px;
                        color: #fff; font-size: 12px; font-weight: 600; cursor: pointer;
                        transition: all 0.2s; font-family: inherit; width: 100%; margin-top: 8px;
                        box-shadow: 0 0 15px rgba(255,102,0,0.3); }}
        .publish-btn:hover {{ background: #ff9050; box-shadow: 0 0 25px rgba(255,102,0,0.5); }}
        .publish-btn:disabled {{ background: #444; box-shadow: none; cursor: not-allowed; color: #888; }}
        .publish-btn.success {{ background: #64dc3c; box-shadow: 0 0 15px rgba(100,220,60,0.3); }}
        .admin-setup {{ display: flex; align-items: center; gap: 6px; margin-top: 8px; }}
        .admin-input {{ flex: 1; padding: 8px 10px; background: rgba(0,0,0,0.3);
                        border: 1px solid rgba(255,255,255,0.1); border-radius: 4px;
                        color: #ccc; font-size: 11px; font-family: inherit; outline: none; }}
        .admin-input:focus {{ border-color: rgba(255,102,0,0.4); }}
        .admin-input::placeholder {{ color: #444; }}
        .admin-save-btn {{ padding: 8px 12px; background: rgba(255,102,0,0.15);
                           border: 1px solid rgba(255,102,0,0.3); border-radius: 4px;
                           color: #FF6600; font-size: 11px; cursor: pointer; font-family: inherit;
                           white-space: nowrap; }}
        .admin-hint {{ font-size: 10px; color: #444; margin-top: 4px; }}

        /* ── Remplacements ── */
        .note-card.replacement {{ border-left: 3px solid #ff5050; }}
        .note-label.replacement {{ color: #ff5050; }}
        .repl-summary {{ font-size: 12px; color: #ccc; line-height: 1.6; }}
        .repl-out {{ color: #ff6b6b; font-weight: 600; }}
        .repl-in {{ color: #64dc3c; font-weight: 600; }}
        .repl-form {{ display: flex; flex-direction: column; gap: 8px; margin-top: 8px; }}
        .repl-form select, .repl-form input {{ padding: 7px 10px; background: rgba(0,0,0,0.3);
            border: 1px solid rgba(255,255,255,0.1); border-radius: 4px;
            color: #ccc; font-size: 11px; font-family: inherit; outline: none; }}
        .repl-form select:focus, .repl-form input:focus {{ border-color: rgba(255,102,0,0.4); }}
        .repl-form select option {{ background: #1f1f1f; color: #ccc; }}
        .repl-row {{ display: flex; gap: 6px; align-items: center; }}
        .repl-row label {{ font-size: 10px; color: #666; min-width: 55px; text-align: right; }}
        .repl-row select, .repl-row input {{ flex: 1; }}
        .repl-add-btn {{ padding: 7px 14px; background: rgba(255,80,80,0.15);
            border: 1px solid rgba(255,80,80,0.3); border-radius: 4px;
            color: #ff5050; font-size: 11px; cursor: pointer; font-family: inherit;
            transition: all 0.2s; align-self: flex-end; }}
        .repl-add-btn:hover {{ background: rgba(255,80,80,0.25); }}
        .add-repl-btn {{ display: flex; align-items: center; justify-content: center; gap: 6px;
                         padding: 8px; background: rgba(255,80,80,0.02);
                         border: 1px dashed rgba(255,80,80,0.15); border-radius: 6px;
                         color: #ff5050; font-size: 11px; cursor: pointer; transition: all 0.2s;
                         font-family: inherit; width: 100%; margin-bottom: 8px; opacity: 0.6; }}
        .add-repl-btn:hover {{ border-color: rgba(255,80,80,0.4); opacity: 1; }}

        /* ── Admin edit mode ── */
        .admin-toolbar {{ display: flex; align-items: center; gap: 8px; margin-bottom: 12px;
                          padding: 8px 12px; background: rgba(255,102,0,0.08);
                          border: 1px solid rgba(255,102,0,0.2); border-radius: 10px; }}
        .edit-toggle {{ padding: 6px 14px; background: rgba(255,102,0,0.15);
                        border: 1px solid rgba(255,102,0,0.3); border-radius: 8px;
                        color: #FF6600; font-size: 11px; font-weight: 600; cursor: pointer;
                        font-family: inherit; transition: all 0.2s; }}
        .edit-toggle:hover {{ background: #FF6600; color: #fff; }}
        .edit-toggle.active {{ background: #FF6600; color: #fff;
                               box-shadow: 0 0 10px rgba(255,102,0,0.4); }}
        .tl-bar.editable {{ cursor: pointer; }}
        .tl-bar.editable:hover {{ outline: 2px solid #FF6600; outline-offset: 1px; }}
        .edit-popup {{ position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%);
                       z-index: 200; background: #1a1a2e; border: 1px solid rgba(255,102,0,0.3);
                       border-radius: 12px; padding: 16px; min-width: 260px;
                       box-shadow: 0 10px 40px rgba(0,0,0,0.6); }}
        .edit-popup h3 {{ font-size: 13px; color: #FF6600; margin-bottom: 10px; }}
        .edit-popup .field {{ display: flex; align-items: center; gap: 8px; margin-bottom: 8px; }}
        .edit-popup .field label {{ font-size: 11px; color: #888; min-width: 45px; }}
        .edit-popup .field input, .edit-popup .field input[type="text"] {{
            flex: 1; padding: 6px 8px; background: rgba(0,0,0,0.3);
            border: 1px solid rgba(255,255,255,0.1); border-radius: 6px;
            color: #fff; font-size: 12px; font-family: inherit; outline: none; }}
        .edit-popup .field input:focus {{ border-color: rgba(255,102,0,0.4); }}
        .edit-popup .actions {{ display: flex; gap: 6px; margin-top: 10px; }}
        .edit-popup .actions button {{ flex: 1; padding: 8px; border: none; border-radius: 8px;
                                       font-size: 11px; font-weight: 600; cursor: pointer;
                                       font-family: inherit; transition: all 0.2s; }}
        .edit-popup .btn-save {{ background: #FF6600; color: #fff; }}
        .edit-popup .btn-save:hover {{ background: #ff9050; }}
        .edit-popup .btn-cancel {{ background: rgba(255,255,255,0.08); color: #888; }}
        .edit-popup .btn-cancel:hover {{ background: rgba(255,255,255,0.15); color: #fff; }}
        .edit-popup .btn-delete {{ background: rgba(220,50,50,0.15); color: #e55; flex: 0.6; }}
        .edit-popup .btn-delete:hover {{ background: rgba(220,50,50,0.3); color: #ff6666; }}
        .edit-popup select {{ flex: 1; padding: 6px 8px; background: rgba(0,0,0,0.3);
                              border: 1px solid rgba(255,255,255,0.1); border-radius: 6px;
                              color: #fff; font-size: 12px; font-family: inherit; outline: none; }}
        .edit-popup select:focus {{ border-color: rgba(255,102,0,0.4); }}
        .day-check {{ display: flex; align-items: center; gap: 8px; padding: 5px 8px;
                      border-radius: 6px; cursor: pointer; font-size: 12px; color: #ccc;
                      transition: background 0.15s; }}
        .day-check:hover {{ background: rgba(255,255,255,0.05); }}
        .day-check em {{ color: #666; font-style: normal; font-size: 10px; }}
        .day-check input[type="checkbox"] {{ accent-color: #FF6600; width: 16px; height: 16px; cursor: pointer; }}
        .day-check.all-check {{ border-bottom: 1px solid rgba(255,255,255,0.06);
                                padding-bottom: 8px; margin-bottom: 4px; }}
        .day-check.all-check span {{ font-weight: 600; color: #FF6600; }}
        .edit-overlay {{ position: fixed; inset: 0; z-index: 199; background: rgba(0,0,0,0.5); }}
        .save-edits-btn {{ display: none; padding: 6px 16px; border: none; border-radius: 8px;
                           font-size: 12px; font-weight: 600; cursor: pointer; font-family: inherit;
                           transition: all 0.2s; margin-left: 8px; }}
        .save-edits-btn.dirty {{ display: inline-block; background: #FF6600; color: #fff;
                                  animation: pulse-save 1.5s ease-in-out infinite; }}
        .save-edits-btn.saving {{ background: #666; color: #fff; cursor: wait; }}
        .save-edits-btn.saved {{ display: inline-block; background: #2a6e2a; color: #64dc3c; }}
        .save-edits-btn.error {{ display: inline-block; background: #6e2a2a; color: #ff6666; }}
        @keyframes pulse-save {{ 0%,100% {{ box-shadow: 0 0 4px rgba(255,102,0,0.3); }}
                                 50% {{ box-shadow: 0 0 14px rgba(255,102,0,0.6); }} }}
        .edit-status {{ font-size: 10px; color: #64dc3c; margin-left: auto; }}

        /* ── Unsaved changes banner ── */
        .unsaved-banner {{ position: fixed; top: 0; left: 0; right: 0; z-index: 9999;
                           background: linear-gradient(135deg, #cc5500, #FF6600);
                           color: #fff; padding: 10px 16px; display: flex; align-items: center;
                           justify-content: center; gap: 12px; font-size: 13px; font-weight: 600;
                           box-shadow: 0 2px 12px rgba(255,102,0,0.4);
                           animation: slideDown 0.3s ease-out; }}
        .unsaved-banner .unsaved-icon {{ font-size: 18px; animation: pulse-icon 1.5s ease-in-out infinite; }}
        .unsaved-banner .unsaved-btn {{ padding: 5px 14px; border: 2px solid #fff; border-radius: 6px;
                                        background: transparent; color: #fff; font-size: 12px;
                                        font-weight: 700; cursor: pointer; font-family: inherit;
                                        transition: all 0.2s; }}
        .unsaved-banner .unsaved-btn:hover {{ background: #fff; color: #FF6600; }}
        .unsaved-banner .unsaved-btn.btn-discard {{ border-color: rgba(255,255,255,0.4);
                                                     color: rgba(255,255,255,0.8); font-weight: 500; }}
        .unsaved-banner .unsaved-btn.btn-discard:hover {{ background: rgba(255,255,255,0.15); color: #fff; }}
        @keyframes slideDown {{ from {{ transform: translateY(-100%); }} to {{ transform: translateY(0); }} }}
        @keyframes pulse-icon {{ 0%,100% {{ opacity: 1; }} 50% {{ opacity: 0.5; }} }}
        body.has-unsaved-banner {{ padding-top: 46px; }}

        /* ── Confirm modal (replaces native confirm) ── */
        .confirm-overlay {{ position: fixed; inset: 0; z-index: 10000;
                            background: rgba(0,0,0,0.7); display: flex; align-items: center;
                            justify-content: center; animation: fadeIn 0.15s ease-out; }}
        .confirm-dialog {{ background: #1a1a2e; border: 1px solid rgba(255,102,0,0.3);
                           border-radius: 12px; padding: 24px; min-width: 300px; max-width: 400px;
                           box-shadow: 0 10px 40px rgba(0,0,0,0.6); text-align: center; }}
        .confirm-dialog .confirm-icon {{ font-size: 36px; margin-bottom: 12px; }}
        .confirm-dialog .confirm-msg {{ font-size: 14px; color: #ddd; margin-bottom: 20px; line-height: 1.5; }}
        .confirm-dialog .confirm-actions {{ display: flex; gap: 10px; }}
        .confirm-dialog .confirm-actions button {{ flex: 1; padding: 10px; border: none; border-radius: 8px;
                                                    font-size: 13px; font-weight: 600; cursor: pointer;
                                                    font-family: inherit; transition: all 0.2s; }}
        .confirm-dialog .btn-confirm-save {{ background: #FF6600; color: #fff; }}
        .confirm-dialog .btn-confirm-save:hover {{ background: #ff9050; }}
        .confirm-dialog .btn-confirm-discard {{ background: rgba(220,50,50,0.15); color: #e55; }}
        .confirm-dialog .btn-confirm-discard:hover {{ background: rgba(220,50,50,0.3); color: #ff6666; }}
        .confirm-dialog .btn-confirm-cancel {{ background: rgba(255,255,255,0.08); color: #888; }}
        .confirm-dialog .btn-confirm-cancel:hover {{ background: rgba(255,255,255,0.15); color: #fff; }}
        @keyframes fadeIn {{ from {{ opacity: 0; }} to {{ opacity: 1; }} }}

        /* ── Drag-resize handles ── */
        .tl-bar.editable .drag-handle {{
            position: absolute; top: 0; bottom: 0; width: 8px;
            cursor: ew-resize; z-index: 10; opacity: 0;
            transition: opacity 0.15s;
        }}
        .tl-bar.editable:hover .drag-handle,
        .tl-bar.editable .drag-handle.active {{ opacity: 1; }}
        .drag-handle.left {{ left: -2px; border-radius: 4px 0 0 4px; background: linear-gradient(90deg, rgba(255,102,0,0.7), transparent); }}
        .drag-handle.right {{ right: -2px; border-radius: 0 4px 4px 0; background: linear-gradient(270deg, rgba(255,102,0,0.7), transparent); }}
        .drag-handle::after {{
            content: ''; position: absolute; top: 50%; transform: translateY(-50%);
            width: 2px; height: 12px; background: rgba(255,255,255,0.7); border-radius: 1px;
        }}
        .drag-handle.left::after {{ left: 2px; }}
        .drag-handle.right::after {{ right: 2px; }}
        .tl-bar-container {{ cursor: default; }}
        .add-staff-row {{ display: flex; justify-content: center; padding: 8px 0; }}
        .add-staff-btn {{ background: rgba(255,102,0,0.1); border: 1px dashed rgba(255,102,0,0.3);
                          color: #FF6600; font-size: 12px; font-weight: 600; padding: 6px 18px;
                          border-radius: 8px; cursor: pointer; font-family: inherit; transition: all 0.2s; }}
        .add-staff-btn:hover {{ background: rgba(255,102,0,0.2); border-color: #FF6600; }}
        .del-staff {{ display: inline-block; margin-left: 4px; color: #e55; font-size: 14px;
                      font-weight: 700; cursor: pointer; opacity: 0.5; transition: opacity 0.15s;
                      line-height: 1; vertical-align: middle; }}
        .del-staff:hover {{ opacity: 1; }}
        .drag-tooltip {{
            position: fixed; z-index: 300; padding: 3px 8px;
            background: #1a1a2e; border: 1px solid rgba(255,102,0,0.5);
            border-radius: 6px; font-size: 11px; font-weight: 600;
            color: #FF6600; pointer-events: none; white-space: nowrap;
            box-shadow: 0 4px 12px rgba(0,0,0,0.5);
        }}

        /* ── Legend ── */
        .legend {{ display: flex; flex-wrap: wrap; gap: 6px; justify-content: center;
                   margin-bottom: 15px; }}
        .legend-item {{ display: flex; align-items: center; gap: 4px; font-size: 10px;
                        color: #666; padding: 3px 8px; background: rgba(255,255,255,0.03);
                        border-radius: 6px; }}
        .legend-dot {{ width: 8px; height: 8px; border-radius: 50%; }}

        /* ── Desktop : planning agrandi ── */
        @media (min-width: 900px) {{
            body {{ padding: 24px 40px; }}
            .container {{ max-width: 1100px; padding: 8px 18px; }}
            h1 {{ font-size: 24px; letter-spacing: 2px; }}
            .subtitle {{ font-size: 14px; }}
            .dates {{ font-size: 16px; padding: 8px 20px; }}
            .week-selector {{ gap: 8px; }}
            .week-tab {{ font-size: 14px; padding: 10px 18px; }}
            .top-bar .admin-btn, .top-bar .psg-btn {{ font-size: 11px; padding: 6px 12px; }}
            .view-toggle {{ max-width: 500px; margin-left: auto; margin-right: auto; }}
            .view-btn {{ font-size: 14px; padding: 10px; }}
            .day-tabs {{ gap: 6px; margin-bottom: 16px; }}
            .day-tab {{ font-size: 13px; padding: 10px 14px; border-radius: 10px; }}
            .timeline {{ margin-bottom: 28px; }}
            .timeline-row {{ margin-bottom: 6px; }}
            .tl-name {{ width: 130px; font-size: 13px; padding-right: 14px; }}
            .tl-bar-container {{ height: 36px; border-radius: 7px; }}
            .tl-bar {{ border-radius: 7px; font-size: 11px; border-left-width: 3px; }}
            .tl-bar .bar-label {{ padding: 0 6px; }}
            .time-marker {{ font-size: 12px; }}
            .legend {{ gap: 10px; margin-bottom: 20px; }}
            .legend-item {{ font-size: 12px; padding: 4px 10px; }}
            .legend-dot {{ width: 10px; height: 10px; }}
            .employee-list {{ gap: 8px; }}
            .employee-btn {{ font-size: 15px; padding: 14px 18px; }}
            .badge {{ font-size: 12px; padding: 4px 10px; }}
            .modal {{ max-width: 650px; }}
            .modal-header h2 {{ font-size: 18px; }}
            .modal-event .ev-time {{ font-size: 13px; min-width: 100px; }}
            .modal-event .ev-label {{ font-size: 14px; }}
            .cal-chooser {{ max-width: 550px; }}
        }}
        @media (min-width: 1300px) {{
            .container {{ max-width: 1400px; padding: 10px 24px; }}
            .tl-name {{ width: 150px; font-size: 14px; }}
            .tl-bar-container {{ height: 40px; }}
            .tl-bar {{ font-size: 12px; }}
            .time-marker {{ font-size: 13px; }}
        }}
        .meta-note {{
            text-align: center; color: #FF6600; font-size: 10px;
            margin: 16px 0 4px; letter-spacing: 0.5px;
        }}
        /* ── Diff popup (collage intelligent) ── */
        .diff-overlay {{ position: fixed; inset: 0; background: rgba(0,0,0,0.7);
                         z-index: 300; display: flex; align-items: center; justify-content: center; }}
        .diff-popup {{ background: #1a1a2e; border: 1px solid rgba(255,102,0,0.3);
                       border-radius: 12px; padding: 16px; max-width: 520px; width: 92%;
                       max-height: 85vh; overflow-y: auto;
                       box-shadow: 0 10px 40px rgba(0,0,0,0.6); }}
        .diff-popup h3 {{ font-size: 13px; color: #FF6600; margin-bottom: 12px; }}
        .diff-lines {{ font-size: 11px; line-height: 1.6; margin-bottom: 12px; }}
        .diff-line {{ padding: 2px 6px; border-radius: 3px; margin-bottom: 2px;
                      white-space: pre-wrap; word-break: break-word; }}
        .diff-line.removed {{ background: rgba(220,50,50,0.15); color: #e88; text-decoration: line-through; }}
        .diff-line.added {{ background: rgba(50,180,50,0.15); color: #8e8; }}
        .diff-line.same {{ color: #666; }}
        .diff-actions {{ display: flex; gap: 8px; margin-top: 12px; }}
        .diff-actions button {{ flex: 1; padding: 10px; border: none; border-radius: 8px;
                                font-size: 12px; font-weight: 600; cursor: pointer;
                                font-family: inherit; transition: all 0.2s; }}
        .diff-btn-apply {{ background: #FF6600; color: #fff; }}
        .diff-btn-apply:hover {{ background: #ff9050; }}
        .diff-btn-cancel {{ background: rgba(255,255,255,0.08); color: #888; }}
        .diff-btn-cancel:hover {{ background: rgba(255,255,255,0.15); color: #fff; }}
        /* ── Action cards dans le popup ── */
        .smart-actions {{ margin-top: 12px; border-top: 1px solid rgba(255,102,0,0.2); padding-top: 10px; }}
        .smart-actions-title {{ font-size: 12px; color: #FF6600; font-weight: 600; margin-bottom: 8px; }}
        .action-card {{ background: rgba(255,102,0,0.06); border: 1px solid rgba(255,102,0,0.15);
                        border-radius: 8px; padding: 8px 10px; margin-bottom: 8px; }}
        .action-card-header {{ display: flex; align-items: center; gap: 8px; margin-bottom: 6px; }}
        .action-card-header input[type="checkbox"] {{ accent-color: #FF6600; width: 16px; height: 16px; }}
        .action-card-header .action-type {{ font-size: 10px; font-weight: 700; text-transform: uppercase;
                                            letter-spacing: 0.5px; padding: 2px 6px; border-radius: 4px; }}
        .action-type-repl {{ background: rgba(255,102,0,0.2); color: #FF6600; }}
        .action-type-sched {{ background: rgba(50,180,50,0.2); color: #8e8; }}
        .action-card-raw {{ font-size: 9px; color: #555; margin-bottom: 6px; font-style: italic; }}
        .action-fields {{ display: grid; grid-template-columns: 1fr 1fr; gap: 4px 8px; }}
        .action-field {{ display: flex; flex-direction: column; gap: 2px; }}
        .action-field label {{ font-size: 9px; color: #666; text-transform: uppercase; letter-spacing: 0.5px; }}
        .action-field select, .action-field input {{
            padding: 5px 6px; background: rgba(0,0,0,0.3); border: 1px solid rgba(255,255,255,0.1);
            border-radius: 5px; color: #fff; font-size: 11px; font-family: inherit; outline: none; }}
        .action-field select:focus, .action-field input:focus {{ border-color: rgba(255,102,0,0.4); }}
    </style>
</head>
<body>
    <div class="container">
        <div class="top-bar">
            <button class="admin-btn" id="admin-btn" title="Mode admin">Admin</button>
            <a href="presences.html" class="psg-btn" title="Suivi des présences PSG Academy">PSG Academy</a>
        </div>
        <div class="header">
            <h1>Planning Urban 7D</h1>
            <p class="subtitle">Semaine {week_num}</p>
            <div class="dates">{date_range}</div>
        </div>

        <div class="week-selector">
{week_tabs.rstrip()}
        </div>

        <div class="week-notes" id="week-notes"></div>

        <div class="view-toggle">
            <button class="view-btn active" data-view="day">Vue quotidienne</button>
            <button class="view-btn" data-view="staff">Vue hebdo par staff</button>
        </div>

        <!-- ── Vue Journée (timeline) ── -->
        <div id="view-day">
            <div class="day-tabs" id="day-tabs"></div>
            <div class="legend" id="legend"></div>
            <div class="timeline" id="timeline"></div>
        </div>

        <!-- ── Vue Staff (liste) ── -->
        <div id="view-staff" style="display:none;">
            <div class="employee-list">
{employee_buttons.rstrip()}
            </div>
        </div>
    </div>

    <!-- ── Modal preview individuelle ── -->
    <div class="modal-overlay" id="modal">
        <div class="modal">
            <div class="modal-header">
                <h2 id="modal-name"></h2>
                <div style="display:flex;align-items:center;gap:8px;">
                    <button class="subscribe-btn" id="modal-subscribe" style="padding:6px 14px;font-size:11px;">
                        S'abonner
                    </button>
                    <button class="modal-close" id="modal-close">&times;</button>
                </div>
            </div>
            <div class="modal-body" id="modal-body"></div>
        </div>
    </div>

    <!-- ── Choix application calendrier ── -->
    <div class="cal-chooser-overlay" id="cal-chooser">
        <div class="cal-chooser">
            <h3>Ajouter au calendrier</h3>
            <div class="cal-sub" id="cal-chooser-name"></div>
            <a class="cal-option" id="cal-google" target="_blank" rel="noopener">
                <span class="cal-icon">G</span>
                <div class="cal-info">
                    <div class="cal-name">Google Agenda</div>
                    <div class="cal-desc">S'abonner (mise \u00e0 jour auto)</div>
                </div>
            </a>
            <div class="cal-option cal-option-small" id="cal-google-manual" style="cursor:pointer;">
                <span class="cal-icon" style="font-size:0.8em;">\u2139</span>
                <div class="cal-info">
                    <div class="cal-name" style="font-size:0.85em;">Google Agenda (manuel)</div>
                    <div class="cal-desc">Si le bouton ci-dessus ouvre l'appli au lieu du navigateur</div>
                </div>
            </div>
            <a class="cal-option" id="cal-apple">
                <span class="cal-icon">\uf8ff</span>
                <div class="cal-info">
                    <div class="cal-name">Apple Calendar</div>
                    <div class="cal-desc">iPhone, iPad, Mac (mise \u00e0 jour auto)</div>
                </div>
            </a>
            <a class="cal-option" id="cal-outlook" target="_blank" rel="noopener">
                <span class="cal-icon">O</span>
                <div class="cal-info">
                    <div class="cal-name">Outlook</div>
                    <div class="cal-desc">Outlook.com / Office 365</div>
                </div>
            </a>
            <a class="cal-option" id="cal-download">
                <span class="cal-icon">\u2b07</span>
                <div class="cal-info">
                    <div class="cal-name">T\u00e9l\u00e9charger .ics</div>
                    <div class="cal-desc">Import ponctuel (ne se met pas \u00e0 jour)</div>
                </div>
            </a>
            <button class="cal-chooser-cancel" id="cal-cancel">Annuler</button>
        </div>
    </div>

    <!-- ── Instructions Google Agenda ── -->
    <div class="cal-chooser-overlay" id="google-instructions" style="display:none;">
        <div class="cal-chooser">
            <h3>Google Agenda</h3>
            <div class="google-steps">
                <div class="step-url" id="google-url-box"></div>
                <button class="copy-url-btn" id="google-copy-btn">Copier le lien</button>
                <div class="step-divider"></div>
                <p class="step-title">Ensuite :</p>
                <ol class="step-list">
                    <li>Ouvrir <a href="https://calendar.google.com/calendar/r/settings/addbyurl" target="_blank" rel="noopener" class="step-link">Google Agenda &gt; Ajouter par URL</a></li>
                    <li>Coller le lien copi\u00e9</li>
                    <li>Cliquer sur <strong>Ajouter l'agenda</strong></li>
                </ol>
                <p class="step-note">L'agenda se mettra \u00e0 jour automatiquement.</p>
                <a class="email-btn" id="google-email">Envoyer les instructions par email</a>
            </div>
            <button class="cal-chooser-cancel" id="google-back">Retour</button>
        </div>
    </div>

    <script>
    (function() {{
        var NOTES_DATA = {notes_json};
        // Nettoyage des anciennes données localStorage (source de désync entre appareils)
        try {{ localStorage.removeItem('planning-notes-S{week_num}'); }} catch(e) {{}}
        var DATA = {events_json};
        // Nettoyage des anciennes données localStorage
        try {{ localStorage.removeItem('planning-edits-S{week_num}'); }} catch(e) {{}}
        var COLORS = {colors_json};
        var DEFAULT_C = {default_color_json};
        var DAYS = {day_labels_json};
        var DAYS_FULL = {day_labels_full_json};
        var WEEK_DATES = {week_dates_json};
        var currentDay = 0;
        // Griser les onglets des semaines passées
        (function() {{
            var today = new Date();
            var yyyy = today.getFullYear();
            var mm = String(today.getMonth()+1).padStart(2,"0");
            var dd = String(today.getDate()).padStart(2,"0");
            var todayStr = yyyy + "-" + mm + "-" + dd;
            document.querySelectorAll('.week-tab[data-end]').forEach(function(tab) {{
                if (tab.getAttribute('data-end') < todayStr && !tab.classList.contains('active')) {{
                    tab.classList.add('past');
                }}
            }});
        }})();
        (function() {{
            var now = new Date();
            var today = now.getFullYear() + '-' + String(now.getMonth()+1).padStart(2,'0') + '-' + String(now.getDate()).padStart(2,'0');
            var idx = WEEK_DATES.indexOf(today);
            if (idx !== -1) currentDay = idx;
        }})();
        var currentView = 'day';

        function getColor(code) {{ return COLORS[code] || DEFAULT_C; }}
        function getFirstName(n) {{ var p=n.split(' '); for(var i=0;i<p.length;i++){{ if(p[i]!==p[i].toUpperCase()) return p.slice(i).join(' '); }} return p[p.length-1]; }}

        // ── Replacement matching ──
        function getReplacements() {{
            return (notesWork && notesWork.replacements) || [];
        }}
        function getReplacementStatus(fullName, dateStr, startH, endH) {{
            var repls = getReplacements();
            for (var i = 0; i < repls.length; i++) {{
                var r = repls[i];
                if (r.date !== dateStr) continue;
                var rStart = parseFloat(r.start.split(':')[0]) + parseFloat(r.start.split(':')[1] || 0) / 60;
                var rEnd = parseFloat(r.end.split(':')[0]) + parseFloat(r.end.split(':')[1] || 0) / 60;
                // Check overlap: bar overlaps with replacement window
                if (startH < rEnd && endH > rStart) {{
                    if (fullName === r.out) return {{status: 'out', other: r['in']}};
                    if (fullName === r['in']) return {{status: 'in', other: r.out}};
                }}
            }}
            return null;
        }}

        // ── Day tabs ──
        var dayTabsEl = document.getElementById('day-tabs');
        DAYS.forEach(function(label, i) {{
            var btn = document.createElement('div');
            btn.className = 'day-tab' + (i === currentDay ? ' active' : '');
            btn.textContent = label;
            btn.onclick = function() {{ selectDay(i); }};
            dayTabsEl.appendChild(btn);
        }});

        function selectDay(i) {{
            currentDay = i;
            dayTabsEl.querySelectorAll('.day-tab').forEach(function(t, j) {{
                t.classList.toggle('active', j === i);
            }});
            renderTimeline();
        }}

        // Scroll active day tab into view
        setTimeout(function() {{
            var activeTab = dayTabsEl.querySelector('.day-tab.active');
            if (activeTab) activeTab.scrollIntoView({{ inline: 'center', block: 'nearest' }});
        }}, 0);

        // ── Legend ── Build code-to-label map from all events
        var CODE_LABELS = {{}};
        Object.keys(DATA).forEach(function(n) {{
            if (n === '_codeNames' || n === '_meta') return;
            DATA[n].events.forEach(function(ev) {{
                if (ev.label && ev.label !== ev.code) CODE_LABELS[ev.code] = ev.label;
            }});
        }});
        function renderLegend(codes) {{
            var el = document.getElementById('legend');
            el.innerHTML = '';
            var seen = {{}};
            codes.forEach(function(code) {{
                if (seen[code]) return;
                seen[code] = true;
                var c = getColor(code);
                var item = document.createElement('div');
                item.className = 'legend-item';
                var displayName = CODE_LABELS[code] || (DATA._codeNames && DATA._codeNames[code]) || code;
                item.innerHTML = '<div class="legend-dot" style="background:' + c.border +
                    ';box-shadow:0 0 6px ' + c.border + '"></div>' + displayName;
                el.appendChild(item);
            }});
        }}

        // ── Timeline rendering ──
        function renderTimeline() {{
            var tl = document.getElementById('timeline');
            tl.innerHTML = '';
            var dateStr = WEEK_DATES[currentDay] || '';

            // Collect events for this day
            var dayEvents = [];
            var allCodes = [];
            Object.keys(DATA).forEach(function(name) {{
                if (name === '_codeNames' || name === '_meta') return;
                var emp = DATA[name];
                emp.events.forEach(function(ev) {{
                    if (ev.day === currentDay) {{
                        dayEvents.push({{ name: name, ev: ev }});
                        allCodes.push(ev.code);
                    }}
                }});
            }});

            // Inject virtual events for replacers not already present this day
            var dayRepls = getReplacements().filter(function(r) {{ return r.date === dateStr; }});
            dayRepls.forEach(function(r) {{
                var replacerName = r['in'];
                if (!replacerName || !DATA[replacerName]) return;
                // Check if replacer already has events this day
                var hasEvents = dayEvents.some(function(d) {{ return d.name === replacerName; }});
                // Find the replaced person's event(s) overlapping the replacement window to get code/label
                var rStart = parseFloat(r.start.split(':')[0]) + parseFloat(r.start.split(':')[1] || 0) / 60;
                var rEnd = parseFloat(r.end.split(':')[0]) + parseFloat(r.end.split(':')[1] || 0) / 60;
                var outName = r.out;
                var refCode = 'VDC';
                var refLabel = 'Vie de centre';
                if (outName && DATA[outName]) {{
                    DATA[outName].events.forEach(function(ev) {{
                        if (ev.day !== currentDay) return;
                        var s2 = new Date(ev.start);
                        var e2 = new Date(ev.end);
                        var sh2 = s2.getHours() + s2.getMinutes()/60;
                        var eh2 = e2.getHours() + e2.getMinutes()/60;
                        if (eh2 <= sh2) eh2 = 24;
                        if (sh2 < rEnd && eh2 > rStart) {{
                            refCode = ev.code;
                            refLabel = ev.label;
                        }}
                    }});
                }}
                if (!hasEvents) {{
                    // Build synthetic ISO dates for the replacement window
                    var synthStart = dateStr + 'T' + r.start.split(':')[0].padStart(2,'0') + ':' + (r.start.split(':')[1] || '00').padStart(2,'0');
                    var synthEnd = dateStr + 'T' + r.end.split(':')[0].padStart(2,'0') + ':' + (r.end.split(':')[1] || '00').padStart(2,'0');
                    var synthEv = {{
                        code: refCode,
                        label: refLabel,
                        start: synthStart,
                        end: synthEnd,
                        day: currentDay,
                        _synthetic: true
                    }};
                    dayEvents.push({{ name: replacerName, ev: synthEv }});
                    allCodes.push(refCode);
                }}
            }});

            if (dayEvents.length === 0) {{
                tl.innerHTML = '<div class="no-events">Aucun cr\u00e9neau ce jour</div>';
                renderLegend([]);
                return;
            }}

            renderLegend(allCodes);

            // Add replacement legend if any replacements exist for this day
            var dayRepls = getReplacements().filter(function(r) {{ return r.date === dateStr; }});
            if (dayRepls.length > 0) {{
                var legendEl = document.getElementById('legend');
                var replOut = document.createElement('div');
                replOut.className = 'legend-item';
                replOut.innerHTML = '<div class="legend-dot" style="background:repeating-linear-gradient(45deg,transparent,transparent 2px,rgba(255,60,60,0.5) 2px,rgba(255,60,60,0.5) 3px);border:1px solid #ff3c3c"></div>Remplac\u00e9(e)';
                legendEl.appendChild(replOut);
                var replIn = document.createElement('div');
                replIn.className = 'legend-item';
                replIn.innerHTML = '<div class="legend-dot" style="background:repeating-linear-gradient(45deg,transparent,transparent 2px,rgba(60,220,80,0.5) 2px,rgba(60,220,80,0.5) 3px);border:1px solid #3cdc50"></div>Rempla\u00e7ant(e)';
                legendEl.appendChild(replIn);
            }}

            // Scrollable inner wrapper
            var inner = document.createElement('div');
            inner.className = 'timeline-inner';

            // Find time range
            var minH = 24, maxH = 0;
            dayEvents.forEach(function(d) {{
                var s = new Date(d.ev.start);
                var e = new Date(d.ev.end);
                var sh = s.getHours() + s.getMinutes()/60;
                var eh = e.getHours() + e.getMinutes()/60;
                if (eh <= sh) eh = 24;
                if (sh < minH) minH = sh;
                if (eh > maxH) maxH = eh;
            }});
            minH = Math.floor(minH);
            maxH = Math.ceil(maxH);
            if (maxH <= minH) maxH = minH + 1;
            var range = maxH - minH;

            // Set inner width: wider on desktop for comfort
            var isDesktop = window.innerWidth >= 900;
            var pxPerHour = isDesktop ? 80 : 40;
            var nameW = isDesktop ? 150 : 70;
            inner.style.minWidth = (nameW + range * pxPerHour) + 'px';

            // Grid line positions for bar containers
            var gridLines = [];
            for (var gh = minH; gh <= maxH; gh++) {{
                var pos = ((gh - minH) / range) * 100;
                gridLines.push({{ pos: pos, cls: 'hour' }});
                if (gh < maxH) {{
                    var halfPos = ((gh + 0.5 - minH) / range) * 100;
                    gridLines.push({{ pos: halfPos, cls: 'half' }});
                }}
            }}

            // Time markers
            var markerRow = document.createElement('div');
            markerRow.className = 'timeline-row';
            var markerSpacer = document.createElement('div');
            markerSpacer.className = 'tl-name';
            markerSpacer.innerHTML = '&nbsp;';
            markerRow.appendChild(markerSpacer);
            var markers = document.createElement('div');
            markers.className = 'time-markers';
            markers.style.flex = '1';
            for (var h = minH; h <= maxH; h++) {{
                var m = document.createElement('span');
                m.className = 'time-marker';
                m.textContent = h + 'h';
                markers.appendChild(m);
            }}
            markerRow.appendChild(markers);
            inner.appendChild(markerRow);

            // Group by employee
            var byName = {{}};
            var nameOrder = [];
            dayEvents.forEach(function(d) {{
                if (!byName[d.name]) {{ byName[d.name] = []; nameOrder.push(d.name); }}
                byName[d.name].push(d.ev);
            }});

            nameOrder.forEach(function(name) {{
                var row = document.createElement('div');
                row.className = 'timeline-row';

                var nameEl = document.createElement('div');
                nameEl.className = 'tl-name';
                nameEl.textContent = getFirstName(name);
                nameEl.title = name;
                nameEl.onclick = function() {{ openModal(name); }};
                row.appendChild(nameEl);

                var barContainer = document.createElement('div');
                barContainer.className = 'tl-bar-container';
                barContainer.dataset.minH = minH;
                barContainer.dataset.range = range;

                // Add grid lines
                gridLines.forEach(function(gl) {{
                    var line = document.createElement('div');
                    line.className = 'tl-grid-line ' + gl.cls;
                    line.style.left = gl.pos + '%';
                    barContainer.appendChild(line);
                }});

                byName[name].forEach(function(ev) {{
                    var s = new Date(ev.start);
                    var e = new Date(ev.end);
                    var sh = s.getHours() + s.getMinutes()/60;
                    var eh = e.getHours() + e.getMinutes()/60;
                    if (eh <= sh) eh = 24;

                    var left = ((sh - minH) / range) * 100;
                    var width = ((eh - sh) / range) * 100;
                    if (left < 0) left = 0;
                    if (left + width > 100) width = 100 - left;

                    var c = getColor(ev.code);
                    var bar = document.createElement('div');
                    bar.className = 'tl-bar';

                    // Check replacement status for this bar
                    var replInfo = getReplacementStatus(name, dateStr, sh, eh);
                    if (replInfo && replInfo.status === 'out') bar.className += ' replaced';
                    if (replInfo && replInfo.status === 'in') bar.className += ' replacer';

                    bar.style.cssText = 'left:' + left + '%;width:' + width + '%;' +
                        'background:' + c.bg + ';border-color:' + c.border + ';color:' + c.text +
                        ';--glow-color:' + c.border + ';' +
                        'box-shadow:inset 0 0 8px rgba(255,255,255,0.05), 0 0 4px ' + c.border + '40;';
                    bar.innerHTML = '<span class="bar-label">' + ev.code + '</span>';
                    var timeStr = s.getHours().toString().padStart(2,'0') + ':' + s.getMinutes().toString().padStart(2,'0') +
                        ' - ' + e.getHours().toString().padStart(2,'0') + ':' + e.getMinutes().toString().padStart(2,'0');
                    bar.title = ev.label + '\\n' + timeStr;
                    if (replInfo && replInfo.status === 'out') bar.title += '\\nRemplacé par ' + getFirstName(replInfo.other);
                    if (replInfo && replInfo.status === 'in') bar.title += '\\nRemplace ' + getFirstName(replInfo.other);
                    barContainer.appendChild(bar);
                }});

                row.appendChild(barContainer);
                inner.appendChild(row);
            }});

            tl.appendChild(inner);

            // Auto-scroll to current hour if viewing today + draw now-line
            var _now = new Date();
            var _today = _now.getFullYear() + '-' + String(_now.getMonth()+1).padStart(2,'0') + '-' + String(_now.getDate()).padStart(2,'0');
            if (WEEK_DATES[currentDay] === _today) {{
                var currentH = _now.getHours() + _now.getMinutes() / 60;
                if (currentH >= minH && currentH <= maxH) {{
                    // Draw now-line on each bar container
                    var nowPct = ((currentH - minH) / range) * 100;
                    inner.querySelectorAll('.tl-bar-container').forEach(function(bc) {{
                        var nl = document.createElement('div');
                        nl.className = 'tl-now-line';
                        nl.style.left = nowPct + '%';
                        bc.appendChild(nl);
                    }});
                    // Draw now-line on time markers row
                    var tmRow = inner.querySelector('.time-markers');
                    if (tmRow) {{
                        tmRow.style.position = 'relative';
                        var nm = document.createElement('div');
                        nm.className = 'tl-now-marker';
                        nm.style.left = nowPct + '%';
                        tmRow.appendChild(nm);
                    }}

                    setTimeout(function() {{
                        var scrollPct = (currentH - minH) / range;
                        var nameColWidth = 70;
                        var scrollableWidth = inner.scrollWidth - nameColWidth;
                        var scrollTarget = nameColWidth + scrollPct * scrollableWidth - tl.clientWidth / 2;
                        tl.scrollLeft = Math.max(0, scrollTarget);
                    }}, 0);
                }}
            }}
        }}

        // Auto-update now-line every 60 seconds
        setInterval(function() {{
            var view = document.getElementById('view-day');
            if (view && view.style.display !== 'none') {{
                renderTimeline();
            }}
        }}, 60000);

        function pad2(n) {{ return n.toString().padStart(2, '0'); }}

        function toICSDate(dt) {{
            return dt.getFullYear().toString() +
                pad2(dt.getMonth() + 1) + pad2(dt.getDate()) + 'T' +
                pad2(dt.getHours()) + pad2(dt.getMinutes()) + '00';
        }}

        function icsEscape(str) {{
            return str.replace(/\\\\/g, '\\\\\\\\').replace(/\\n/g, '\\\\n').replace(/,/g, '\\\\,').replace(/;/g, '\\\\;');
        }}

        function generateICSForNames(names) {{
            // Build notes description from NOTES_DATA (notes only, no label)
            var noteDesc = '';
            if (NOTES_DATA.comment) {{
                noteDesc += NOTES_DATA.comment;
            }}
            (NOTES_DATA.updates || []).forEach(function(u) {{
                if (u.text) {{
                    var prefix = u.date ? ('MAJ ' + u.date + ': ') : 'MAJ: ';
                    if (noteDesc) noteDesc += '\\n';
                    noteDesc += prefix + u.text;
                }}
            }});

            var lines = [
                'BEGIN:VCALENDAR', 'VERSION:2.0',
                'PRODID:-//Planning Urban 7D//FR',
                'CALSCALE:GREGORIAN', 'METHOD:PUBLISH',
                'X-WR-CALNAME:Planning Urban 7D',
                'X-WR-TIMEZONE:Europe/Paris'
            ];
            names.forEach(function(name) {{
                var emp = DATA[name];
                if (!emp) return;
                emp.events.forEach(function(ev, i) {{
                    var s = new Date(ev.start);
                    var e = new Date(ev.end);
                    lines.push('BEGIN:VEVENT');
                    lines.push('UID:export-' + emp.slug + '-' + i + '@urban7d');
                    lines.push('DTSTART;TZID=Europe/Paris:' + toICSDate(s));
                    lines.push('DTEND;TZID=Europe/Paris:' + toICSDate(e));
                    lines.push('SUMMARY:' + getFirstName(name) + ' - ' + ev.label);
                    if (noteDesc) lines.push('DESCRIPTION:' + icsEscape(noteDesc));
                    lines.push('END:VEVENT');
                }});
            }});
            lines.push('END:VCALENDAR');
            return lines.join('\\r\\n');
        }}

        // ── Calendar chooser (universel tous navigateurs / OS) ──
        var _currentIcsUrl = '';

        function openCalendarChooser(slug, displayName) {{
            var base = window.location.href.replace(/[^/]*$/, '');
            var icsPath = 'ics/' + slug + '.ics';
            var fullUrl = new URL(icsPath, base).href;
            var webcalUrl = 'webcal://' + new URL(icsPath, base).host + new URL(icsPath, base).pathname;
            var calName = encodeURIComponent('Planning ' + displayName);
            _currentIcsUrl = fullUrl;

            document.getElementById('cal-chooser-name').textContent = displayName;
            document.getElementById('cal-google').href =
                'https://calendar.google.com/calendar/render?cid=' + encodeURIComponent(webcalUrl);
            document.getElementById('cal-apple').href = webcalUrl;
            document.getElementById('cal-outlook').href =
                'https://outlook.live.com/calendar/0/addfromweb?url=' + encodeURIComponent(fullUrl) + '&name=' + calName;
            document.getElementById('cal-download').href = icsPath;
            document.getElementById('cal-download').setAttribute('download', slug + '.ics');
            document.getElementById('google-url-box').textContent = fullUrl;

            document.getElementById('cal-chooser').style.display = '';
            document.getElementById('cal-chooser').classList.add('open');
        }}

        function closeCalendarChooser() {{
            document.getElementById('cal-chooser').classList.remove('open');
            document.getElementById('cal-chooser').style.display = '';
            document.getElementById('google-instructions').style.display = 'none';
        }}
        document.getElementById('cal-cancel').onclick = closeCalendarChooser;
        document.getElementById('cal-chooser').onclick = function(e) {{
            if (e.target === this) closeCalendarChooser();
        }};

        // Google Agenda (manuel) : ouvre le panneau d'instructions
        document.getElementById('cal-google-manual').onclick = function() {{
            document.getElementById('cal-chooser').style.display = 'none';
            document.getElementById('cal-chooser').classList.remove('open');
            document.getElementById('google-instructions').style.display = '';
            // Copie auto du lien
            if (navigator.clipboard) {{
                navigator.clipboard.writeText(_currentIcsUrl).then(function() {{
                    var btn = document.getElementById('google-copy-btn');
                    btn.textContent = 'Lien copi\u00e9 !';
                    btn.classList.add('copied');
                    setTimeout(function() {{ btn.textContent = 'Copier le lien'; btn.classList.remove('copied'); }}, 3000);
                }});
            }}
        }};
        document.getElementById('google-copy-btn').onclick = function() {{
            var btn = this;
            if (navigator.clipboard) {{
                navigator.clipboard.writeText(_currentIcsUrl).then(function() {{
                    btn.textContent = 'Lien copi\u00e9 !';
                    btn.classList.add('copied');
                    setTimeout(function() {{ btn.textContent = 'Copier le lien'; btn.classList.remove('copied'); }}, 3000);
                }});
            }} else {{
                prompt('Copier ce lien :', _currentIcsUrl);
            }}
        }};
        document.getElementById('google-back').onclick = function() {{
            document.getElementById('google-instructions').style.display = 'none';
            document.getElementById('cal-chooser').style.display = '';
            document.getElementById('cal-chooser').classList.add('open');
        }};
        document.getElementById('google-instructions').onclick = function(e) {{
            if (e.target === this) closeCalendarChooser();
        }};

        // Bouton email : ouvre un mailto avec instructions
        var _currentDisplayName = '';
        var _origOpenChooser = openCalendarChooser;
        openCalendarChooser = function(slug, displayName) {{
            _currentDisplayName = displayName;
            _origOpenChooser(slug, displayName);
        }};
        document.getElementById('google-email').onclick = function(e) {{
            e.preventDefault();
            var addUrl = 'https://calendar.google.com/calendar/r/settings/addbyurl';
            var subject = 'Planning Urban 7D - ' + _currentDisplayName;
            var body = 'Salut ' + _currentDisplayName + ' !\\n\\n'
                + 'Pour ajouter ton planning dans Google Agenda :\\n\\n'
                + '1. Copie ce lien :\\n' + _currentIcsUrl + '\\n\\n'
                + '2. Ouvre cette page (depuis un ordi ou navigateur web) :\\n' + addUrl + '\\n\\n'
                + '3. Colle le lien et clique "Ajouter l\\'agenda"\\n\\n'
                + 'Ton planning se mettra \u00e0 jour automatiquement !';
            window.location.href = 'mailto:?subject=' + encodeURIComponent(subject) + '&body=' + encodeURIComponent(body);
        }};

        // ── View toggle ──
        document.querySelectorAll('.view-btn').forEach(function(btn) {{
            btn.onclick = function() {{
                currentView = btn.getAttribute('data-view');
                document.querySelectorAll('.view-btn').forEach(function(b) {{
                    b.classList.toggle('active', b === btn);
                }});
                document.getElementById('view-day').style.display = currentView === 'day' ? '' : 'none';
                document.getElementById('view-staff').style.display = currentView === 'staff' ? '' : 'none';
            }};
        }});

        // ── Modal ──
        var modalEl = document.getElementById('modal');
        document.getElementById('modal-close').onclick = closeModal;
        modalEl.onclick = function(e) {{ if (e.target === modalEl) closeModal(); }};

        function closeModal() {{ modalEl.classList.remove('open'); }}

        function openModal(name) {{
            var emp = DATA[name];
            if (!emp) return;

            document.getElementById('modal-name').textContent = getFirstName(name);
            var body = document.getElementById('modal-body');
            body.innerHTML = '';

            // Group events by day
            var byDay = {{}};
            emp.events.forEach(function(ev) {{ if (!byDay[ev.day]) byDay[ev.day] = []; byDay[ev.day].push(ev); }});

            // Inject virtual events for days where this person is a replacer but has no events
            var repls = getReplacements();
            repls.forEach(function(r) {{
                if (r['in'] !== name) return;
                // Find the day index for this replacement date
                var dayIdx = WEEK_DATES.indexOf(r.date);
                if (dayIdx < 0) return;
                var hasEventsThisDay = byDay[dayIdx] && byDay[dayIdx].length > 0;
                if (hasEventsThisDay) return;
                // Find code/label from replaced person's event
                var outName = r.out;
                var refCode = 'VDC';
                var refLabel = 'Vie de centre';
                var rStart = parseFloat(r.start.split(':')[0]) + parseFloat(r.start.split(':')[1] || 0) / 60;
                var rEnd = parseFloat(r.end.split(':')[0]) + parseFloat(r.end.split(':')[1] || 0) / 60;
                if (outName && DATA[outName]) {{
                    DATA[outName].events.forEach(function(ev) {{
                        if (ev.day !== dayIdx) return;
                        var s2 = new Date(ev.start);
                        var e2 = new Date(ev.end);
                        var sh2 = s2.getHours() + s2.getMinutes()/60;
                        var eh2 = e2.getHours() + e2.getMinutes()/60;
                        if (eh2 <= sh2) eh2 = 24;
                        if (sh2 < rEnd && eh2 > rStart) {{
                            refCode = ev.code;
                            refLabel = ev.label;
                        }}
                    }});
                }}
                var synthStart = r.date + 'T' + r.start.split(':')[0].padStart(2,'0') + ':' + (r.start.split(':')[1] || '00').padStart(2,'0');
                var synthEnd = r.date + 'T' + r.end.split(':')[0].padStart(2,'0') + ':' + (r.end.split(':')[1] || '00').padStart(2,'0');
                if (!byDay[dayIdx]) byDay[dayIdx] = [];
                byDay[dayIdx].push({{
                    code: refCode,
                    label: refLabel,
                    start: synthStart,
                    end: synthEnd,
                    day: dayIdx,
                    _synthetic: true
                }});
            }});

            var hasDays = false;
            for (var d = 0; d < 7; d++) {{
                if (!byDay[d] || byDay[d].length === 0) continue;
                hasDays = true;
                var dayDiv = document.createElement('div');
                dayDiv.className = 'modal-day';

                var title = document.createElement('div');
                title.className = 'modal-day-title';
                title.textContent = DAYS_FULL[d];
                dayDiv.appendChild(title);

                byDay[d].forEach(function(ev) {{
                    var c = getColor(ev.code);
                    var s = new Date(ev.start);
                    var e = new Date(ev.end);
                    var sh = s.getHours() + s.getMinutes()/60;
                    var eh = e.getHours() + e.getMinutes()/60;
                    if (eh <= sh) eh = 24;
                    var evDateStr = ev.start.substring(0, 10);

                    var evDiv = document.createElement('div');
                    evDiv.className = 'modal-event';

                    // Check replacement status
                    var replInfo = getReplacementStatus(name, evDateStr, sh, eh);
                    if (replInfo && replInfo.status === 'out') evDiv.className += ' replaced';
                    if (replInfo && replInfo.status === 'in') evDiv.className += ' replacer';

                    evDiv.style.cssText = 'background:' + c.bg + ';border-color:' + c.border +
                        ';box-shadow:0 0 8px ' + c.border + '30;';

                    var timeSpan = document.createElement('span');
                    timeSpan.className = 'ev-time';
                    timeSpan.style.color = c.text;
                    timeSpan.textContent = s.getHours().toString().padStart(2,'0') + ':' +
                        s.getMinutes().toString().padStart(2,'0') + ' \u2192 ' +
                        e.getHours().toString().padStart(2,'0') + ':' +
                        e.getMinutes().toString().padStart(2,'0');

                    var labelSpan = document.createElement('span');
                    labelSpan.className = 'ev-label';
                    labelSpan.style.color = c.text;
                    labelSpan.textContent = ev.label;

                    evDiv.appendChild(timeSpan);
                    evDiv.appendChild(labelSpan);

                    // Add replacement annotation text
                    if (replInfo) {{
                        var replSpan = document.createElement('span');
                        replSpan.className = 'ev-repl';
                        if (replInfo.status === 'out') {{
                            replSpan.textContent = '\u2194 ' + getFirstName(replInfo.other);
                            replSpan.style.color = '#ff6b6b';
                        }} else {{
                            replSpan.textContent = '\u2194 ' + getFirstName(replInfo.other);
                            replSpan.style.color = '#51cf66';
                        }}
                        evDiv.appendChild(replSpan);
                    }}

                    dayDiv.appendChild(evDiv);
                }});
                body.appendChild(dayDiv);
            }}

            if (!hasDays) {{
                body.innerHTML = '<div class="no-events">Repos cette semaine</div>';
            }} else {{
                // Total weekly hours footer (brut / net)
                var totalH = computeWeeklyHours(emp);
                var footer = document.createElement('div');
                footer.className = 'modal-hours-total';
                if (totalH.pause > 0) {{
                    footer.innerHTML =
                        '<div class="hours-line">Heures brut : <strong>' + formatHours(totalH.brut) + '</strong></div>' +
                        '<div class="hours-line pause">Pauses (20min / 6h) : <strong>\u2212' + formatHours(totalH.pause) + '</strong></div>' +
                        '<div class="hours-line net">Heures net : <strong>' + formatHours(totalH.net) + '</strong></div>';
                }} else {{
                    footer.innerHTML = 'Total semaine : <strong>' + formatHours(totalH.brut) + '</strong>';
                }}
                body.appendChild(footer);
            }}

            // Subscribe button → opens calendar chooser
            var subBtn = document.getElementById('modal-subscribe');
            subBtn.onclick = function(e) {{
                e.preventDefault();
                openCalendarChooser(emp.slug, getFirstName(name));
            }};

            modalEl.classList.add('open');
        }}

        // ── Compute weekly hours per employee and display badges ──
        function computeWeeklyHours(emp) {{
            var brut = 0;
            emp.events.forEach(function(ev) {{
                var s = new Date(ev.start);
                var e = new Date(ev.end);
                brut += (e - s) / (1000 * 60 * 60);
            }});
            var pauseH = computeWeeklyPause(emp);
            return {{ brut: brut, net: brut - pauseH, pause: pauseH }};
        }}
        // Compute total pause deduction for an employee's week.
        // Rule: per day, merge consecutive/overlapping shifts into continuous
        // blocks, then for each block every 6h worked → 20 min pause.
        function computeWeeklyPause(emp) {{
            // Group events by day
            var byDay = {{}};
            emp.events.forEach(function(ev) {{
                var d = ev.day;
                if (!byDay[d]) byDay[d] = [];
                byDay[d].push({{ s: new Date(ev.start).getTime(), e: new Date(ev.end).getTime() }});
            }});
            var totalPause = 0; // in hours
            Object.keys(byDay).forEach(function(d) {{
                var intervals = byDay[d].slice().sort(function(a,b) {{ return a.s - b.s; }});
                // Merge strictly consecutive intervals (fin == début du suivant)
                var merged = [intervals[0]];
                for (var i = 1; i < intervals.length; i++) {{
                    var last = merged[merged.length - 1];
                    if (intervals[i].s <= last.e) {{
                        last.e = Math.max(last.e, intervals[i].e);
                    }} else {{
                        merged.push({{ s: intervals[i].s, e: intervals[i].e }});
                    }}
                }}
                // For each continuous block, count pauses (every 6h → 20min)
                merged.forEach(function(block) {{
                    var durationH = (block.e - block.s) / (1000 * 60 * 60);
                    var pauses = Math.floor(durationH / 6);
                    totalPause += pauses * (20 / 60); // 20 min in hours
                }});
            }});
            return totalPause;
        }}
        function formatHours(h) {{
            var hrs = Math.floor(h);
            var mins = Math.round((h - hrs) * 60);
            return mins > 0 ? hrs + 'h' + (mins < 10 ? '0' : '') + mins : hrs + 'h';
        }}
        function updateHoursBadges() {{
            document.querySelectorAll('.employee-btn[data-name]').forEach(function(btn) {{
                var emp = DATA[btn.getAttribute('data-name')];
                if (!emp) return;
                // Remove old badges
                btn.querySelectorAll('.hours-badge').forEach(function(b) {{ b.remove(); }});
                var h = computeWeeklyHours(emp);
                var badge = document.createElement('span');
                badge.className = 'badge hours-badge';
                if (h.pause > 0) {{
                    badge.innerHTML = formatHours(h.net) + ' <span class="hours-brut">(' + formatHours(h.brut) + ')</span>';
                    badge.title = 'Net : ' + formatHours(h.net) + ' | Brut : ' + formatHours(h.brut) + ' | Pauses : ' + formatHours(h.pause);
                }} else {{
                    badge.textContent = formatHours(h.brut);
                }}
                btn.appendChild(badge);
            }});
        }}
        updateHoursBadges();

        // ── Staff list click ──
        document.querySelectorAll('.employee-btn[data-name]').forEach(function(btn) {{
            btn.onclick = function() {{ openModal(btn.getAttribute('data-name')); }};
        }});

        // ── Notes de semaine (injectées depuis notes/SXX.json) ──
        var REPO = 'OhLaPey/planning-urbansoccer';
        var NOTES_PATH = 'notes/S{week_num}.json';
        var TOKEN_KEY = 'planning-admin-token';
        var ADMIN_KEY = 'planning-admin-unlocked';
        var notesEl = document.getElementById('week-notes');
        var notesWork = JSON.parse(JSON.stringify(NOTES_DATA));
        var notesDirty = false;
        function saveNotesLocal() {{
            // Plus de localStorage — les notes sont en mémoire et persistées via "Publier"
            updateUnsavedBanner();
        }}

        // ── Unsaved changes banner ──
        var _unsavedBannerEl = null;
        function hasUnsavedChanges() {{ return notesDirty || _editsDirty; }}
        function updateUnsavedBanner() {{
            if (hasUnsavedChanges()) {{
                if (!_unsavedBannerEl) {{
                    _unsavedBannerEl = document.createElement('div');
                    _unsavedBannerEl.className = 'unsaved-banner';
                    _unsavedBannerEl.innerHTML =
                        '<span class="unsaved-icon">⚠</span>' +
                        '<span>Modifications non enregistrées</span>';
                    document.body.prepend(_unsavedBannerEl);
                    document.body.classList.add('has-unsaved-banner');
                }}
            }} else {{
                if (_unsavedBannerEl) {{
                    _unsavedBannerEl.remove();
                    _unsavedBannerEl = null;
                    document.body.classList.remove('has-unsaved-banner');
                }}
            }}
        }}
        // ── beforeunload protection ──
        window.addEventListener('beforeunload', function(e) {{
            if (hasUnsavedChanges()) {{ e.preventDefault(); e.returnValue = ''; }}
        }});
        // ── Custom confirm dialog (replaces native confirm) ──
        function showConfirmDialog(msg, onSave, onDiscard) {{
            var overlay = document.createElement('div');
            overlay.className = 'confirm-overlay';
            overlay.innerHTML =
                '<div class="confirm-dialog">' +
                    '<div class="confirm-icon">⚠️</div>' +
                    '<div class="confirm-msg">' + msg + '</div>' +
                    '<div class="confirm-actions">' +
                        '<button class="btn-confirm-save">Enregistrer</button>' +
                        '<button class="btn-confirm-discard">Quitter sans sauver</button>' +
                        '<button class="btn-confirm-cancel">Annuler</button>' +
                    '</div>' +
                '</div>';
            document.body.appendChild(overlay);
            overlay.querySelector('.btn-confirm-save').onclick = function() {{
                overlay.remove(); if (onSave) onSave();
            }};
            overlay.querySelector('.btn-confirm-discard').onclick = function() {{
                overlay.remove(); if (onDiscard) onDiscard();
            }};
            overlay.querySelector('.btn-confirm-cancel').onclick = function() {{
                overlay.remove();
            }};
            overlay.onclick = function(e) {{ if (e.target === overlay) overlay.remove(); }};
        }}

        // ── Parsing intelligent du texte collé ──
        var _staffNames = Object.keys(DATA).filter(function(n) {{ return n !== '_codeNames' && n !== '_meta'; }}).sort();
        var _nameMap = {{}};
        _staffNames.forEach(function(full) {{
            var parts = full.split(' ');
            parts.forEach(function(p) {{ if (p.length > 2) _nameMap[p.toLowerCase()] = full; }});
            if (parts.length > 2) _nameMap[parts.slice(1).join(' ').toLowerCase()] = full;
        }});
        var _dayMap = {{ 'lundi': 0, 'mardi': 1, 'mercredi': 2, 'jeudi': 3,
                         'vendredi': 4, 'samedi': 5, 'dimanche': 6 }};
        var _dayLabelsShort = ['Lun','Mar','Mer','Jeu','Ven','Sam','Dim'];
        var _codeList = Object.keys(COLORS);

        function _findStaff(word) {{ return _nameMap[word.toLowerCase()] || null; }}

        function _detectDay(str) {{
            var dm = str.match(/\\b(lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)\\b/i);
            if (dm) return _dayMap[dm[1].toLowerCase()];
            var ddm = str.match(/(\\d{{1,2}})[/](\\d{{1,2}})/);
            if (ddm) {{
                var dd = ddm[1].padStart(2,'0'), mm = ddm[2].padStart(2,'0');
                for (var i = 0; i < WEEK_DATES.length; i++) {{
                    if (WEEK_DATES[i].endsWith('-' + mm + '-' + dd)) return i;
                }}
            }}
            return null;
        }}

        function _detectTimes(str) {{
            var m = str.match(/(\\d{{1,2}})[hH:](\\d{{0,2}})\\s*[-\u2013\u00e0]\\s*(\\d{{1,2}})[hH:](\\d{{0,2}})/);
            if (m) return [m[1].padStart(2,'0') + ':' + (m[2]||'00').padStart(2,'0'),
                           m[3].padStart(2,'0') + ':' + (m[4]||'00').padStart(2,'0')];
            return null;
        }}

        function _detectCode(str) {{
            var upper = str.toUpperCase();
            for (var i = 0; i < _codeList.length; i++) {{
                if (upper.indexOf(_codeList[i]) >= 0) return _codeList[i];
            }}
            // Fuzzy match common words
            var fuzzy = {{'vie de centre':'VDC','edf':'EDF-C','stage':'STAGE','padel':'C-PAD',
                         'cours padel':'C-PAD','anniversaire':'ANNIV','maladie':'MAL','r\u00e9union':'REU',
                         'cup':'CUP-L','r\u00e9gisseur':'L-REG','formation':'FOR-E','inventaire':'INVEN'}};
            var lower = str.toLowerCase();
            for (var key in fuzzy) {{ if (lower.indexOf(key) >= 0) return fuzzy[key]; }}
            return null;
        }}

        function parseSmartActions(text) {{
            var actions = [];
            var lines = text.split('\\n');
            lines.forEach(function(line) {{
                var trimmed = line.trim();
                var lower = trimmed.toLowerCase();
                if (!lower) return;

                // 1) Remplacement : "X remplacé par Y" / "Y remplace X"
                var rm = lower.match(/(\\w+)\\s+remplac[e\u00e9]e?\\s+par\\s+(\\w+)/i) ||
                         lower.match(/(\\w+)\\s+remplace\\s+(\\w+)/i);
                if (rm) {{
                    var n1 = _findStaff(rm[1]), n2 = _findStaff(rm[2]);
                    if (n1 && n2 && n1 !== n2) {{
                        var isPassive = lower.indexOf('remplac\u00e9') >= 0;
                        var dayIdx = _detectDay(lower);
                        var times = _detectTimes(lower);
                        actions.push({{
                            type: 'replacement',
                            out: isPassive ? n1 : n2,
                            'in': isPassive ? n2 : n1,
                            dayIdx: dayIdx,
                            start: times ? times[0] : null,
                            end: times ? times[1] : null,
                            raw: trimmed
                        }});
                        return;
                    }}
                }}

                // 2) Format multi-staff : "Jour DD/MM : Staff1 XXh-XXh en CODE, Staff2 XXh-XXh en CODE"
                //    Détecte un préfixe jour puis split par virgule
                var lineDay = _detectDay(lower);
                var colonIdx = trimmed.indexOf(':');
                var parts;
                if (lineDay !== null && colonIdx > 0 && colonIdx < 30) {{
                    // Le jour est dans le préfixe avant le ":"
                    parts = trimmed.substring(colonIdx + 1).split(',');
                }} else {{
                    parts = [trimmed];
                }}

                parts.forEach(function(part) {{
                    var partLower = part.toLowerCase().trim();
                    if (!partLower) return;
                    // Chercher un nom de staff dans ce segment
                    var foundStaff = null;
                    var partWords = partLower.replace(/[^a-z\u00e0-\u00ff\\s]/g, ' ').split(/\\s+/);
                    for (var w = 0; w < partWords.length; w++) {{
                        var s = _findStaff(partWords[w]);
                        if (s) {{ foundStaff = s; break; }}
                    }}
                    if (!foundStaff) return;
                    // Détecter horaires et code dans ce segment
                    var segTimes = _detectTimes(partLower);
                    var segDay = _detectDay(partLower);
                    if (segDay === null) segDay = lineDay; // hériter du jour de la ligne
                    var segCode = _detectCode(partLower);
                    // Ignorer les "RAS" (rien à signaler)
                    if (partLower.indexOf('ras') >= 0 && !segTimes && !segCode) return;
                    if (segTimes || segCode) {{
                        actions.push({{
                            type: 'schedule',
                            staff: foundStaff,
                            dayIdx: segDay,
                            start: segTimes ? segTimes[0] : null,
                            end: segTimes ? segTimes[1] : null,
                            code: segCode || 'VDC',
                            raw: part.trim()
                        }});
                    }}
                }});
            }});
            return actions;
        }}

        // ── Diff popup : comparaison visuelle + actions éditables ──
        function showDiffPopup(oldText, newText, smartActions, callback) {{
            var oldLines = oldText.split('\\n');
            var newLines = newText.split('\\n');
            var oldSet = {{}};
            oldLines.forEach(function(l) {{ if (l.trim()) oldSet[l.trim()] = true; }});
            var newSet = {{}};
            newLines.forEach(function(l) {{ if (l.trim()) newSet[l.trim()] = true; }});
            var diffHtml = '';
            var hasChanges = false;
            oldLines.forEach(function(line) {{
                if (!line.trim()) return;
                if (!newSet[line.trim()]) {{
                    diffHtml += '<div class="diff-line removed">\u2212 ' + escHtml(line) + '</div>';
                    hasChanges = true;
                }}
            }});
            newLines.forEach(function(line) {{
                if (!line.trim()) return;
                if (!oldSet[line.trim()]) {{
                    diffHtml += '<div class="diff-line added">+ ' + escHtml(line) + '</div>';
                    hasChanges = true;
                }}
            }});
            newLines.forEach(function(line) {{
                if (!line.trim()) return;
                if (oldSet[line.trim()]) {{
                    diffHtml += '<div class="diff-line same">&nbsp; ' + escHtml(line) + '</div>';
                }}
            }});
            if (!hasChanges && (!smartActions || smartActions.length === 0)) {{
                callback(true, []);
                return;
            }}

            // Build staff <option> list
            var staffOpts = '';
            _staffNames.forEach(function(n) {{
                staffOpts += '<option value="' + escHtml(n) + '">' + escHtml(n) + '</option>';
            }});
            // Build day <option> list
            var dayOpts = '<option value="">--</option>';
            WEEK_DATES.forEach(function(d, i) {{
                dayOpts += '<option value="' + i + '">' + _dayLabelsShort[i] + ' ' + d.split('-')[2] + '</option>';
            }});
            // Build code <option> list
            var codeOpts = '';
            _codeList.forEach(function(c) {{
                codeOpts += '<option value="' + c + '">' + c + '</option>';
            }});

            var overlay = document.createElement('div');
            overlay.className = 'diff-overlay';
            var popupEl = document.createElement('div');
            popupEl.className = 'diff-popup';

            // Title
            popupEl.innerHTML = '<h3>Modifications d\u00e9tect\u00e9es</h3>';

            // Diff section
            if (hasChanges) {{
                var diffDiv = document.createElement('div');
                diffDiv.className = 'diff-lines';
                diffDiv.innerHTML = diffHtml;
                popupEl.appendChild(diffDiv);
            }}

            // Smart actions section (editable cards)
            var actionCards = [];
            if (smartActions && smartActions.length > 0) {{
                var actDiv = document.createElement('div');
                actDiv.className = 'smart-actions';
                actDiv.innerHTML = '<div class="smart-actions-title">\U0001f504 Actions d\u00e9tect\u00e9es</div>';

                smartActions.forEach(function(a, idx) {{
                    var card = document.createElement('div');
                    card.className = 'action-card';
                    card.dataset.idx = idx;

                    // Header with checkbox + type badge
                    var hdr = document.createElement('div');
                    hdr.className = 'action-card-header';
                    var cb = document.createElement('input');
                    cb.type = 'checkbox'; cb.checked = true; cb.dataset.actionIdx = idx;
                    hdr.appendChild(cb);
                    var badge = document.createElement('span');
                    badge.className = 'action-type ' + (a.type === 'replacement' ? 'action-type-repl' : 'action-type-sched');
                    badge.textContent = a.type === 'replacement' ? 'Remplacement' : 'Horaire';
                    hdr.appendChild(badge);
                    card.appendChild(hdr);

                    // Raw text (source)
                    var rawDiv = document.createElement('div');
                    rawDiv.className = 'action-card-raw';
                    rawDiv.textContent = '\u00ab ' + a.raw + ' \u00bb';
                    card.appendChild(rawDiv);

                    // Editable fields
                    var fields = document.createElement('div');
                    fields.className = 'action-fields';

                    if (a.type === 'replacement') {{
                        // Staff out
                        fields.innerHTML +=
                            '<div class="action-field"><label>Sort</label>' +
                            '<select data-field="out">' + staffOpts.replace('value="' + escHtml(a.out) + '"', 'value="' + escHtml(a.out) + '" selected') + '</select></div>';
                        // Staff in
                        fields.innerHTML +=
                            '<div class="action-field"><label>Entre</label>' +
                            '<select data-field="in">' + staffOpts.replace('value="' + escHtml(a['in']) + '"', 'value="' + escHtml(a['in']) + '" selected') + '</select></div>';
                    }} else {{
                        // Staff
                        fields.innerHTML +=
                            '<div class="action-field"><label>Staff</label>' +
                            '<select data-field="staff">' + staffOpts.replace('value="' + escHtml(a.staff) + '"', 'value="' + escHtml(a.staff) + '" selected') + '</select></div>';
                        // Code
                        fields.innerHTML +=
                            '<div class="action-field"><label>Activit\u00e9</label>' +
                            '<select data-field="code">' + codeOpts.replace('value="' + (a.code||'VDC') + '"', 'value="' + (a.code||'VDC') + '" selected') + '</select></div>';
                    }}
                    // Day
                    var selDay = (a.dayIdx !== null && a.dayIdx !== undefined) ? a.dayIdx.toString() : '';
                    fields.innerHTML +=
                        '<div class="action-field"><label>Jour</label>' +
                        '<select data-field="dayIdx">' + dayOpts.replace('value="' + selDay + '">', 'value="' + selDay + '" selected>') + '</select></div>';
                    // Times
                    fields.innerHTML +=
                        '<div class="action-field"><label>Horaires</label>' +
                        '<div style="display:flex;gap:4px;align-items:center;">' +
                        '<input type="time" data-field="start" value="' + (a.start || '09:00') + '" style="flex:1;">' +
                        '<span style="color:#666;">-</span>' +
                        '<input type="time" data-field="end" value="' + (a.end || '17:00') + '" style="flex:1;">' +
                        '</div></div>';

                    card.appendChild(fields);
                    actDiv.appendChild(card);
                    actionCards.push(card);
                }});
                popupEl.appendChild(actDiv);
            }}

            // Action buttons
            var btns = document.createElement('div');
            btns.className = 'diff-actions';
            btns.innerHTML = '<button class="diff-btn-cancel">Annuler</button><button class="diff-btn-apply">Appliquer</button>';
            popupEl.appendChild(btns);

            overlay.appendChild(popupEl);
            document.body.appendChild(overlay);

            // Collect edited values on apply
            btns.querySelector('.diff-btn-apply').onclick = function() {{
                var selected = [];
                actionCards.forEach(function(card, idx) {{
                    var cb = card.querySelector('input[type="checkbox"]');
                    if (!cb.checked) return;
                    var a = JSON.parse(JSON.stringify(smartActions[idx]));
                    // Read edited values from form fields
                    card.querySelectorAll('[data-field]').forEach(function(el) {{
                        var field = el.dataset.field;
                        if (field === 'dayIdx') {{
                            a.dayIdx = el.value !== '' ? parseInt(el.value) : null;
                        }} else {{
                            a[field] = el.value;
                        }}
                    }});
                    selected.push(a);
                }});
                overlay.remove(); callback(true, selected);
            }};
            btns.querySelector('.diff-btn-cancel').onclick = function() {{
                overlay.remove(); callback(false, []);
            }};
            overlay.onclick = function(e) {{ if (e.target === overlay) {{ overlay.remove(); callback(false, []); }} }};
        }}
        function escHtml(s) {{
            return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
        }}

        // ── Appliquer les actions validées (remplacements + horaires) ──
        function applySmartActions(selectedActions) {{
            if (!selectedActions || selectedActions.length === 0) return;
            var data = notesWork;
            if (!data.replacements) data.replacements = [];
            var hasTimelineChanges = false;
            selectedActions.forEach(function(a) {{
                var dayIdx = (a.dayIdx !== null && a.dayIdx !== undefined) ? a.dayIdx : currentDay;
                var dateStr = WEEK_DATES[dayIdx] || WEEK_DATES[currentDay];
                if (a.type === 'replacement') {{
                    data.replacements.push({{
                        date: dateStr,
                        out: a.out,
                        'in': a['in'],
                        start: a.start || '09:00',
                        end: a.end || '17:00'
                    }});
                    hasTimelineChanges = true;
                }} else if (a.type === 'schedule') {{
                    var staff = a.staff;
                    if (DATA[staff]) {{
                        var startISO = dateStr + 'T' + (a.start || '09:00');
                        var endISO = dateStr + 'T' + (a.end || '17:00');
                        DATA[staff].events.push({{
                            code: a.code || 'VDC',
                            label: CODE_LABELS[a.code] || a.code || 'VDC',
                            start: startISO,
                            end: endISO,
                            day: dayIdx
                        }});
                        _editsDirty = true;
                        hasTimelineChanges = true;
                    }}
                }}
            }});
            if (hasTimelineChanges) {{
                updateSaveButton();
                renderTimeline();
            }}
        }}

        var STAFF_CODE = '1937';
        var STAFF_KEY = 'planning-staff-ok';
        var _p = ['Z2l0aHViX3BhdF8xMUJWTEZMVl','EwNGFQeEFvQWZzYktvX2lZOHZF','cVhqaUx1ZzNmOVQ5cUhUcUJKan','NkMWhKR2tGYXl0c28xMDJmYXRV','SFhYS1pWWks4MXZGUkpE'];
        function isStaffVerified() {{ return sessionStorage.getItem(STAFF_KEY) === '1'; }}
        function verifyStaff() {{
            if (isStaffVerified()) return true;
            var code = prompt('Code staff requis :');
            if (code && code.trim() === STAFF_CODE) {{ sessionStorage.setItem(STAFF_KEY, '1'); return true; }}
            alert('Code staff incorrect.'); return false;
        }}
        function getToken() {{
            if (!isStaffVerified()) return '';
            return localStorage.getItem(TOKEN_KEY) || atob(_p.join(''));
        }}
        function setToken(t) {{ localStorage.setItem(TOKEN_KEY, t); }}
        function isAdminUnlocked() {{ return sessionStorage.getItem(ADMIN_KEY) === '1' || isStaffVerified(); }}
        function unlockAdmin() {{ sessionStorage.setItem(ADMIN_KEY, '1'); }}
        function ensureToken() {{
            if (!verifyStaff()) return '';
            return getToken();
        }}

        function renderNotes() {{
            var data = notesWork;
            notesEl.innerHTML = '';

            // Comment card
            var card = document.createElement('div');
            card.className = 'note-card comment';
            var hdr = document.createElement('div');
            hdr.className = 'note-header';
            hdr.innerHTML = '<span class="note-label comment">Note de semaine</span>';
            var editBtn = document.createElement('button');
            editBtn.className = 'note-btn';
            editBtn.innerHTML = '\u270e';
            editBtn.title = '\u00c9diter';
            hdr.appendChild(editBtn);
            card.appendChild(hdr);
            var txt = document.createElement('div');
            txt.className = 'note-text';
            txt.textContent = data.comment || '';
            card.appendChild(txt);
            notesEl.appendChild(card);

            editBtn.onclick = function() {{
                if (txt.contentEditable === 'true') {{
                    txt.contentEditable = 'false';
                    data.comment = txt.innerText;
                    editBtn.innerHTML = '\u270e';
                    notesDirty = true; saveNotesLocal();
                    renderNotes();
                }} else {{
                    txt.contentEditable = 'true';
                    txt.focus();
                    editBtn.innerHTML = '\u2714';
                }}
            }};

            // ── Collage intelligent : détecter les changements et demander confirmation ──
            txt.addEventListener('paste', function(e) {{
                if (txt.contentEditable !== 'true') return;
                var pasted = (e.clipboardData || window.clipboardData).getData('text');
                if (!pasted || !pasted.trim()) return;
                var oldText = data.comment || '';
                var newText = pasted.trim();
                if (oldText.trim() === newText) return;
                e.preventDefault();
                var smartActions = parseSmartActions(newText);
                showDiffPopup(oldText, newText, smartActions, function(accepted, selectedActions) {{
                    if (accepted) {{
                        data.comment = newText;
                        applySmartActions(selectedActions);
                        notesDirty = true; saveNotesLocal();
                        renderNotes();
                    }}
                }});
            }});

            // Update cards
            data.updates.forEach(function(u, idx) {{
                var ucard = document.createElement('div');
                ucard.className = 'note-card update';
                var uhdr = document.createElement('div');
                uhdr.className = 'note-header';
                var dateLabel = '';
                if (u.date) {{
                    var _dp = u.date.split(/[\\-T :]/);
                    var _dd = new Date(parseInt(_dp[0]), parseInt(_dp[1])-1, parseInt(_dp[2]),
                        _dp.length > 3 ? parseInt(_dp[3]) : 0, _dp.length > 4 ? parseInt(_dp[4]) : 0);
                    var _jours = ['Dimanche','Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'];
                    var _mois = ['janvier','f\u00e9vrier','mars','avril','mai','juin','juillet','ao\u00fbt','septembre','octobre','novembre','d\u00e9cembre'];
                    var _timePart = (_dp.length > 3) ? ' \u00e0 ' + _dp[3] + 'h' + (_dp[4] || '00') : '';
                    dateLabel = ' \u2014 ' + _jours[_dd.getDay()] + ' ' + _dd.getDate() + ' ' + _mois[_dd.getMonth()] + _timePart;
                }}
                uhdr.innerHTML = '<span class="note-label update">Mise \u00e0 jour' + dateLabel + '</span>';
                var uactions = document.createElement('div');
                uactions.className = 'note-actions';
                var uedit = document.createElement('button');
                uedit.className = 'note-btn';
                uedit.innerHTML = '\u270e';
                uedit.title = '\u00c9diter';
                var udel = document.createElement('button');
                udel.className = 'note-btn del';
                udel.innerHTML = '\u2716';
                udel.title = 'Supprimer';
                uactions.appendChild(uedit);
                uactions.appendChild(udel);
                uhdr.appendChild(uactions);
                ucard.appendChild(uhdr);
                var utxt = document.createElement('div');
                utxt.className = 'note-text';
                utxt.textContent = u.text || '';
                ucard.appendChild(utxt);
                notesEl.appendChild(ucard);

                uedit.onclick = function() {{
                    if (utxt.contentEditable === 'true') {{
                        utxt.contentEditable = 'false';
                        data.updates[idx].text = utxt.innerText;
                        uedit.innerHTML = '\u270e';
                        notesDirty = true; saveNotesLocal();
                        renderNotes();
                    }} else {{
                        utxt.contentEditable = 'true';
                        utxt.focus();
                        uedit.innerHTML = '\u2714';
                    }}
                }};

                // ── Collage intelligent sur les mises à jour ──
                utxt.addEventListener('paste', function(e) {{
                    if (utxt.contentEditable !== 'true') return;
                    var pasted = (e.clipboardData || window.clipboardData).getData('text');
                    if (!pasted || !pasted.trim()) return;
                    var smartActions = parseSmartActions(pasted.trim());
                    if (smartActions.length === 0) return; // pas d'action détectée → collage normal
                    e.preventDefault();
                    var oldText = data.updates[idx].text || '';
                    showDiffPopup(oldText, pasted.trim(), smartActions, function(accepted, selectedActions) {{
                        if (accepted) {{
                            data.updates[idx].text = pasted.trim();
                            applySmartActions(selectedActions);
                            notesDirty = true; saveNotesLocal();
                            renderNotes();
                        }}
                    }});
                }});

                udel.onclick = function() {{
                    data.updates.splice(idx, 1);
                    notesDirty = true; saveNotesLocal();
                    renderNotes();
                }};
            }});

            // ── Replacement cards ──
            var repls = data.replacements || [];
            repls.forEach(function(r, idx) {{
                var rcard = document.createElement('div');
                rcard.className = 'note-card replacement';
                var rhdr = document.createElement('div');
                rhdr.className = 'note-header';
                var rDateLabel = '';
                if (r.date) {{
                    var _rp = r.date.split('-');
                    var _rd = new Date(parseInt(_rp[0]), parseInt(_rp[1])-1, parseInt(_rp[2]));
                    var _rjours = ['Dim','Lun','Mar','Mer','Jeu','Ven','Sam'];
                    rDateLabel = ' \u2014 ' + _rjours[_rd.getDay()] + ' ' + _rd.getDate() + '/' + _rp[1];
                }}
                rhdr.innerHTML = '<span class="note-label replacement">Remplacement' + rDateLabel + '</span>';
                var ractions = document.createElement('div');
                ractions.className = 'note-actions';
                var rdel = document.createElement('button');
                rdel.className = 'note-btn del';
                rdel.innerHTML = '\u2716';
                rdel.title = 'Supprimer';
                ractions.appendChild(rdel);
                rhdr.appendChild(ractions);
                rcard.appendChild(rhdr);
                var rsummary = document.createElement('div');
                rsummary.className = 'repl-summary';
                rsummary.innerHTML = '<span class="repl-out">' + getFirstName(r.out) + '</span> \u2192 <span class="repl-in">' + getFirstName(r.in) + '</span>' +
                    '  <span style="color:#666;font-size:10px">' + (r.start || '') + ' \u2013 ' + (r.end || '') + '</span>';
                rcard.appendChild(rsummary);
                notesEl.appendChild(rcard);

                rdel.onclick = function() {{
                    data.replacements.splice(idx, 1);
                    notesDirty = true; saveNotesLocal();
                    renderNotes();
                    renderTimeline();
                }};
            }});

            // Add replacement button
            var addReplBtn = document.createElement('button');
            addReplBtn.className = 'add-repl-btn';
            addReplBtn.textContent = '+ Ajouter un remplacement';
            addReplBtn.onclick = function() {{
                // Build employee list from DATA
                var names = Object.keys(DATA).filter(function(n) {{ return n !== '_codeNames'; }}).sort();
                var form = document.createElement('div');
                form.className = 'note-card replacement';
                form.innerHTML = '<div class="note-header"><span class="note-label replacement">Nouveau remplacement</span></div>';
                var formBody = document.createElement('div');
                formBody.className = 'repl-form';

                // Date row
                var dateRow = document.createElement('div');
                dateRow.className = 'repl-row';
                dateRow.innerHTML = '<label>Jour</label>';
                var dateSel = document.createElement('select');
                WEEK_DATES.forEach(function(d, i) {{
                    var opt = document.createElement('option');
                    opt.value = d;
                    var _jrs = ['Dim','Lun','Mar','Mer','Jeu','Ven','Sam'];
                    var _dp = d.split('-');
                    var _dt = new Date(parseInt(_dp[0]), parseInt(_dp[1])-1, parseInt(_dp[2]));
                    opt.textContent = _jrs[_dt.getDay()] + ' ' + _dt.getDate() + '/' + _dp[1];
                    if (i === currentDay) opt.selected = true;
                    dateSel.appendChild(opt);
                }});
                dateRow.appendChild(dateSel);
                formBody.appendChild(dateRow);

                // "Out" row – only people working on the selected day
                var outRow = document.createElement('div');
                outRow.className = 'repl-row';
                outRow.innerHTML = '<label>Sort</label>';
                var outSel = document.createElement('select');
                function refreshOutOptions() {{
                    var prev = outSel.value;
                    outSel.innerHTML = '';
                    var outDef = document.createElement('option');
                    outDef.value = ''; outDef.textContent = 'Personne remplac\u00e9e...';
                    outSel.appendChild(outDef);
                    var dayIdx = WEEK_DATES.indexOf(dateSel.value);
                    names.forEach(function(n) {{
                        if (dayIdx < 0) return;
                        var emp = DATA[n];
                        if (!emp || !emp.events) return;
                        var works = emp.events.some(function(ev) {{ return ev.day === dayIdx; }});
                        if (!works) return;
                        var opt = document.createElement('option');
                        opt.value = n; opt.textContent = n;
                        if (n === prev) opt.selected = true;
                        outSel.appendChild(opt);
                    }});
                }}
                refreshOutOptions();
                dateSel.addEventListener('change', refreshOutOptions);
                outRow.appendChild(outSel);
                formBody.appendChild(outRow);

                // "In" row
                var inRow = document.createElement('div');
                inRow.className = 'repl-row';
                inRow.innerHTML = '<label>Entre</label>';
                var inSel = document.createElement('select');
                var inDef = document.createElement('option');
                inDef.value = ''; inDef.textContent = 'Rempla\u00e7ant(e)...';
                inSel.appendChild(inDef);
                names.forEach(function(n) {{
                    var opt = document.createElement('option');
                    opt.value = n; opt.textContent = n;
                    inSel.appendChild(opt);
                }});
                inRow.appendChild(inSel);
                formBody.appendChild(inRow);

                // Time range row
                var timeRow = document.createElement('div');
                timeRow.className = 'repl-row';
                timeRow.innerHTML = '<label>Cr\u00e9neau</label>';
                var startInput = document.createElement('input');
                startInput.type = 'time'; startInput.value = '18:00';
                startInput.style.flex = '1';
                var sep = document.createElement('span');
                sep.textContent = ' \u2192 '; sep.style.color = '#666';
                var endInput = document.createElement('input');
                endInput.type = 'time'; endInput.value = '23:00';
                endInput.style.flex = '1';
                timeRow.appendChild(startInput);
                timeRow.appendChild(sep);
                timeRow.appendChild(endInput);
                formBody.appendChild(timeRow);

                // Auto-fill time range from replaced person's schedule
                function autoFillTimes() {{
                    var selName = outSel.value;
                    var selDate = dateSel.value;
                    if (!selName || !selDate || !DATA[selName]) return;
                    var emp = DATA[selName];
                    var dayIdx = WEEK_DATES.indexOf(selDate);
                    if (dayIdx < 0) return;
                    // Find earliest start and latest end for this person on this day
                    var earliest = null, latest = null;
                    emp.events.forEach(function(ev) {{
                        if (ev.day !== dayIdx) return;
                        var s = new Date(ev.start);
                        var e = new Date(ev.end);
                        var sStr = s.getHours().toString().padStart(2,'0') + ':' + s.getMinutes().toString().padStart(2,'0');
                        var eStr = e.getHours().toString().padStart(2,'0') + ':' + e.getMinutes().toString().padStart(2,'0');
                        if (eStr === '00:00') eStr = '23:59';
                        if (!earliest || sStr < earliest) earliest = sStr;
                        if (!latest || eStr > latest) latest = eStr;
                    }});
                    if (earliest && latest) {{
                        startInput.value = earliest;
                        endInput.value = latest;
                    }}
                }}
                outSel.addEventListener('change', autoFillTimes);
                dateSel.addEventListener('change', autoFillTimes);

                // Submit
                var submitBtn = document.createElement('button');
                submitBtn.className = 'repl-add-btn';
                submitBtn.textContent = 'Valider';
                submitBtn.onclick = function() {{
                    if (!outSel.value || !inSel.value) return;
                    if (!data.replacements) data.replacements = [];
                    data.replacements.push({{
                        date: dateSel.value,
                        out: outSel.value,
                        in: inSel.value,
                        start: startInput.value,
                        end: endInput.value
                    }});
                    notesDirty = true; saveNotesLocal();
                    renderNotes();
                    renderTimeline();
                }};
                formBody.appendChild(submitBtn);
                form.appendChild(formBody);

                // Replace button with form
                addReplBtn.replaceWith(form);
            }};
            notesEl.appendChild(addReplBtn);

            // Add update button
            var addBtn = document.createElement('button');
            addBtn.className = 'add-note-btn';
            addBtn.textContent = '+ Ajouter une mise \u00e0 jour';
            addBtn.onclick = function() {{
                var today = new Date();
                var ds = today.getFullYear() + '-' +
                    (today.getMonth()+1).toString().padStart(2,'0') + '-' +
                    today.getDate().toString().padStart(2,'0') + ' ' +
                    today.getHours().toString().padStart(2,'0') + ':' +
                    today.getMinutes().toString().padStart(2,'0');
                data.updates.push({{ date: ds, text: '' }});
                notesDirty = true; saveNotesLocal();
                renderNotes();
                var cards = notesEl.querySelectorAll('.note-card.update .note-text');
                if (cards.length > 0) {{
                    var last = cards[cards.length - 1];
                    last.contentEditable = 'true';
                    last.focus();
                    var editBtns = notesEl.querySelectorAll('.note-card.update .note-btn:not(.del)');
                    if (editBtns.length > 0) editBtns[editBtns.length - 1].innerHTML = '\u2714';
                }}
            }};
            notesEl.appendChild(addBtn);

            // Publish button (only if admin token is set and notes changed)
            var token = getToken();
            if (token && notesDirty) {{
                var pubBtn = document.createElement('button');
                pubBtn.className = 'publish-btn';
                pubBtn.textContent = 'Publier les notes';
                pubBtn.onclick = function() {{
                    pubBtn.disabled = true;
                    pubBtn.textContent = 'Publication en cours...';
                    pushNotesToGitHub(data, pubBtn);
                }};
                notesEl.appendChild(pubBtn);
            }}
        }}

        function pushNotesToGitHub(data, btn) {{
            var token = ensureToken();
            if (!token) {{ btn.textContent = 'Token requis pour publier'; btn.disabled = false; return; }}
            var content = btoa(unescape(encodeURIComponent(JSON.stringify(data, null, 2) + '\\n')));
            var apiUrl = 'https://api.github.com/repos/' + REPO + '/contents/' + NOTES_PATH;

            // First get the current file SHA (required for update)
            fetch(apiUrl, {{
                headers: {{ 'Authorization': 'Bearer ' + token, 'Accept': 'application/vnd.github.v3+json' }}
            }})
            .then(function(r) {{ return r.ok ? r.json() : {{ sha: null }}; }})
            .then(function(file) {{
                var body = {{
                    message: 'MAJ notes S{week_num} depuis la page',
                    content: content,
                    branch: 'main'
                }};
                if (file.sha) body.sha = file.sha;

                return fetch(apiUrl, {{
                    method: 'PUT',
                    headers: {{
                        'Authorization': 'Bearer ' + token,
                        'Accept': 'application/vnd.github.v3+json',
                        'Content-Type': 'application/json'
                    }},
                    body: JSON.stringify(body)
                }});
            }})
            .then(function(r) {{
                if (r.ok) {{
                    notesDirty = false;
                    updateUnsavedBanner();
                    showRefreshCountdown(btn);
                }} else {{
                    return r.json().then(function(err) {{
                        btn.disabled = false;
                        btn.textContent = 'Erreur : ' + (err.message || 'v\u00e9rifier le token');
                        btn.classList.remove('success');
                    }});
                }}
            }})
            .catch(function(e) {{
                btn.disabled = false;
                btn.textContent = 'Erreur r\u00e9seau, r\u00e9essayer';
            }});
        }}

        function showRefreshCountdown(btn) {{
            var seconds = 90;
            btn.classList.add('success');
            btn.disabled = true;

            function tick() {{
                if (seconds > 0) {{
                    btn.textContent = 'Publi\u00e9 \u2714 En ligne dans ~' + seconds + 's \u2014 Rafra\u00eechir';
                    seconds--;
                    setTimeout(tick, 1000);
                }} else {{
                    btn.textContent = "C'est en ligne ! Rafra\u00eechir la page";
                }}
                btn.disabled = false;
                btn.onclick = function() {{ location.reload(); }};
            }}
            tick();
        }}

        renderNotes();

        // Initial render
        renderTimeline();

        // ── Admin edit mode ──
        var editMode = false;
        var adminToolbarEl = null;

        function initAdminToolbar() {{
            if (!isAdminUnlocked()) return;
            if (adminToolbarEl) return;
            adminToolbarEl = document.createElement('div');
            adminToolbarEl.className = 'admin-toolbar';
            adminToolbarEl.innerHTML = '';
            var toggleBtn = document.createElement('button');
            toggleBtn.className = 'edit-toggle';
            toggleBtn.textContent = 'Mode \u00e9dition';
            toggleBtn.onclick = function() {{
                if (editMode && _editsDirty) {{
                    showConfirmDialog(
                        'Vous avez des modifications non enregistrées.<br>Que souhaitez-vous faire ?',
                        function() {{ publishAllEdits(); }},
                        function() {{
                            _editsDirty = false;
                            updateSaveButton();
                            updateUnsavedBanner();
                            editMode = false;
                            toggleBtn.classList.remove('active');
                            toggleBtn.textContent = 'Mode édition';
                            renderTimeline();
                        }}
                    );
                    return;
                }}
                editMode = !editMode;
                toggleBtn.classList.toggle('active', editMode);
                toggleBtn.textContent = editMode ? 'Quitter édition' : 'Mode édition';
                renderTimeline();
            }};
            adminToolbarEl.appendChild(toggleBtn);
            var saveBtn = document.createElement('button');
            saveBtn.className = 'save-edits-btn';
            saveBtn.id = 'save-edits-btn';
            saveBtn.textContent = 'Enregistrer';
            saveBtn.style.display = 'none';
            saveBtn.onclick = function() {{ publishAllEdits(); }};
            adminToolbarEl.appendChild(saveBtn);
            var statusEl = document.createElement('span');
            statusEl.className = 'edit-status';
            statusEl.id = 'edit-status';
            adminToolbarEl.appendChild(statusEl);
            var viewDay = document.getElementById('view-day');
            viewDay.insertBefore(adminToolbarEl, viewDay.firstChild);
        }}

        // Override renderTimeline to add editable class when editMode
        var _origRenderTimeline = renderTimeline;
        var _dragState = null;

        // Snap a decimal hour to the nearest 15-min increment
        function snapHour(h) {{ return Math.round(h * 4) / 4; }}
        function hourToStr(h) {{
            var hh = Math.floor(h); var mm = Math.round((h - hh) * 60);
            if (mm === 60) {{ hh++; mm = 0; }}
            return hh.toString().padStart(2,'0') + ':' + mm.toString().padStart(2,'0');
        }}

        function startDrag(e, bar, side, empName, ev, container) {{
            e.preventDefault(); e.stopPropagation();
            var rect = container.getBoundingClientRect();
            // Compute current timeline range from rendered grid
            var tl = document.getElementById('timeline');
            var inner = tl.querySelector('.timeline-inner');
            // Read minH / range from container data attributes
            var minH = parseFloat(container.dataset.minH);
            var range = parseFloat(container.dataset.range);

            _dragState = {{ bar: bar, side: side, empName: empName, ev: ev,
                            rect: rect, minH: minH, range: range }};
            bar.querySelector('.drag-handle.' + side).classList.add('active');

            // Create tooltip
            var tip = document.createElement('div');
            tip.className = 'drag-tooltip';
            tip.id = 'drag-tooltip';
            document.body.appendChild(tip);
            updateDrag(e);
        }}

        function updateDrag(e) {{
            if (!_dragState) return;
            var clientX = e.touches ? e.touches[0].clientX : e.clientX;
            var clientY = e.touches ? e.touches[0].clientY : e.clientY;
            var ds = _dragState;
            var pct = ((clientX - ds.rect.left) / ds.rect.width) * 100;
            pct = Math.max(0, Math.min(100, pct));
            var newH = snapHour(ds.minH + (pct / 100) * ds.range);

            var s = new Date(ds.ev.start);
            var eDate = new Date(ds.ev.end);
            var sh = s.getHours() + s.getMinutes() / 60;
            var eh = eDate.getHours() + eDate.getMinutes() / 60;
            if (eh <= sh) eh = 24;

            if (ds.side === 'left') {{
                newH = Math.min(newH, eh - 0.25);  // min 15 min
                var newLeft = ((newH - ds.minH) / ds.range) * 100;
                var newWidth = ((eh - newH) / ds.range) * 100;
                ds.bar.style.left = newLeft + '%';
                ds.bar.style.width = newWidth + '%';
                ds.currentH = newH;
            }} else {{
                newH = Math.max(newH, sh + 0.25);
                var origLeft = ((sh - ds.minH) / ds.range) * 100;
                var newWidth = ((newH - sh) / ds.range) * 100;
                ds.bar.style.width = newWidth + '%';
                ds.currentH = newH;
            }}

            var tip = document.getElementById('drag-tooltip');
            if (tip) {{
                tip.textContent = hourToStr(newH);
                tip.style.left = (clientX + 12) + 'px';
                tip.style.top = (clientY - 28) + 'px';
            }}
        }}

        function endDrag() {{
            if (!_dragState) return;
            var ds = _dragState;
            var s = new Date(ds.ev.start);
            var eDate = new Date(ds.ev.end);
            var sh = s.getHours() + s.getMinutes() / 60;
            var eh = eDate.getHours() + eDate.getMinutes() / 60;
            if (eh <= sh) eh = 24;

            var newStart, newEnd;
            if (ds.side === 'left') {{
                newStart = hourToStr(ds.currentH);
                newEnd = hourToStr(eh);
            }} else {{
                newStart = hourToStr(sh);
                newEnd = hourToStr(ds.currentH);
            }}

            _dragState = null;
            var tip = document.getElementById('drag-tooltip');
            if (tip) tip.remove();

            applyTimeEdit(ds.empName, ds.ev, newStart, newEnd);
        }}

        document.addEventListener('mousemove', function(e) {{ updateDrag(e); }});
        document.addEventListener('mouseup', function() {{ endDrag(); }});
        document.addEventListener('touchmove', function(e) {{ if (_dragState) {{ e.preventDefault(); updateDrag(e); }} }}, {{ passive: false }});
        document.addEventListener('touchend', function() {{ endDrag(); }});

        renderTimeline = function() {{
            _origRenderTimeline();
            if (!editMode) return;

            document.querySelectorAll('#timeline .tl-bar').forEach(function(bar) {{
                bar.classList.add('editable');
                // Add drag handles
                if (!bar.querySelector('.drag-handle')) {{
                    var handleL = document.createElement('div');
                    handleL.className = 'drag-handle left';
                    var handleR = document.createElement('div');
                    handleR.className = 'drag-handle right';
                    bar.appendChild(handleL);
                    bar.appendChild(handleR);
                }}
            }});

            // Add click + drag handlers + delete staff buttons
            var rows = document.querySelectorAll('#timeline .timeline-row');
            rows.forEach(function(row) {{
                var nameEl = row.querySelector('.tl-name');
                if (!nameEl || !nameEl.title) return;
                var empName = nameEl.title;
                var container = row.querySelector('.tl-bar-container');

                // Add delete staff button next to name
                if (!nameEl.querySelector('.del-staff')) {{
                    var delBtn = document.createElement('span');
                    delBtn.className = 'del-staff';
                    delBtn.textContent = '\u00d7';
                    delBtn.title = 'Supprimer ' + empName;
                    delBtn.onclick = function(e) {{ e.stopPropagation(); deleteStaff(empName); }};
                    nameEl.appendChild(delBtn);
                }}
                row.querySelectorAll('.tl-bar').forEach(function(bar, idx) {{
                    var emp = DATA[empName];
                    if (!emp) return;
                    var dayEvts = emp.events.filter(function(ev) {{ return ev.day === currentDay; }});
                    if (!dayEvts[idx]) return;
                    var ev = dayEvts[idx];

                    bar.onclick = function(e) {{
                        if (!editMode || _dragState) return;
                        // Don't open popup if click was on a drag handle
                        if (e.target.classList.contains('drag-handle')) return;
                        e.stopPropagation();
                        openEditPopup(empName, ev, idx);
                    }};

                    bar.querySelector('.drag-handle.left').onmousedown = function(e) {{ startDrag(e, bar, 'left', empName, ev, container); }};
                    bar.querySelector('.drag-handle.right').onmousedown = function(e) {{ startDrag(e, bar, 'right', empName, ev, container); }};
                    bar.querySelector('.drag-handle.left').ontouchstart = function(e) {{ startDrag(e, bar, 'left', empName, ev, container); }};
                    bar.querySelector('.drag-handle.right').ontouchstart = function(e) {{ startDrag(e, bar, 'right', empName, ev, container); }};
                }});

                // Click on empty area of bar container → create new event
                (function(cont, eName) {{
                    cont.addEventListener('click', function(e) {{
                        if (!editMode || _dragState) return;
                        if (e.target !== cont && !e.target.classList.contains('tl-grid-line')) return;
                        var rect = cont.getBoundingClientRect();
                        var minH = parseFloat(cont.dataset.minH);
                        var range = parseFloat(cont.dataset.range);
                        var pct = ((e.clientX - rect.left) / rect.width) * 100;
                        var clickH = snapHour(minH + (pct / 100) * range);
                        openAddEventPopup(eName, clickH);
                    }});
                }})(container, empName);
            }});

            // Add staff button at bottom of timeline
            var addStaffRow = document.getElementById('add-staff-row');
            if (addStaffRow) addStaffRow.remove();
            var tl = document.getElementById('timeline');
            var addRow = document.createElement('div');
            addRow.className = 'timeline-row add-staff-row';
            addRow.id = 'add-staff-row';
            addRow.innerHTML = '<button class="add-staff-btn">+ Ajouter un employ\u00e9</button>';
            addRow.querySelector('button').onclick = function() {{ openAddStaffPopup(); }};
            tl.querySelector('.timeline-inner').appendChild(addRow);
        }};

        // Build activity code options for select
        function buildCodeOptions() {{
            var codes = {{}};
            Object.keys(DATA).forEach(function(n) {{
                if (n === '_codeNames' || n === '_meta') return;
                DATA[n].events.forEach(function(ev) {{
                    if (!codes[ev.code]) codes[ev.code] = ev.label || ev.code;
                }});
            }});
            if (DATA._codeNames) {{
                Object.keys(DATA._codeNames).forEach(function(c) {{
                    if (!codes[c]) codes[c] = DATA._codeNames[c];
                }});
            }}
            return codes;
        }}

        function openEditPopup(empName, ev, evIdx) {{
            // Remove existing popup
            var old = document.getElementById('edit-overlay');
            if (old) old.remove();
            old = document.getElementById('edit-popup');
            if (old) old.remove();

            var s = new Date(ev.start);
            var e = new Date(ev.end);
            var sh = s.getHours().toString().padStart(2, '0') + ':' + s.getMinutes().toString().padStart(2, '0');
            var eh = e.getHours().toString().padStart(2, '0') + ':' + e.getMinutes().toString().padStart(2, '0');

            // Build code selector
            var codes = buildCodeOptions();
            var codeOpts = '';
            Object.keys(codes).sort().forEach(function(c) {{
                var sel = (c === ev.code) ? ' selected' : '';
                codeOpts += '<option value="' + c + '"' + sel + '>' + c + ' \u2014 ' + codes[c] + '</option>';
            }});

            var overlay = document.createElement('div');
            overlay.className = 'edit-overlay';
            overlay.id = 'edit-overlay';
            overlay.onclick = function() {{ closeEditPopup(); }};
            document.body.appendChild(overlay);

            var popup = document.createElement('div');
            popup.className = 'edit-popup';
            popup.id = 'edit-popup';
            popup.innerHTML =
                '<h3>' + getFirstName(empName) + ' \u2014 ' + ev.label + '</h3>' +
                '<div class="field"><label>Activit\u00e9</label><select id="edit-code">' + codeOpts + '</select></div>' +
                '<div class="field"><label>D\u00e9but</label><input type="time" id="edit-start" value="' + sh + '"></div>' +
                '<div class="field"><label>Fin</label><input type="time" id="edit-end" value="' + eh + '"></div>' +
                '<div class="actions">' +
                '<button class="btn-delete" id="edit-delete">Supprimer</button>' +
                '<button class="btn-cancel" id="edit-cancel">Annuler</button>' +
                '<button class="btn-save" id="edit-save">Enregistrer</button>' +
                '</div>';
            document.body.appendChild(popup);

            document.getElementById('edit-cancel').onclick = closeEditPopup;
            document.getElementById('edit-delete').onclick = function() {{
                if (!confirm('Supprimer ce cr\u00e9neau ?')) return;
                deleteEvent(empName, ev);
                closeEditPopup();
            }};
            document.getElementById('edit-save').onclick = function() {{
                var newStart = document.getElementById('edit-start').value;
                var newEnd = document.getElementById('edit-end').value;
                var newCode = document.getElementById('edit-code').value;
                if (!newStart || !newEnd) return;
                var codes = buildCodeOptions();
                ev.code = newCode;
                ev.label = codes[newCode] || newCode;
                applyTimeEdit(empName, ev, newStart, newEnd);
                closeEditPopup();
            }};
        }}

        function deleteEvent(empName, ev) {{
            var emp = DATA[empName];
            if (!emp) return;
            var idx = emp.events.indexOf(ev);
            if (idx !== -1) emp.events.splice(idx, 1);
            renderTimeline();
            updateHoursBadges();
            pushDataAfterEdit();
        }}

        function openAddEventPopup(empName, defaultHour) {{
            closeEditPopup();
            var dateStr = WEEK_DATES[currentDay];
            var startH = hourToStr(defaultHour);
            var endH = hourToStr(defaultHour + 1);

            var codes = buildCodeOptions();
            var codeOpts = '';
            Object.keys(codes).sort().forEach(function(c) {{
                codeOpts += '<option value="' + c + '">' + c + ' \u2014 ' + codes[c] + '</option>';
            }});

            var overlay = document.createElement('div');
            overlay.className = 'edit-overlay';
            overlay.id = 'edit-overlay';
            overlay.onclick = function() {{ closeEditPopup(); }};
            document.body.appendChild(overlay);

            var popup = document.createElement('div');
            popup.className = 'edit-popup';
            popup.id = 'edit-popup';
            popup.innerHTML =
                '<h3>Nouveau cr\u00e9neau \u2014 ' + getFirstName(empName) + '</h3>' +
                '<div class="field"><label>Activit\u00e9</label><select id="edit-code">' + codeOpts + '</select></div>' +
                '<div class="field"><label>D\u00e9but</label><input type="time" id="edit-start" value="' + startH + '"></div>' +
                '<div class="field"><label>Fin</label><input type="time" id="edit-end" value="' + endH + '"></div>' +
                '<div class="actions">' +
                '<button class="btn-cancel" id="edit-cancel">Annuler</button>' +
                '<button class="btn-save" id="edit-save">Ajouter</button>' +
                '</div>';
            document.body.appendChild(popup);

            document.getElementById('edit-cancel').onclick = closeEditPopup;
            document.getElementById('edit-save').onclick = function() {{
                var newStart = document.getElementById('edit-start').value;
                var newEnd = document.getElementById('edit-end').value;
                var newCode = document.getElementById('edit-code').value;
                if (!newStart || !newEnd) return;
                var codes = buildCodeOptions();
                var newEv = {{
                    code: newCode,
                    label: codes[newCode] || newCode,
                    start: dateStr + 'T' + newStart,
                    end: dateStr + 'T' + newEnd,
                    day: currentDay
                }};
                DATA[empName].events.push(newEv);
                renderTimeline();
                updateHoursBadges();
                closeEditPopup();
                pushDataAfterEdit();
            }};
        }}

        function openAddStaffPopup() {{
            closeEditPopup();
            // Find staff who have NO events on the current day (repos)
            var reposStaff = [];
            Object.keys(DATA).forEach(function(name) {{
                if (name === '_codeNames' || name === '_meta') return;
                var hasEventToday = DATA[name].events.some(function(ev) {{ return ev.day === currentDay; }});
                if (!hasEventToday) reposStaff.push(name);
            }});

            if (reposStaff.length === 0) {{
                alert('Tous les employ\u00e9s travaillent d\u00e9j\u00e0 ce jour.');
                return;
            }}

            reposStaff.sort();

            var overlay = document.createElement('div');
            overlay.className = 'edit-overlay';
            overlay.id = 'edit-overlay';
            overlay.onclick = function() {{ closeEditPopup(); }};
            document.body.appendChild(overlay);

            var popup = document.createElement('div');
            popup.className = 'edit-popup';
            popup.id = 'edit-popup';

            var listHtml = '<h3>Ajouter un employ\u00e9 en repos</h3>';
            listHtml += '<p style="font-size:11px;color:#888;margin-bottom:10px;">S\u00e9lectionnez l\\'employ\u00e9 \u00e0 ajouter pour ' + DAYS_FULL[currentDay] + ' :</p>';
            listHtml += '<div style="max-height:250px;overflow-y:auto;">';
            reposStaff.forEach(function(name, idx) {{
                listHtml += '<button class="btn-save" id="pick-staff-' + idx + '" style="display:block;width:100%;margin-bottom:6px;text-align:left;padding:8px 12px;font-size:13px;">' + name + '</button>';
            }});
            listHtml += '</div>';
            listHtml += '<div class="actions" style="margin-top:12px;">';
            listHtml += '<button class="btn-cancel" id="edit-cancel">Annuler</button>';
            listHtml += '</div>';
            popup.innerHTML = listHtml;
            document.body.appendChild(popup);

            document.getElementById('edit-cancel').onclick = closeEditPopup;
            reposStaff.forEach(function(name, idx) {{
                document.getElementById('pick-staff-' + idx).onclick = function() {{
                    closeEditPopup();
                    openAddEventForReposStaff(name);
                }};
            }});
        }}

        function openAddEventForReposStaff(empName) {{
            var dateStr = WEEK_DATES[currentDay];
            var codes = buildCodeOptions();
            var codeOpts = '';
            Object.keys(codes).sort().forEach(function(c) {{
                codeOpts += '<option value="' + c + '">' + c + ' \u2014 ' + codes[c] + '</option>';
            }});

            var overlay = document.createElement('div');
            overlay.className = 'edit-overlay';
            overlay.id = 'edit-overlay';
            overlay.onclick = function() {{ closeEditPopup(); }};
            document.body.appendChild(overlay);

            var popup = document.createElement('div');
            popup.className = 'edit-popup';
            popup.id = 'edit-popup';
            popup.innerHTML =
                '<h3>Cr\u00e9neau \u2014 ' + getFirstName(empName) + '</h3>' +
                '<div class="field"><label>Activit\u00e9</label><select id="edit-code">' + codeOpts + '</select></div>' +
                '<div class="field"><label>D\u00e9but</label><input type="time" id="edit-start" value="09:00"></div>' +
                '<div class="field"><label>Fin</label><input type="time" id="edit-end" value="17:00"></div>' +
                '<div class="actions">' +
                '<button class="btn-cancel" id="edit-cancel">Annuler</button>' +
                '<button class="btn-save" id="edit-save">Ajouter</button>' +
                '</div>';
            document.body.appendChild(popup);

            document.getElementById('edit-cancel').onclick = closeEditPopup;
            document.getElementById('edit-save').onclick = function() {{
                var newStart = document.getElementById('edit-start').value;
                var newEnd = document.getElementById('edit-end').value;
                var newCode = document.getElementById('edit-code').value;
                if (!newStart || !newEnd) return;
                var codes = buildCodeOptions();
                var newEv = {{
                    code: newCode,
                    label: codes[newCode] || newCode,
                    start: dateStr + 'T' + newStart,
                    end: dateStr + 'T' + newEnd,
                    day: currentDay
                }};
                DATA[empName].events.push(newEv);
                renderTimeline();
                updateHoursBadges();
                closeEditPopup();
                pushDataAfterEdit();
            }};
        }}

        function deleteStaff(empName) {{
            var emp = DATA[empName];
            if (!emp) return;
            closeEditPopup();

            // Find which days this employee has events
            var daySet = {{}};
            emp.events.forEach(function(ev) {{ daySet[ev.day] = true; }});
            var daysWithEvents = Object.keys(daySet).map(Number).sort(function(a,b){{ return a-b; }});

            // If no events, just delete the empty entry
            if (daysWithEvents.length === 0) {{
                if (!confirm('Supprimer ' + empName + ' (aucun cr\u00e9neau) ?')) return;
                delete DATA[empName];
                renderTimeline();
                updateHoursBadges();
                pushDataAfterEdit();
                return;
            }}

            var overlay = document.createElement('div');
            overlay.className = 'edit-overlay';
            overlay.id = 'edit-overlay';
            overlay.onclick = function() {{ closeEditPopup(); }};
            document.body.appendChild(overlay);

            var popup = document.createElement('div');
            popup.className = 'edit-popup';
            popup.id = 'edit-popup';

            var html = '<h3>Supprimer ' + getFirstName(empName) + '</h3>';
            html += '<p style="font-size:11px;color:#888;margin-bottom:10px;">S\u00e9lectionner les jours \u00e0 supprimer :</p>';

            // "Select all" checkbox
            html += '<label class="day-check all-check"><input type="checkbox" id="del-all"> <span>Toute la semaine</span></label>';

            // One checkbox per day
            for (var i = 0; i < 7; i++) {{
                var hasEvts = daySet[i];
                var count = emp.events.filter(function(ev) {{ return ev.day === i; }}).length;
                var label = DAYS_FULL[i] + ' ' + (WEEK_DATES[i] || '').substring(8,10) + '/' + (WEEK_DATES[i] || '').substring(5,7);
                if (hasEvts) {{
                    html += '<label class="day-check"><input type="checkbox" value="' + i + '" class="del-day-cb"' +
                        (daysWithEvents.length === 1 ? ' checked' : '') +
                        '> <span>' + label + ' <em>(' + count + ' cr\u00e9neau' + (count > 1 ? 'x' : '') + ')</em></span></label>';
                }}
            }}

            html += '<div class="actions" style="margin-top:12px;">' +
                '<button class="btn-cancel" id="edit-cancel">Annuler</button>' +
                '<button class="btn-delete" id="del-confirm" style="flex:1;">Supprimer</button>' +
                '</div>';

            popup.innerHTML = html;
            document.body.appendChild(popup);

            // "Select all" toggles all checkboxes
            document.getElementById('del-all').onchange = function() {{
                var checked = this.checked;
                popup.querySelectorAll('.del-day-cb').forEach(function(cb) {{ cb.checked = checked; }});
            }};
            // If all individual checkboxes are checked, check "select all" too
            popup.querySelectorAll('.del-day-cb').forEach(function(cb) {{
                cb.onchange = function() {{
                    var allCbs = popup.querySelectorAll('.del-day-cb');
                    var allChecked = true;
                    allCbs.forEach(function(c) {{ if (!c.checked) allChecked = false; }});
                    document.getElementById('del-all').checked = allChecked;
                }};
            }});

            document.getElementById('edit-cancel').onclick = closeEditPopup;
            document.getElementById('del-confirm').onclick = function() {{
                var selectedDays = [];
                popup.querySelectorAll('.del-day-cb:checked').forEach(function(cb) {{
                    selectedDays.push(parseInt(cb.value));
                }});
                if (selectedDays.length === 0) return;

                // If all days selected, remove employee entirely
                if (selectedDays.length === daysWithEvents.length) {{
                    delete DATA[empName];
                }} else {{
                    // Remove only events from selected days
                    emp.events = emp.events.filter(function(ev) {{
                        return selectedDays.indexOf(ev.day) === -1;
                    }});
                }}
                renderTimeline();
                updateHoursBadges();
                closeEditPopup();
                pushDataAfterEdit();
            }};
        }}

        var _editsDirty = false;

        function pushDataAfterEdit() {{
            _editsDirty = true;
            updateSaveButton();
            updateUnsavedBanner();
        }}

        function updateSaveButton() {{
            var btn = document.getElementById('save-edits-btn');
            if (!btn) return;
            if (_editsDirty) {{
                btn.style.display = '';
                btn.disabled = false;
                btn.textContent = 'Enregistrer';
                btn.className = 'save-edits-btn dirty';
            }} else {{
                btn.className = 'save-edits-btn';
                btn.style.display = 'none';
            }}
        }}

        function publishAllEdits() {{
            var btn = document.getElementById('save-edits-btn');
            if (btn) {{
                btn.disabled = true;
                btn.textContent = 'Sauvegarde...';
                btn.className = 'save-edits-btn saving';
            }}
            var statusEl = document.getElementById('edit-status');
            pushDataToGitHub(function(ok) {{
                if (ok) {{
                    _editsDirty = false;
                    updateUnsavedBanner();
                    // Mettre à jour la note de source
                    var metaEl = document.querySelector('.meta-note');
                    if (metaEl) {{
                        var raw = weekData._meta.updated_at;
                        var fmt = raw;
                        var m2 = raw.match(/^(\d{{4}})-(\d{{2}})-(\d{{2}})\s+(\d{{2}}:\d{{2}})$/);
                        if (m2) fmt = m2[3] + '/' + m2[2] + '/' + m2[1] + ' \u00e0 ' + m2[4];
                        metaEl.innerHTML = 'S{week_num} \u00b7 MAJ ' + fmt + ' \u00b7 Modif admin';
                    }}
                    if (btn) {{
                        btn.textContent = 'Sauvegard\u00e9 \u2714';
                        btn.className = 'save-edits-btn saved';
                        setTimeout(function() {{ updateSaveButton(); }}, 2000);
                    }}
                }} else {{
                    if (btn) {{
                        btn.disabled = false;
                        btn.textContent = 'Erreur \u2014 R\u00e9essayer';
                        btn.className = 'save-edits-btn error';
                    }}
                }}
            }});
        }}

        function closeEditPopup() {{
            var el = document.getElementById('edit-overlay');
            if (el) el.remove();
            el = document.getElementById('edit-popup');
            if (el) el.remove();
        }}

        function applyTimeEdit(empName, ev, newStart, newEnd) {{
            var dateStr = ev.start.substring(0, 11);
            ev.start = dateStr + newStart;
            ev.end = dateStr + newEnd;
            renderTimeline();
            updateHoursBadges();
            pushDataAfterEdit();
        }}

        function pushDataToGitHub(cb) {{
            var token = ensureToken();
            if (!token) {{ cb(false); return; }}

            // Build the updated data JSON for this week
            var weekData = {{}};
            Object.keys(DATA).forEach(function(name) {{
                if (name === '_codeNames') return;
                if (name === '_meta') return;
                weekData[name] = DATA[name];
            }});
            // Ajouter _meta avec source et timestamp
            var now = new Date();
            var pad = function(n) {{ return n < 10 ? '0' + n : '' + n; }};
            weekData._meta = {{
                source: 'Modif admin',
                updated_at: now.getFullYear() + '-' + pad(now.getMonth()+1) + '-' + pad(now.getDate()) + ' ' + pad(now.getHours()) + ':' + pad(now.getMinutes())
            }};
            var content = btoa(unescape(encodeURIComponent(JSON.stringify(weekData, null, 2) + '\\n')));
            var dataPath = 'data/S{week_num}-events.json';
            var apiUrl = 'https://api.github.com/repos/' + REPO + '/contents/' + dataPath;

            // Get current SHA
            fetch(apiUrl, {{
                headers: {{ 'Authorization': 'Bearer ' + token, 'Accept': 'application/vnd.github.v3+json' }}
            }})
            .then(function(r) {{ return r.ok ? r.json() : {{ sha: null }}; }})
            .then(function(file) {{
                var body = {{
                    message: 'MAJ cr\u00e9neaux S{week_num} depuis la page',
                    content: content,
                    branch: 'main'
                }};
                if (file.sha) body.sha = file.sha;

                return fetch(apiUrl, {{
                    method: 'PUT',
                    headers: {{
                        'Authorization': 'Bearer ' + token,
                        'Accept': 'application/vnd.github.v3+json',
                        'Content-Type': 'application/json'
                    }},
                    body: JSON.stringify(body)
                }});
            }})
            .then(function(r) {{ cb(r.ok); }})
            .catch(function() {{ cb(false); }});
        }}

        // Init admin toolbar if already unlocked
        initAdminToolbar();

        // ── Note discrète : dernière MAJ + source ──
        var _meta = {meta_json};
        (function() {{
            if (!_meta.updated_at) return;
            var src = _meta.source || 'Excel';
            // Formater la date en français (DD/MM/YYYY à HH:MM)
            var raw = _meta.updated_at;
            var formatted = raw;
            var m = raw.match(/^(\d{{4}})-(\d{{2}})-(\d{{2}})\s+(\d{{2}}:\d{{2}})$/);
            if (m) formatted = m[3] + '/' + m[2] + '/' + m[1] + ' à ' + m[4];
            var note = document.createElement('div');
            note.className = 'meta-note';
            note.innerHTML = 'S{week_num} · MAJ ' + formatted + ' · ' + src;
            document.querySelector('.container').appendChild(note);
        }})();

        // Admin button in header (combines auth + edit toggle)
        var adminBtnEl = document.getElementById('admin-btn');
        if (isAdminUnlocked()) {{
            adminBtnEl.classList.add('unlocked');
        }}
        adminBtnEl.onclick = function() {{
            if (!isAdminUnlocked()) {{
                if (verifyStaff()) {{
                    unlockAdmin();
                    adminBtnEl.classList.add('unlocked');
                    initAdminToolbar();
                    renderNotes();
                    var toggleBtn = adminToolbarEl.querySelector('.edit-toggle');
                    if (toggleBtn) toggleBtn.click();
                }}
            }} else {{
                // Already unlocked — toggle edit mode
                var toggleBtn = adminToolbarEl ? adminToolbarEl.querySelector('.edit-toggle') : null;
                if (toggleBtn) toggleBtn.click();
            }}
        }};

    }})();
    </script>
</body>
</html>"""


# ── Attendance pages (PSG Academy) ─────────────────────────────────────────


def generate_attendance_pages():
    """Génère les pages statiques de présences PSG Academy."""
    try:
        from attendance import find_attendance_file, parse_attendance
    except ImportError:
        print("\n⚠ Module attendance.py introuvable — pages présences non générées")
        return

    filepath = find_attendance_file()
    if not filepath:
        print("\n⚠ Aucun fichier Excel de présences trouvé — pages présences non générées")
        return

    data = parse_attendance(filepath)
    if not data["creneaux"]:
        print("\n⚠ Aucun créneau trouvé dans le fichier de présences")
        return

    os.makedirs("data", exist_ok=True)
    print(f"\n── Génération pages présences ({len(data['creneaux'])} créneaux) ──")

    creneaux_index = []

    for creneau in data["creneaux"]:
        slug = creneau["slug"]
        title = creneau["title"]
        sessions = creneau["sessions"]
        groups = creneau["groups"]
        total_kids = sum(len(g["kids"]) for g in groups)

        # ── Write JSON data ──
        json_path = f"data/presences-{slug}.json"
        json_data = {
            "title": title,
            "slug": slug,
            "sheet_name": creneau["sheet_name"],
            "sessions": sessions,
            "groups": [
                {
                    "name": g["name"],
                    "kids": [
                        {
                            "row": k["row"],
                            "num": k["num"],
                            "name": k["name"],
                            "category": k["category"],
                            "attendance": k["attendance"],
                        }
                        for k in g["kids"]
                    ],
                }
                for g in groups
            ],
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

        creneaux_index.append({"title": title, "slug": slug, "total_kids": total_kids,
                                "sessions_count": len([s for s in sessions if not s["is_vacation"]])})

        # ── Generate HTML page ──
        _write_attendance_html(slug, title, sessions, groups, json_data)
        print(f"  Écrit : presences-{slug}.html + {json_path}")

    # ── Generate index page ──
    _write_attendance_index(creneaux_index)
    print(f"  Écrit : presences.html (index, {len(creneaux_index)} créneaux)")


def _write_attendance_html(slug, title, sessions, groups, json_data):
    """Génère une page HTML statique de présences pour un créneau."""
    # Build session headers
    session_headers = ""
    for s in sessions:
        cls = "vacation" if s["is_vacation"] else ""
        label = s["label"]
        date_str = s.get("date", "") or ""
        session_headers += (
            f'<th class="session-col {cls}" data-label="{label}" '
            f'title="{date_str}">{label}</th>\n'
        )

    # Build group tables
    groups_html = ""
    for g_idx, group in enumerate(groups):
        groups_html += f'<div class="group-section" data-group="{g_idx}">\n'
        groups_html += f'<h3 class="group-title">{group["name"]}</h3>\n'
        groups_html += '<table class="attendance-table"><thead><tr>\n'
        groups_html += '<th class="num-col">#</th><th class="name-col">Nom</th>'
        groups_html += '<th class="cat-col">Cat.</th>\n'
        groups_html += session_headers
        groups_html += '<th class="total-col">Total</th></tr></thead><tbody>\n'

        for kid in group["kids"]:
            groups_html += (
                f'<tr data-row="{kid["row"]}" data-name="{kid["name"]}">\n'
                f'<td class="num-col">{kid["num"]}</td>'
                f'<td class="name-col">{kid["name"]}</td>'
                f'<td class="cat-col">{kid["category"]}</td>\n'
            )
            for s in sessions:
                val = kid["attendance"].get(s["label"])
                cls = "vacation" if s["is_vacation"] else ""
                if val == 1:
                    cls += " present"
                elif val == 0:
                    cls += " absent"
                display = "1" if val == 1 else ("0" if val == 0 else "")
                groups_html += (
                    f'<td class="session-cell {cls}" data-label="{s["label"]}" '
                    f'data-row="{kid["row"]}" data-col="{s["col"]}">{display}</td>\n'
                )
            # Total present
            total_p = sum(1 for s in sessions if kid["attendance"].get(s["label"]) == 1)
            groups_html += f'<td class="total-col total-val">{total_p}</td></tr>\n'

        # Total row
        groups_html += '</tbody><tfoot><tr class="total-row">\n'
        groups_html += '<td></td><td class="name-col"><strong>TOTAL</strong></td><td></td>\n'
        for s in sessions:
            total = sum(
                1 for k in group["kids"]
                if k["attendance"].get(s["label"]) == 1
            )
            cls = "vacation" if s["is_vacation"] else ""
            groups_html += f'<td class="session-cell {cls} total-cell">{total if total else ""}</td>\n'
        grand = sum(
            sum(1 for s in sessions if k["attendance"].get(s["label"]) == 1)
            for k in group["kids"]
        )
        groups_html += f'<td class="total-col"><strong>{grand}</strong></td></tr>\n'
        groups_html += '</tfoot></table></div>\n'

    json_embedded = json.dumps(json_data, ensure_ascii=False)

    html = f'''<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PSG Academy - {title}</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Montserrat', sans-serif;
            background: #0C1C3E;
            min-height: 100vh;
            padding: 10px;
            color: #fff;
            position: relative;
        }}
        body::before {{
            content: '';
            position: fixed; inset: 0; z-index: 0; pointer-events: none;
            background: url('bg-psg.jpg') center center / cover no-repeat;
            opacity: 0.15;
        }}
        .container {{
            position: relative; z-index: 1; max-width: 900px; margin: 0 auto;
            background: rgba(12,28,62,0.95); border-radius: 8px;
            padding: 12px; margin-top: 6px; margin-bottom: 6px;
            border-top: 4px solid #E30613; overflow: visible;
        }}

        /* ── Header ── */
        .header {{
            text-align: center; margin-bottom: 12px;
            padding: 14px 10px 8px; position: relative;
        }}
        .back-btn {{
            position: absolute; top: 8px; left: 8px;
            background: none; border: 1px solid rgba(255,255,255,0.15);
            color: #aaa; font-size: 11px; cursor: pointer; padding: 6px 12px;
            border-radius: 6px; text-decoration: none; font-family: inherit;
            font-weight: 600; transition: all 0.2s;
        }}
        .back-btn:hover {{ color: #E30613; border-color: #E30613; }}
        h1 {{
            color: #fff; font-size: 20px; font-weight: 900; margin-bottom: 2px;
            text-transform: uppercase; letter-spacing: 2px;
        }}
        .subtitle {{
            color: #E30613; font-size: 11px; font-weight: 700;
            text-transform: uppercase; letter-spacing: 3px;
        }}

        /* ── Créneau tabs ── */
        .creneau-tabs {{
            display: flex; justify-content: center; gap: 6px;
            margin-bottom: 15px; flex-wrap: wrap;
        }}
        .creneau-tab {{
            padding: 8px 14px; background: rgba(255,255,255,0.04);
            border: 1px solid rgba(255,255,255,0.08); border-radius: 6px;
            color: #666; text-decoration: none; font-weight: 600; font-size: 12px;
            transition: all 0.2s; text-transform: uppercase; letter-spacing: 0.5px;
        }}
        .creneau-tab:hover {{ background: rgba(227,6,19,0.1); border-color: rgba(227,6,19,0.3); color: #E30613; }}
        .creneau-tab.active {{ background: #E30613; border-color: #E30613; color: white;
                                box-shadow: 0 0 15px rgba(227,6,19,0.4); }}

        /* ── Session selector ── */
        .session-selector {{
            display: flex; align-items: center; gap: 10px;
            justify-content: center; margin-bottom: 15px; flex-wrap: wrap;
        }}
        .session-selector label {{ font-size: 12px; font-weight: 700; color: #aaa; text-transform: uppercase; }}
        .session-selector select {{
            padding: 8px 12px; background: rgba(0,0,0,0.3);
            border: 1px solid rgba(255,255,255,0.15); border-radius: 6px;
            color: #fff; font-size: 13px; font-family: inherit; font-weight: 600;
        }}
        .session-selector select:focus {{ border-color: #E30613; outline: none; }}

        /* ── Tables ── */
        .group-section {{ margin-bottom: 20px; }}
        .group-title {{
            font-size: 13px; font-weight: 700; color: #E30613;
            text-transform: uppercase; letter-spacing: 1px;
            margin-bottom: 8px; padding-left: 4px;
        }}
        .attendance-table {{
            width: 100%; border-collapse: collapse; font-size: 11px;
        }}
        .attendance-table th {{
            background: rgba(255,255,255,0.06); padding: 6px 4px;
            text-align: center; font-weight: 700; color: #aaa;
            border-bottom: 2px solid rgba(255,255,255,0.1);
            white-space: nowrap; font-size: 10px;
        }}
        .attendance-table td {{
            padding: 5px 4px; text-align: center;
            border-bottom: 1px solid rgba(255,255,255,0.05);
        }}
        .name-col {{ text-align: left !important; white-space: nowrap; min-width: 140px; }}
        .num-col {{ width: 28px; color: #555; }}
        .cat-col {{ width: 40px; color: #888; font-size: 10px; }}
        .total-col {{ width: 40px; font-weight: 700; color: #E30613; }}
        .session-col {{ min-width: 28px; }}
        .session-cell {{ cursor: default; transition: all 0.15s; min-width: 28px; }}
        .session-cell.editable {{ cursor: pointer; }}
        .session-cell.editable:hover {{ background: rgba(255,255,255,0.1); }}
        .session-cell.present {{ background: rgba(0,200,120,0.25); color: #4fc6a0; font-weight: 700; }}
        .session-cell.absent {{ background: rgba(227,6,19,0.2); color: #ff6b6b; font-weight: 700; }}
        .session-cell.vacation {{ background: rgba(0,180,220,0.15); color: #0bb4dc; }}
        .total-row td {{ font-weight: 700; border-top: 2px solid rgba(255,255,255,0.15);
                          background: rgba(255,255,255,0.03); }}

        /* ── Edit mode ── */
        .edit-bar {{
            display: none; justify-content: center; gap: 10px;
            margin-bottom: 12px; padding: 10px;
            background: rgba(227,6,19,0.1); border: 1px solid rgba(227,6,19,0.3);
            border-radius: 8px;
        }}
        .edit-bar.active {{ display: flex; flex-wrap: wrap; align-items: center; }}
        .edit-bar .badge {{ font-size: 11px; font-weight: 700; color: #E30613;
                             text-transform: uppercase; letter-spacing: 1px; }}
        .save-btn {{
            padding: 8px 20px; background: #E30613; border: none; border-radius: 6px;
            color: white; font-weight: 700; font-size: 12px; cursor: pointer;
            font-family: inherit; text-transform: uppercase; transition: all 0.2s;
        }}
        .save-btn:hover {{ background: #ff1a2a; box-shadow: 0 0 15px rgba(227,6,19,0.4); }}
        .save-btn:disabled {{ opacity: 0.5; cursor: not-allowed; }}
        .cancel-btn {{
            padding: 8px 16px; background: none; border: 1px solid rgba(255,255,255,0.2);
            border-radius: 6px; color: #aaa; font-weight: 600; font-size: 12px;
            cursor: pointer; font-family: inherit; transition: all 0.2s;
        }}
        .cancel-btn:hover {{ border-color: #fff; color: #fff; }}

        /* ── Admin button ── */
        .admin-btn {{
            position: absolute; top: 8px; right: 8px;
            background: none; border: 1px solid rgba(255,255,255,0.15);
            color: #555; font-size: 16px; cursor: pointer; padding: 4px 10px;
            border-radius: 6px; transition: all 0.2s;
        }}
        .admin-btn:hover {{ color: #E30613; border-color: #E30613; }}
        .admin-btn.unlocked {{ color: #E30613; border-color: #E30613; }}

        /* ── Responsive ── */
        .table-wrapper {{ overflow-x: auto; -webkit-overflow-scrolling: touch; }}
        @media (max-width: 600px) {{
            body {{ padding: 4px; }}
            .container {{ padding: 6px; }}
            .attendance-table {{ font-size: 10px; }}
            .name-col {{ min-width: 100px; }}
        }}

        /* ── Status messages ── */
        .status-msg {{
            position: fixed; bottom: 20px; left: 50%; transform: translateX(-50%);
            padding: 12px 24px; border-radius: 8px; font-size: 13px; font-weight: 600;
            z-index: 100; opacity: 0; transition: opacity 0.3s;
            font-family: 'Montserrat', sans-serif;
        }}
        .status-msg.show {{ opacity: 1; }}
        .status-msg.success {{ background: rgba(0,200,120,0.9); color: #fff; }}
        .status-msg.error {{ background: rgba(227,6,19,0.9); color: #fff; }}
    </style>
</head>
<body>
    <div class="container">
        <a href="presences.html" class="back-btn">&larr; Créneaux</a>
        <button class="admin-btn" id="admin-btn" title="Mode édition">&#9881;</button>

        <div class="header">
            <h1>PSG Academy</h1>
            <p class="subtitle">{title}</p>
        </div>

        <div class="session-selector">
            <label>Séance :</label>
            <select id="session-select"></select>
        </div>

        <div class="edit-bar" id="edit-bar">
            <span class="badge">Mode édition</span>
            <button class="save-btn" id="save-btn" disabled>Enregistrer</button>
            <button class="cancel-btn" id="cancel-btn">Annuler</button>
        </div>

        <div class="table-wrapper">
{groups_html}
        </div>
    </div>

    <div class="status-msg" id="status-msg"></div>

    <script>
    (function() {{
        var DATA = {json_embedded};
        var REPO = 'OhLaPey/planning-urbansoccer';
        var JSON_PATH = 'data/presences-{slug}.json';
        var TOKEN_KEY = 'planning-admin-token';
        var STAFF_CODE = '1937';
        var STAFF_KEY = 'planning-staff-ok';
        var _p = ['Z2l0aHViX3BhdF8xMUJWTEZMVl','EwNGFQeEFvQWZzYktvX2lZOHZF','cVhqaUx1ZzNmOVQ5cUhUcUJKan','NkMWhKR2tGYXl0c28xMDJmYXRV','SFhYS1pWWks4MXZGUkpE'];

        var editMode = false;
        var dirty = false;
        var pendingChanges = {{}};  // {{"row-col": value}}

        // ── Auth ──
        function isStaffVerified() {{ return sessionStorage.getItem(STAFF_KEY) === '1'; }}
        function verifyStaff() {{
            if (isStaffVerified()) return true;
            var code = prompt('Code staff requis :');
            if (code && code.trim() === STAFF_CODE) {{ sessionStorage.setItem(STAFF_KEY, '1'); return true; }}
            alert('Code staff incorrect.'); return false;
        }}
        function getToken() {{
            if (!isStaffVerified()) return '';
            return localStorage.getItem(TOKEN_KEY) || atob(_p.join(''));
        }}

        // ── Session selector ──
        var sessionSelect = document.getElementById('session-select');
        var sessions = DATA.sessions.filter(function(s) {{ return !s.is_vacation; }});
        var vacations = DATA.sessions.filter(function(s) {{ return s.is_vacation; }});

        // Determine current session
        var today = new Date();
        var todayStr = today.getFullYear() + '-' +
            String(today.getMonth()+1).padStart(2,'0') + '-' +
            String(today.getDate()).padStart(2,'0');
        var currentSession = sessions.length > 0 ? sessions[sessions.length - 1].label : null;
        for (var i = 0; i < sessions.length; i++) {{
            if (sessions[i].date) {{
                // Parse dd/mm/yy
                var parts = sessions[i].date.split('/');
                if (parts.length === 3) {{
                    var yr = parseInt(parts[2]);
                    if (yr < 100) yr += 2000;
                    var sDate = yr + '-' + parts[1].padStart(2,'0') + '-' + parts[0].padStart(2,'0');
                    if (sDate >= todayStr) {{ currentSession = sessions[i].label; break; }}
                }}
            }}
        }}

        // Populate select
        sessions.forEach(function(s) {{
            var opt = document.createElement('option');
            opt.value = s.label;
            opt.textContent = s.label + (s.date ? ' (' + s.date + ')' : '');
            if (s.label === currentSession) opt.selected = true;
            sessionSelect.appendChild(opt);
        }});

        function getSelectedSession() {{ return sessionSelect.value; }}

        function highlightSession() {{
            var sel = getSelectedSession();
            // Hide all session columns except selected + show name/num/cat/total
            document.querySelectorAll('.session-col, .session-cell').forEach(function(el) {{
                var label = el.getAttribute('data-label');
                el.style.display = (label === sel) ? '' : 'none';
            }});
            updateTotals();
        }}

        sessionSelect.addEventListener('change', function() {{
            if (dirty) {{
                if (!confirm('Modifications non enregistrées. Changer de séance ?')) {{
                    sessionSelect.value = currentSession;
                    return;
                }}
                cancelEdit();
            }}
            highlightSession();
        }});

        // ── Totals ──
        function updateTotals() {{
            document.querySelectorAll('.group-section').forEach(function(section) {{
                var rows = section.querySelectorAll('tbody tr');
                rows.forEach(function(row) {{
                    // Count all visible present cells for this kid
                    var total = 0;
                    DATA.sessions.forEach(function(s) {{
                        var cell = row.querySelector('.session-cell[data-label="' + s.label + '"]');
                        if (cell && cell.textContent.trim() === '1') total++;
                    }});
                    var totalCell = row.querySelector('.total-val');
                    if (totalCell) totalCell.textContent = total;
                }});
                // Footer total for selected session
                var sel = getSelectedSession();
                var footCells = section.querySelectorAll('tfoot .session-cell');
                footCells.forEach(function(fc) {{
                    var label = fc.getAttribute('data-label');
                    if (label === sel) {{
                        var count = 0;
                        rows.forEach(function(row) {{
                            var cell = row.querySelector('.session-cell[data-label="' + label + '"]');
                            if (cell && cell.textContent.trim() === '1') count++;
                        }});
                        fc.textContent = count || '';
                    }}
                }});
            }});
        }}

        // ── Edit mode ──
        var adminBtn = document.getElementById('admin-btn');
        var editBar = document.getElementById('edit-bar');
        var saveBtn = document.getElementById('save-btn');
        var cancelBtn = document.getElementById('cancel-btn');

        adminBtn.addEventListener('click', function() {{
            if (editMode) {{
                cancelEdit();
                return;
            }}
            if (!verifyStaff()) return;
            editMode = true;
            adminBtn.classList.add('unlocked');
            editBar.classList.add('active');
            enableCellEditing();
        }});

        function enableCellEditing() {{
            var sel = getSelectedSession();
            document.querySelectorAll('.session-cell[data-label="' + sel + '"]').forEach(function(cell) {{
                if (cell.closest('tfoot')) return;  // skip total row
                cell.classList.add('editable');
                cell.addEventListener('click', toggleCell);
            }});
        }}

        function disableCellEditing() {{
            document.querySelectorAll('.session-cell.editable').forEach(function(cell) {{
                cell.classList.remove('editable');
                cell.removeEventListener('click', toggleCell);
            }});
        }}

        function toggleCell(e) {{
            if (!editMode) return;
            var cell = e.currentTarget;
            var current = cell.textContent.trim();
            var newVal;
            if (current === '1') {{
                newVal = 0;
                cell.textContent = '0';
                cell.classList.remove('present');
                cell.classList.add('absent');
            }} else {{
                newVal = 1;
                cell.textContent = '1';
                cell.classList.remove('absent');
                cell.classList.add('present');
            }}
            var key = cell.getAttribute('data-row') + '-' + cell.getAttribute('data-col');
            pendingChanges[key] = newVal;
            dirty = true;
            saveBtn.disabled = false;
            updateTotals();
        }}

        function cancelEdit() {{
            if (dirty && !confirm('Annuler les modifications ?')) return;
            // Revert changes
            Object.keys(pendingChanges).forEach(function(key) {{
                var parts = key.split('-');
                var row = parts[0], col = parts[1];
                var cell = document.querySelector('.session-cell[data-row="' + row + '"][data-col="' + col + '"]');
                if (cell) {{
                    // Find original value in DATA
                    var origVal = findOriginalValue(parseInt(row), getSelectedSession());
                    cell.textContent = origVal === 1 ? '1' : (origVal === 0 ? '0' : '');
                    cell.classList.remove('present', 'absent');
                    if (origVal === 1) cell.classList.add('present');
                    else if (origVal === 0) cell.classList.add('absent');
                }}
            }});
            pendingChanges = {{}};
            dirty = false;
            editMode = false;
            saveBtn.disabled = true;
            adminBtn.classList.remove('unlocked');
            editBar.classList.remove('active');
            disableCellEditing();
            updateTotals();
        }}

        function findOriginalValue(row, sessionLabel) {{
            for (var gi = 0; gi < DATA.groups.length; gi++) {{
                for (var ki = 0; ki < DATA.groups[gi].kids.length; ki++) {{
                    if (DATA.groups[gi].kids[ki].row === row) {{
                        return DATA.groups[gi].kids[ki].attendance[sessionLabel];
                    }}
                }}
            }}
            return null;
        }}

        // ── Save via GitHub API ──
        saveBtn.addEventListener('click', function() {{
            if (!dirty) return;
            var token = getToken();
            if (!token) {{ showStatus('Token manquant', 'error'); return; }}

            saveBtn.disabled = true;
            saveBtn.textContent = 'Enregistrement...';

            // Update DATA in memory
            var sel = getSelectedSession();
            Object.keys(pendingChanges).forEach(function(key) {{
                var parts = key.split('-');
                var row = parseInt(parts[0]);
                var val = pendingChanges[key];
                for (var gi = 0; gi < DATA.groups.length; gi++) {{
                    for (var ki = 0; ki < DATA.groups[gi].kids.length; ki++) {{
                        if (DATA.groups[gi].kids[ki].row === row) {{
                            DATA.groups[gi].kids[ki].attendance[sel] = val;
                        }}
                    }}
                }}
            }});

            // Push to GitHub
            var url = 'https://api.github.com/repos/' + REPO + '/contents/' + JSON_PATH;
            fetch(url, {{
                headers: {{ 'Authorization': 'token ' + token, 'Accept': 'application/vnd.github.v3+json' }}
            }})
            .then(function(r) {{ return r.json(); }})
            .then(function(info) {{
                var sha = info.sha;
                var content = btoa(unescape(encodeURIComponent(JSON.stringify(DATA, null, 2))));
                return fetch(url, {{
                    method: 'PUT',
                    headers: {{
                        'Authorization': 'token ' + token,
                        'Accept': 'application/vnd.github.v3+json',
                        'Content-Type': 'application/json'
                    }},
                    body: JSON.stringify({{
                        message: 'Présences ' + DATA.title + ' ' + sel + ' — mise à jour',
                        content: content,
                        sha: sha
                    }})
                }});
            }})
            .then(function(r) {{
                if (!r.ok) throw new Error('HTTP ' + r.status);
                return r.json();
            }})
            .then(function() {{
                pendingChanges = {{}};
                dirty = false;
                saveBtn.textContent = 'Enregistrer';
                saveBtn.disabled = true;
                showStatus('Enregistré !', 'success');
            }})
            .catch(function(err) {{
                saveBtn.textContent = 'Enregistrer';
                saveBtn.disabled = false;
                showStatus('Erreur : ' + err.message, 'error');
            }});
        }});

        cancelBtn.addEventListener('click', cancelEdit);

        // ── Status messages ──
        function showStatus(msg, type) {{
            var el = document.getElementById('status-msg');
            el.textContent = msg;
            el.className = 'status-msg show ' + type;
            setTimeout(function() {{ el.classList.remove('show'); }}, 3000);
        }}

        // ── Init ──
        highlightSession();
        // If staff already verified, show admin as ready
        if (isStaffVerified()) {{
            adminBtn.style.color = '#E30613';
            adminBtn.style.borderColor = 'rgba(227,6,19,0.3)';
        }}
    }})();
    </script>
</body>
</html>'''

    with open(f"presences-{slug}.html", "w", encoding="utf-8") as f:
        f.write(html)


def _write_attendance_index(creneaux_index):
    """Génère la page index des créneaux de présences."""
    cards_html = ""
    tabs_html = ""
    for c in creneaux_index:
        cards_html += (
            f'<a href="presences-{c["slug"]}.html" class="creneau-card">\n'
            f'  <div class="creneau-title">{c["title"]}</div>\n'
            f'  <div class="creneau-info">{c["total_kids"]} enfants &middot; '
            f'{c["sessions_count"]} séances</div>\n'
            f'</a>\n'
        )
        tabs_html += (
            f'<a href="presences-{c["slug"]}.html" class="creneau-tab">{c["title"]}</a>\n'
        )

    html = f'''<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PSG Academy - Présences</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Montserrat', sans-serif;
            background: #0C1C3E;
            min-height: 100vh;
            padding: 15px;
            color: #fff;
            position: relative;
        }}
        body::before {{
            content: '';
            position: fixed; inset: 0; z-index: 0; pointer-events: none;
            background: url('bg-psg.jpg') center center / cover no-repeat;
            opacity: 0.15;
        }}
        .container {{
            position: relative; z-index: 1; max-width: 600px; margin: 0 auto;
            background: rgba(12,28,62,0.95); border-radius: 8px;
            padding: 20px; margin-top: 20px;
            border-top: 4px solid #E30613;
        }}
        .header {{
            text-align: center; margin-bottom: 24px; padding: 14px 10px 8px;
            position: relative;
        }}
        .back-btn {{
            position: absolute; top: 8px; left: 0;
            background: none; border: 1px solid rgba(255,255,255,0.15);
            color: #aaa; font-size: 11px; cursor: pointer; padding: 6px 12px;
            border-radius: 6px; text-decoration: none; font-family: inherit;
            font-weight: 600; transition: all 0.2s;
        }}
        .back-btn:hover {{ color: #E30613; border-color: #E30613; }}
        h1 {{
            color: #fff; font-size: 24px; font-weight: 900; margin-bottom: 4px;
            text-transform: uppercase; letter-spacing: 2px;
        }}
        .subtitle {{
            color: #E30613; font-size: 11px; font-weight: 700;
            text-transform: uppercase; letter-spacing: 3px;
        }}
        .creneau-card {{
            display: block; padding: 18px 20px; margin-bottom: 10px;
            background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08);
            border-radius: 8px; text-decoration: none; color: #fff;
            transition: all 0.2s;
        }}
        .creneau-card:hover {{
            background: rgba(227,6,19,0.1); border-color: rgba(227,6,19,0.3);
            transform: translateY(-2px);
        }}
        .creneau-title {{
            font-size: 16px; font-weight: 800; text-transform: uppercase;
            letter-spacing: 1px; margin-bottom: 4px;
        }}
        .creneau-info {{
            font-size: 12px; color: #888; font-weight: 500;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <a href="index.html" class="back-btn">&larr; Planning</a>
            <h1>PSG Academy</h1>
            <p class="subtitle">Suivi des présences</p>
        </div>

{cards_html}
    </div>
</body>
</html>'''

    with open("presences.html", "w", encoding="utf-8") as f:
        f.write(html)


# ── Main ───────────────────────────────────────────────────────────────────


def main():
    excel_files = discover_excel_files()
    if not excel_files:
        print("Aucun fichier 'Plannings YYYY SXX.xlsx' trouv\u00e9.")
        return

    print(f"Fichiers Excel trouv\u00e9s : {len(excel_files)}")
    for ef in excel_files:
        print(f"  - {ef['filename']} (S{ef['week']}, {ef['year']})")

    # ── Collecter tous les événements par employé, toutes semaines ──
    all_employee_events = {}   # {name: [events]}
    week_data = {}             # {week_num: {employees, year}}
    all_weeks = set()

    for ef in excel_files:
        year, week_num = ef["year"], ef["week"]
        dates = week_dates(year, week_num)

        wb = openpyxl.load_workbook(ef["filename"])
        ws = wb["Planning"] if "Planning" in wb.sheetnames else wb.active

        employees = parse_employees(ws, dates, week_num)
        all_weeks.add(week_num)
        week_data[week_num] = {"employees": employees, "year": year}

        active_count = 0
        print(f"\nSemaine {week_num} ({year}) :")
        for name, evts in employees.items():
            if name not in all_employee_events:
                all_employee_events[name] = []
            all_employee_events[name].extend(evts)
            if evts:
                active_count += 1
                print(f"  {name} ({len(evts)} \u00e9v\u00e9nements)")
                for e in evts:
                    end_str = e["end"].strftime("%H:%M")
                    if e["end"].date() > e["start"].date():
                        end_str += " (+1j)"
                    print(f"    {e['start'].strftime('%a %d/%m %H:%M')} - "
                          f"{end_str} : {e['label']}")
        print(f"  \u2192 {active_count} employ\u00e9s actifs")

    # ── Charger les notes par semaine ──
    all_week_notes = {}
    for wn in all_weeks:
        all_week_notes[wn] = load_week_notes(wn)

    # ── Injecter les créneaux virtuels pour les remplaçants sans événement ce jour ──
    for wn in all_weeks:
        notes = all_week_notes.get(wn, {})
        for r in notes.get("replacements", []):
            in_name = r.get("in", "")
            out_name = r.get("out", "")
            if not in_name:
                continue
            repl_date_str = r.get("date", "")
            r_parts_s = r.get("start", "0:00").split(":")
            r_parts_e = r.get("end", "0:00").split(":")
            r_sh = int(r_parts_s[0]) + int(r_parts_s[1] if len(r_parts_s) > 1 else 0) / 60
            r_eh = int(r_parts_e[0]) + int(r_parts_e[1] if len(r_parts_e) > 1 else 0) / 60
            # Check if replacer already has events on this date
            in_evts = all_employee_events.get(in_name, [])
            has_on_date = any(e["start"].strftime("%Y-%m-%d") == repl_date_str for e in in_evts)
            if has_on_date:
                continue
            # Find code/label from replaced person's events
            ref_code = "VDC"
            ref_label = "Vie de centre"
            out_evts = all_employee_events.get(out_name, [])
            for oev in out_evts:
                if oev["start"].strftime("%Y-%m-%d") != repl_date_str:
                    continue
                o_sh = oev["start"].hour + oev["start"].minute / 60
                o_eh = oev["end"].hour + oev["end"].minute / 60
                if o_eh <= o_sh:
                    o_eh = 24
                if o_sh < r_eh and o_eh > r_sh:
                    ref_code = oev.get("code", ref_code)
                    ref_label = oev.get("label", ref_label)
                    break
            from datetime import datetime as _dt2
            synth_start = _dt2.strptime(f"{repl_date_str} {int(r_parts_s[0]):02d}:{int(r_parts_s[1] if len(r_parts_s)>1 else 0):02d}", "%Y-%m-%d %H:%M")
            synth_end = _dt2.strptime(f"{repl_date_str} {int(r_parts_e[0]):02d}:{int(r_parts_e[1] if len(r_parts_e)>1 else 0):02d}", "%Y-%m-%d %H:%M")
            synth_evt = {
                "code": ref_code,
                "label": ref_label,
                "start": synth_start,
                "end": synth_end,
                "week": wn,
            }
            if in_name not in all_employee_events:
                all_employee_events[in_name] = []
            all_employee_events[in_name].append(synth_evt)

    # ── Générer les fichiers ICS (cumulatifs, toutes semaines) ──
    os.makedirs("ics", exist_ok=True)
    ics_count = 0
    for name, events in all_employee_events.items():
        if events:
            events.sort(key=lambda e: e["start"])
            ics_content = generate_ics(name, events, week_notes=all_week_notes)
            filename = f"ics/{slug(name)}.ics"
            with open(filename, "w", encoding="utf-8") as f:
                f.write(ics_content)
            ics_count += 1
    print(f"\n{ics_count} fichiers ICS g\u00e9n\u00e9r\u00e9s dans ics/")

    # ── Générer HTML + JSON par semaine ──
    os.makedirs("data", exist_ok=True)
    for week_num in sorted(all_weeks):
        wd = week_data[week_num]
        year = wd["year"]
        employees = wd["employees"]

        # JSON
        active_names = sorted([n for n, e in employees.items() if e])
        monday = datetime.fromisocalendar(year, week_num, 1)
        sunday = monday + timedelta(days=6)
        json_data = {
            "semaine": week_num,
            "annee": year,
            "date_debut": f"{monday.day} {FRENCH_MONTHS[monday.month]}",
            "date_fin": f"{sunday.day} {FRENCH_MONTHS[sunday.month]}",
            "employesActifs": active_names,
        }
        json_path = f"data/S{week_num}.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        print(f"\u00c9crit : {json_path}")

        # Events JSON — toujours régénérer depuis l'Excel (source de vérité)
        events_path = f"data/S{week_num}-events.json"
        ef_info = next((e for e in excel_files if e["week"] == week_num), None)
        excel_ver = ef_info["version"] if ef_info else 0
        events_data = json.loads(build_events_json(employees))
        source = f"Excel v{excel_ver}" if excel_ver > 0 else "Excel"
        events_data["_meta"] = {
            "source": source,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        with open(events_path, "w", encoding="utf-8") as f:
            json.dump(events_data, f, ensure_ascii=False, indent=2)
            f.write("\n")
        print(f"Écrit : {events_path}")


        # HTML
        html_content = generate_html(employees, week_num, year, all_weeks, excel_version=excel_ver)
        html_path = f"S{week_num}.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"\u00c9crit : {html_path}")

    # ── Mettre à jour index.html → redirection intelligente vers la bonne semaine ──
    latest_week = max(all_weeks)

    # Construire la table des semaines avec leurs dates de début
    weeks_info = []
    for wn in sorted(all_weeks):
        wd = week_data[wn]
        y = wd["year"]
        mon = datetime.fromisocalendar(y, wn, 1)
        sun = mon + timedelta(days=6)
        weeks_info.append({
            "week": wn,
            "start": mon.strftime("%Y-%m-%d"),
            "end": sun.strftime("%Y-%m-%d"),
        })
    weeks_json = json.dumps(weeks_info)

    with open("index.html", "w", encoding="utf-8") as f:
        f.write(
            '<!DOCTYPE html>\n'
            '<html lang="fr">\n'
            '<head>\n'
            '    <meta charset="UTF-8">\n'
            '    <meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
            '    <meta name="apple-mobile-web-app-capable" content="yes">\n'
            '    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">\n'
            '    <meta name="apple-mobile-web-app-title" content="Planning U7D">\n'
            '    <link rel="manifest" href="manifest.json">\n'
            '    <link rel="apple-touch-icon" href="data:image/svg+xml,<svg xmlns=\'http://www.w3.org/2000/svg\' viewBox=\'0 0 100 100\'><text y=\'.9em\' font-size=\'90\'>&#9917;</text></svg>">\n'
            '    <meta name="theme-color" content="#FF6600">\n'
            '    <title>Planning Urban 7D</title>\n'
            '    <script>\n'
            '    (function() {\n'
            f'        var weeks = {weeks_json};\n'
            '        var today = new Date();\n'
            '        var yyyy = today.getFullYear();\n'
            '        var mm = String(today.getMonth()+1).padStart(2,"0");\n'
            '        var dd = String(today.getDate()).padStart(2,"0");\n'
            '        var todayStr = yyyy + "-" + mm + "-" + dd;\n'
            '        var target = null;\n'
            '        for (var i = 0; i < weeks.length; i++) {\n'
            '            if (todayStr >= weeks[i].start && todayStr <= weeks[i].end) {\n'
            '                target = weeks[i].week; break;\n'
            '            }\n'
            '        }\n'
            '        if (!target) target = weeks[weeks.length - 1].week;\n'
            '        fetch("S" + target + ".html")\n'
            '            .then(function(r) { return r.text(); })\n'
            '            .then(function(html) {\n'
            '                document.open();\n'
            '                document.write(html);\n'
            '                document.close();\n'
            '            })\n'
            '            .catch(function() {\n'
            '                window.location.replace("S" + target + ".html");\n'
            '            });\n'
            '    })();\n'
            '    </script>\n'
            f'    <noscript><meta http-equiv="refresh" content="0;url=S{latest_week}.html"></noscript>\n'
            '</head>\n'
            '<body>\n'
            '</body>\n'
            '</html>'
        )
    print(f"\u00c9crit : index.html (semaines : {', '.join(f'S{w}' for w in sorted(all_weeks))})")

    # ── Générer les pages de présences PSG Academy ──
    generate_attendance_pages()

    print("\nTermin\u00e9 !")
    print("\n\u2500\u2500 Abonnement calendrier \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500")
    print("1. H\u00e9bergez ces fichiers (GitHub Pages, Netlify, etc.)")
    print("2. Les employ\u00e9s ouvrent la page et cliquent sur leur nom")
    print("3. Le calendrier se met \u00e0 jour automatiquement")
    print(f"\nPour ajouter une nouvelle semaine :")
    print(f"  1. Ajoutez le fichier Excel « Plannings {year} SXX.xlsx »")
    print(f"  2. Relancez : python generate.py")
    print(f"  3. Publiez les fichiers mis \u00e0 jour")


if __name__ == "__main__":
    main()
