"""
Génère un fichier Excel pour le budget lots du tournoi de Padel (84€).
Design aux couleurs Urban Padel (noir, vert fluo, orange, rouge).
- Tableau gauche : catalogue des lots disponibles avec prix
- Tableau droit : attribution des lots par place avec listes déroulantes
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XlImage
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Budget Lots"

# ============================================================
# CHARTE URBAN PADEL (extraite du site)
# ============================================================
BLACK = "1A1A1A"
DARK_GREY = "2D2D2D"
MID_GREY = "3A3A3A"
LIGHT_GREY = "4A4A4A"
GREEN_FLUO = "7BED5E"      # Vert fluo (boutons, accents)
GREEN_DARK = "5CB842"       # Vert foncé
ORANGE = "F5A623"           # Orange (formes géométriques)
RED = "E84057"              # Rouge/Rose (accent)
WHITE = "FFFFFF"
OFF_WHITE = "E8E8E8"

# Or / Argent / Bronze pour les places
GOLD = "FFD966"
SILVER = "C0C0C0"
BRONZE = "CD7F32"

currency_fmt = '#,##0.00 €'

# --- Fills ---
fill_black = PatternFill(start_color=BLACK, end_color=BLACK, fill_type="solid")
fill_dark = PatternFill(start_color=DARK_GREY, end_color=DARK_GREY, fill_type="solid")
fill_mid = PatternFill(start_color=MID_GREY, end_color=MID_GREY, fill_type="solid")
fill_light_grey = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
fill_green = PatternFill(start_color=GREEN_FLUO, end_color=GREEN_FLUO, fill_type="solid")
fill_green_dark = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
fill_orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
fill_red = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
fill_gold = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
fill_silver = PatternFill(start_color=SILVER, end_color=SILVER, fill_type="solid")
fill_bronze = PatternFill(start_color=BRONZE, end_color=BRONZE, fill_type="solid")

# --- Fonts ---
font_title = Font(bold=True, size=14, color=GREEN_FLUO)
font_subtitle = Font(bold=True, size=11, color=ORANGE)
font_header = Font(bold=True, size=10, color=BLACK)
font_white = Font(size=10, color=WHITE)
font_white_bold = Font(bold=True, size=10, color=WHITE)
font_green = Font(bold=True, size=10, color=GREEN_FLUO)
font_green_big = Font(bold=True, size=14, color=GREEN_FLUO)
font_orange = Font(bold=True, size=10, color=ORANGE)
font_red = Font(bold=True, size=11, color=RED)
font_black = Font(size=10, color=BLACK)
font_black_bold = Font(bold=True, size=10, color=BLACK)
font_off_white = Font(size=10, color=OFF_WHITE)

# --- Borders ---
border_grey = Border(
    left=Side(style='thin', color=LIGHT_GREY),
    right=Side(style='thin', color=LIGHT_GREY),
    top=Side(style='thin', color=LIGHT_GREY),
    bottom=Side(style='thin', color=LIGHT_GREY)
)
border_green = Border(
    left=Side(style='thin', color=GREEN_FLUO),
    right=Side(style='thin', color=GREEN_FLUO),
    top=Side(style='thin', color=GREEN_FLUO),
    bottom=Side(style='thin', color=GREEN_FLUO)
)
border_bottom_green = Border(
    bottom=Side(style='medium', color=GREEN_FLUO)
)

# --- Alignments ---
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

BUDGET = 84

def style_cell(cell, font=None, fill=None, alignment=None, number_format=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if number_format: cell.number_format = number_format
    if border: cell.border = border

def fill_bg(ws, row_start, row_end, col_start, col_end, fill=fill_black):
    """Remplir une zone en fond noir"""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).fill = fill

# ============================================================
# Fond noir sur toute la zone de travail
# ============================================================
fill_bg(ws, 1, 30, 1, 8, fill_black)

# ============================================================
# LOGO Urban Padel (en haut à gauche)
# ============================================================
logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo_urbanpadel_white.png")
if os.path.exists(logo_path):
    img = XlImage(logo_path)
    img.width = 140
    img.height = 84
    ws.add_image(img, 'A1')

# Espace pour le logo : lignes 1-4, cols A-B
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22
ws.row_dimensions[3].height = 22
ws.row_dimensions[4].height = 22

# Titre principal à droite du logo
ws.merge_cells('C1:D3')
c = ws['C1']
c.value = "TOURNOI\nPADEL"
style_cell(c, font=Font(bold=True, size=20, color=GREEN_FLUO), fill=fill_black, alignment=center)

# Budget en vert fluo
ws.merge_cells('C4:D4')
c = ws['C4']
c.value = f"BUDGET LOTS : {BUDGET}€"
style_cell(c, font=Font(bold=True, size=12, color=ORANGE), fill=fill_black, alignment=center)

# Titre droit
ws.merge_cells('F1:H3')
c = ws['F1']
c.value = "ATTRIBUTION\nDES LOTS"
style_cell(c, font=Font(bold=True, size=18, color=ORANGE), fill=fill_black, alignment=center)

ws.merge_cells('F4:H4')
c = ws['F4']
c.value = f"BUDGET : {BUDGET}€"
style_cell(c, font=Font(bold=True, size=12, color=GREEN_FLUO), fill=fill_black, alignment=center)

# ============================================================
# Ligne séparatrice verte (ligne 5)
# ============================================================
for c in range(1, 9):
    cell = ws.cell(row=5, column=c)
    cell.fill = PatternFill(start_color=GREEN_FLUO, end_color=GREEN_FLUO, fill_type="solid")
ws.row_dimensions[5].height = 4

# ============================================================
# TABLEAU GAUCHE : Catalogue des lots (colonnes A-D)
# ============================================================
row_header = 6
headers_left = ["#", "LOT DISPONIBLE", "  ", "PRIX"]
cols_left = [1, 2, 3, 4]
for i, h in enumerate(headers_left):
    cell = ws.cell(row=row_header, column=cols_left[i], value=h)
    style_cell(cell, font=font_header, fill=fill_green, alignment=center, border=border_grey)

ws.row_dimensions[row_header].height = 28

# Catalogue de lots
lots = [
    ("Tube balles Padel (x3)", 8),
    ("Surgrips (lot de 3)", 6),
    ("Bandeau sport", 5),
    ("Bidon isotherme", 12),
    ("Chaussettes sport", 10),
    ("Serviette microfibre", 9),
    ("Casquette / Visière", 15),
    ("Sac à chaussures", 18),
    ("Protège-raquette", 16),
    ("T-shirt technique", 22),
    ("Pochette de raquette", 25),
    ("Sac de padel compact", 30),
]

lot_start_row = row_header + 1
for idx, (nom, prix) in enumerate(lots):
    r = lot_start_row + idx
    ws.row_dimensions[r].height = 22
    fill = fill_dark if idx % 2 == 0 else fill_mid

    cell_num = ws.cell(row=r, column=1, value=idx + 1)
    style_cell(cell_num, font=font_green, fill=fill, alignment=center, border=border_grey)

    cell_nom = ws.cell(row=r, column=2, value=nom)
    style_cell(cell_nom, font=font_white, fill=fill, alignment=left_align, border=border_grey)

    # Col C = espace séparateur
    ws.cell(row=r, column=3).fill = fill
    ws.cell(row=r, column=3).border = border_grey

    cell_prix = ws.cell(row=r, column=4, value=prix)
    style_cell(cell_prix, font=font_orange, fill=fill, alignment=center, number_format=currency_fmt, border=border_grey)

lot_end_row = lot_start_row + len(lots) - 1

# ============================================================
# COLONNE SÉPARATRICE (E)
# ============================================================
for r in range(1, 30):
    ws.cell(row=r, column=5).fill = fill_black

# ============================================================
# TABLEAU DROIT : Attribution (colonnes F-H)
# ============================================================
headers_right = ["PLACE", "LOT ATTRIBUÉ", "COÛT"]
cols_right = [6, 7, 8]
for i, h in enumerate(headers_right):
    cell = ws.cell(row=row_header, column=cols_right[i], value=h)
    style_cell(cell, font=font_header, fill=fill_orange, alignment=center, border=border_grey)

# Validation : liste déroulante
dv = DataValidation(
    type="list",
    formula1=f"=$B${lot_start_row}:$B${lot_end_row}",
    allow_blank=True
)
dv.error = "Choisissez un lot du catalogue"
dv.errorTitle = "Lot invalide"
dv.prompt = "Sélectionnez un lot du catalogue"
dv.promptTitle = "Choix du lot"
ws.add_data_validation(dv)

# Places : 1er (4 lots), 2ème (3 lots), 3ème (2 lots)
places = [
    ("1ER", 4, fill_gold, Font(bold=True, size=10, color=BLACK)),
    ("2ÈME", 3, fill_silver, Font(bold=True, size=10, color=BLACK)),
    ("3ÈME", 2, fill_bronze, Font(bold=True, size=10, color=WHITE)),
]

current_row = row_header + 1
all_cost_rows = []

for place_name, nb_lots, fill_place, font_place in places:
    for i in range(nb_lots):
        r = current_row
        all_cost_rows.append(r)
        ws.row_dimensions[r].height = 22

        # Place label
        if i == 0:
            label = f"{place_name} - Lot {i+1}"
        else:
            label = f"Lot {i+1}"

        cell_place = ws.cell(row=r, column=6, value=label)
        style_cell(cell_place, font=font_place, fill=fill_place, alignment=center, border=border_grey)

        # Lot attribué (dropdown) - fond plus sombre pour indiquer saisie
        cell_lot = ws.cell(row=r, column=7)
        cell_lot.value = None
        style_cell(cell_lot, font=font_white, fill=fill_dark, alignment=left_align, border=border_green)
        dv.add(cell_lot)

        # Coût = RECHERCHEV
        cell_cost = ws.cell(row=r, column=8)
        cell_cost.value = f'=IF(G{r}="",0,VLOOKUP(G{r},$B${lot_start_row}:$D${lot_end_row},3,FALSE))'
        style_cell(cell_cost, font=font_orange, fill=fill_dark, alignment=center, number_format=currency_fmt, border=border_grey)

        current_row += 1

    # Sous-total par place
    r_sub = current_row
    ws.row_dimensions[r_sub].height = 20
    sub_start = r_sub - nb_lots
    sub_end = r_sub - 1

    cell_sub_label = ws.cell(row=r_sub, column=6, value=f"S/Total {place_name}")
    style_cell(cell_sub_label, font=font_white_bold, fill=fill_light_grey, alignment=center, border=border_grey)
    ws.cell(row=r_sub, column=7).fill = fill_light_grey
    ws.cell(row=r_sub, column=7).border = border_grey
    cell_sub_total = ws.cell(row=r_sub, column=8)
    cell_sub_total.value = f'=SUM(H{sub_start}:H{sub_end})'
    style_cell(cell_sub_total, font=font_green, fill=fill_light_grey, alignment=center, number_format=currency_fmt, border=border_grey)

    current_row += 1

# ============================================================
# Ligne séparatrice verte avant total
# ============================================================
for c in range(6, 9):
    cell = ws.cell(row=current_row, column=c)
    cell.fill = PatternFill(start_color=GREEN_FLUO, end_color=GREEN_FLUO, fill_type="solid")
ws.row_dimensions[current_row].height = 4
current_row += 1

# --- TOTAL ---
total_row = current_row
ws.row_dimensions[total_row].height = 30
first_cost = row_header + 1
last_cost = all_cost_rows[-1]

cell_total_label = ws.cell(row=total_row, column=6, value="TOTAL")
style_cell(cell_total_label, font=Font(bold=True, size=13, color=GREEN_FLUO), fill=fill_black, alignment=center, border=border_green)
ws.cell(row=total_row, column=7).fill = fill_black
ws.cell(row=total_row, column=7).border = border_green

# Total = somme des sous-totaux (plus fiable)
sub_total_rows = []
cursor = row_header + 1
for _, nb, _, _ in places:
    cursor += nb
    sub_total_rows.append(cursor)
    cursor += 1

sum_formula = "+".join([f"H{r}" for r in sub_total_rows])
cell_total = ws.cell(row=total_row, column=8)
cell_total.value = f'={sum_formula}'
style_cell(cell_total, font=Font(bold=True, size=13, color=GREEN_FLUO), fill=fill_black, alignment=center, number_format=currency_fmt, border=border_green)

# --- RESTE ---
reste_row = total_row + 1
ws.row_dimensions[reste_row].height = 30

cell_reste_label = ws.cell(row=reste_row, column=6, value="RESTE")
style_cell(cell_reste_label, font=Font(bold=True, size=13, color=RED), fill=fill_black, alignment=center, border=Border(
    left=Side(style='thin', color=RED), right=Side(style='thin', color=RED),
    top=Side(style='thin', color=RED), bottom=Side(style='thin', color=RED)))
ws.cell(row=reste_row, column=7).fill = fill_black
ws.cell(row=reste_row, column=7).border = Border(
    left=Side(style='thin', color=RED), right=Side(style='thin', color=RED),
    top=Side(style='thin', color=RED), bottom=Side(style='thin', color=RED))
cell_reste = ws.cell(row=reste_row, column=8)
cell_reste.value = f'={BUDGET}-H{total_row}'
style_cell(cell_reste, font=Font(bold=True, size=13, color=RED), fill=fill_black, alignment=center, number_format=currency_fmt,
           border=Border(left=Side(style='thin', color=RED), right=Side(style='thin', color=RED),
                         top=Side(style='thin', color=RED), bottom=Side(style='thin', color=RED)))

# ============================================================
# Fond noir restant
# ============================================================
max_row = max(lot_end_row, reste_row) + 3
fill_bg(ws, reste_row + 1, max_row, 1, 8, fill_black)

# ============================================================
# Largeurs de colonnes
# ============================================================
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 3
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 3   # séparateur
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 14

# Masquer le quadrillage
ws.sheet_view.showGridLines = False

# ============================================================
# Sauvegarde
# ============================================================
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "budget_lots_padel.xlsx")
wb.save(output)
print(f"Fichier généré : {output}")
print(f"  - Catalogue : {len(lots)} lots")
print(f"  - Attribution : 4 lots 1er + 3 lots 2ème + 2 lots 3ème")
print(f"  - Design : charte Urban Padel (noir/vert fluo/orange/rouge)")
